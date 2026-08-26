"""XLSX template filling for Markdown table exports."""

from __future__ import annotations

import html
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote

import yaml
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter, range_boundaries
from PIL import Image

from docwen_core.detection import detect_content_format
from docwen_core.formats import CATEGORY_IMAGE, get_category
from docwen_core.yaml_tools import extract_yaml
from docwen_plugin_markdown.common_utils import parse_md_tables, parse_raw_md_tables
from docwen_plugin_markdown.yaml_processor import ensure_title_fallback

_PLACEHOLDER_RE = re.compile(r"\{\{([^}]+)\}\}")
_IMAGE_PLACEHOLDER_RE = re.compile(r"\{\{IMAGE:([^\r\n]*?)\}\}")
_HTML_BREAK_RE = re.compile(r"<br\s*/?>", re.IGNORECASE)
_HTML_COMMENT_RE = re.compile(r"<!--.*?-->", re.DOTALL)
_HTML_TAG_RE = re.compile(r"<[^>]+>")
_ANGLE_URL_RE = re.compile(r"<(https?://[^>]+)>")
_ANGLE_EMAIL_RE = re.compile(r"<([^@\s<>]+@[^@\s<>]+\.[^@\s<>]+)>")
_ZERO_WIDTH_TRANSLATION = str.maketrans("", "", "\u200b\u200c\u200d\ufeff\u2060")
_COLUMN_PREFIX = "↓"
_ROW_PREFIX = "→"


def build_template_workbook(
    markdown: str,
    template_path: str | Path,
    *,
    source_stem: str = "",
    image_scope: str | None = None,
    list_separator: str = "、",
) -> tuple[Any, dict[str, int]]:
    """Fill an XLSX template using Markdown YAML and table data.

    This restores the core old-project contract without reintroducing the old
    global loader stack: plain ``{{field}}`` placeholders read YAML front matter,
    while whole-cell ``{{↓Header}}`` and ``{{→Header}}`` placeholders consume
    Markdown table columns vertically or horizontally.
    """
    yaml_text, body = extract_yaml(markdown)
    yaml_data = _parse_yaml_mapping(yaml_text)
    tables = parse_md_tables(body)
    raw_tables = parse_raw_md_tables(body, preserve_merge_marker_escapes=True)
    table_columns = _columns_from_tables(tables)

    with Path(template_path).open("rb") as stream:
        workbook = load_workbook(stream)
    if image_scope is not None:
        _promote_template_image_placeholders(workbook, image_scope)
    template_merged_ranges = _merged_ranges_by_sheet(workbook)
    merge_plans, merge_plan_warnings = _collect_markdown_table_merge_plans(workbook, raw_tables or tables)
    yaml_replacements = _yaml_replacements(
        workbook,
        yaml_data,
        source_stem=source_stem,
        list_separator=list_separator,
    )
    yaml_replaced = _replace_yaml_placeholders(workbook, yaml_replacements)
    column_filled, column_merged_skipped, column_protected_skipped = _fill_column_placeholders(workbook, table_columns)
    row_filled, row_merged_skipped, row_protected_skipped = _fill_row_placeholders(workbook, table_columns)
    markdown_merges_applied, markdown_merge_warnings = _apply_markdown_table_merge_plans(
        workbook, merge_plans, template_merged_ranges
    )
    images_inserted = process_image_placeholders(
        workbook,
        template_path,
        image_scope=image_scope,
    )
    remaining_cleaned = _clean_remaining_placeholders(workbook)

    return workbook, {
        "yaml_placeholders": yaml_replaced,
        "column_placeholders": column_filled,
        "row_placeholders": row_filled,
        "merged_cells_skipped": column_merged_skipped + row_merged_skipped,
        "protected_cells_skipped": column_protected_skipped + row_protected_skipped,
        "markdown_table_merges": markdown_merges_applied,
        "markdown_table_merge_warnings": merge_plan_warnings + markdown_merge_warnings,
        "image_placeholders": images_inserted,
        "remaining_placeholders_cleaned": remaining_cleaned,
        "table_count": len(tables),
    }


def _parse_yaml_mapping(yaml_text: str) -> dict[str, Any]:
    if not yaml_text.strip():
        return {}
    data = yaml.safe_load(yaml_text) or {}
    return data if isinstance(data, dict) else {}


def _columns_from_tables(tables: list[dict[str, Any]]) -> dict[str, list[tuple[Any, str | None]]]:
    columns: dict[str, list[tuple[Any, str | None]]] = {}
    for table in tables:
        headers = [str(header) for header in table.get("headers", [])]
        for row in table.get("rows", []):
            for index, header in enumerate(headers):
                if index >= len(row):
                    continue
                columns.setdefault(header, []).append(_coerce_cell_value(row[index]))
    return columns


def _merged_ranges_by_sheet(workbook: Any) -> dict[str, list[str]]:
    return {
        worksheet.title: [str(merged_range) for merged_range in worksheet.merged_cells.ranges]
        for worksheet in workbook.worksheets
    }


def _yaml_replacements(
    workbook: Any,
    yaml_data: dict[str, Any],
    *,
    source_stem: str,
    list_separator: str,
) -> dict[str, tuple[Any, str | None]]:
    field_names: set[str] = set()
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                for match in _PLACEHOLDER_RE.findall(cell.value):
                    if match.startswith((_COLUMN_PREFIX, _ROW_PREFIX, "IMAGE:", "IMAGE@")):
                        continue
                    field_names.add(match)

    ensure_title_fallback(yaml_data, placeholder_names=field_names, source_stem=source_stem)

    replacements: dict[str, tuple[Any, str | None]] = {}
    for field_name in field_names:
        if field_name not in yaml_data:
            continue
        replacements[f"{{{{{field_name}}}}}"] = _coerce_cell_value(
            yaml_data[field_name],
            list_separator=list_separator,
        )
    return replacements


def _replace_yaml_placeholders(workbook: Any, replacements: dict[str, tuple[Any, str | None]]) -> int:
    replaced = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value:
                    continue
                original = cell.value
                stripped = original.strip()
                if stripped in replacements:
                    value, number_format = replacements[stripped]
                    cell.value = value
                    if number_format:
                        cell.number_format = number_format
                    replaced += 1
                    continue

                next_value = original
                text_format = False
                for placeholder, (value, number_format) in replacements.items():
                    if placeholder not in next_value:
                        continue
                    next_value = next_value.replace(placeholder, str(value))
                    text_format = text_format or number_format == numbers.FORMAT_TEXT
                if next_value != original:
                    cell.value = next_value
                    if text_format:
                        cell.number_format = numbers.FORMAT_TEXT
                    replaced += 1
    return replaced


def _fill_column_placeholders(workbook: Any, columns: dict[str, list[tuple[Any, str | None]]]) -> tuple[int, int, int]:
    filled = 0
    merged_skipped = 0
    protected_skipped = 0
    for worksheet in workbook.worksheets:
        cells_to_fill: list[tuple[Any, list[tuple[Any, str | None]]]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                field_name = _whole_cell_placeholder_name(cell.value)
                if field_name is None:
                    continue
                if not field_name.startswith(_COLUMN_PREFIX):
                    continue
                header = field_name[len(_COLUMN_PREFIX) :]
                if header in columns:
                    cells_to_fill.append((cell, columns[header]))

        for anchor, values in cells_to_fill:
            current_row = anchor.row
            data_index = 0
            while data_index < len(values):
                if not _is_cell_writable(worksheet, current_row, anchor.column):
                    current_row += 1
                    merged_skipped += 1
                    continue

                target = worksheet.cell(row=current_row, column=anchor.column)
                value, number_format = values[data_index]
                if _is_cell_protected(worksheet, target):
                    _add_protected_comment(target, value)
                    data_index += 1
                    current_row += 1
                    protected_skipped += 1
                    continue

                _assign_cell_value(target, value, number_format)
                filled += 1
                data_index += 1
                current_row += 1
    return filled, merged_skipped, protected_skipped


def _fill_row_placeholders(workbook: Any, columns: dict[str, list[tuple[Any, str | None]]]) -> tuple[int, int, int]:
    filled = 0
    merged_skipped = 0
    protected_skipped = 0
    for worksheet in workbook.worksheets:
        cells_to_fill: list[tuple[Any, list[tuple[Any, str | None]]]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                field_name = _whole_cell_placeholder_name(cell.value)
                if field_name is None:
                    continue
                if not field_name.startswith(_ROW_PREFIX):
                    continue
                header = field_name[len(_ROW_PREFIX) :]
                if header in columns:
                    cells_to_fill.append((cell, columns[header]))

        for anchor, values in cells_to_fill:
            current_col = anchor.column
            data_index = 0
            while data_index < len(values):
                if not _is_cell_writable(worksheet, anchor.row, current_col):
                    current_col += 1
                    merged_skipped += 1
                    continue

                target = worksheet.cell(row=anchor.row, column=current_col)
                value, number_format = values[data_index]
                if _is_cell_protected(worksheet, target):
                    _add_protected_comment(target, value)
                    data_index += 1
                    current_col += 1
                    protected_skipped += 1
                    continue

                _assign_cell_value(target, value, number_format)
                filled += 1
                data_index += 1
                current_col += 1
    return filled, merged_skipped, protected_skipped


def _is_cell_writable(worksheet: Any, row: int, column: int) -> bool:
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return row == merged_range.min_row and column == merged_range.min_col
    return True


def _is_cell_protected(worksheet: Any, cell: Any) -> bool:
    try:
        return bool(worksheet.protection.sheet and cell.protection.locked)
    except Exception:
        return False


def _add_protected_comment(cell: Any, consumed_value: Any) -> None:
    cell.comment = Comment(f"Protected cell skipped; consumed value: {consumed_value}", "DocWen")


def _assign_cell_value(cell: Any, value: Any, number_format: str | None) -> None:
    cell.value = value
    if number_format:
        cell.number_format = number_format


def _collect_markdown_table_merge_plans(
    workbook: Any, tables: list[dict[str, Any]]
) -> tuple[list[dict[str, int]], int]:
    merge_plans: list[dict[str, int]] = []
    warnings = 0
    row_offsets_by_group: dict[tuple[str, int, int, tuple[str, ...]], int] = defaultdict(int)

    for table in tables:
        headers = [str(header) for header in table.get("headers", [])]
        rows = [[str(value) for value in row] for row in table.get("rows", [])]
        if not headers or not rows:
            continue

        candidate_regions = _find_column_placeholder_regions(workbook, headers)
        if not candidate_regions:
            continue

        merge_regions, planning_warnings = _plan_rectangular_merges(rows)
        warnings += planning_warnings

        for region in candidate_regions:
            group_key = region["group_key"]
            row_offset = row_offsets_by_group[group_key]
            for start_row, start_col, end_row, end_col in merge_regions:
                merge_plans.append(
                    {
                        "sheet_name": region["sheet_name"],
                        "start_row": region["start_row"] + row_offset + start_row,
                        "start_col": region["start_col"] + start_col,
                        "end_row": region["start_row"] + row_offset + end_row,
                        "end_col": region["start_col"] + end_col,
                    }
                )
            row_offsets_by_group[group_key] += len(rows)

    return merge_plans, warnings


def _find_column_placeholder_regions(workbook: Any, headers: list[str]) -> list[dict[str, Any]]:
    candidate_regions: list[dict[str, Any]] = []
    placeholders = {header: f"{_COLUMN_PREFIX}{header}" for header in headers}

    for worksheet in workbook.worksheets:
        grouped: dict[int, dict[str, Any]] = defaultdict(dict)
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                field_name = _whole_cell_placeholder_name(cell.value)
                if field_name is None:
                    continue
                for header, placeholder_name in placeholders.items():
                    if field_name == placeholder_name:
                        grouped[cell.row][header] = cell

        for row_index, header_cells in grouped.items():
            if any(header not in header_cells for header in headers):
                continue
            ordered_cells = [header_cells[header] for header in headers]
            ordered_cols = [cell.column for cell in ordered_cells]
            expected_cols = list(range(ordered_cols[0], ordered_cols[0] + len(headers)))
            if ordered_cols != expected_cols:
                continue
            candidate_regions.append(
                {
                    "sheet_name": worksheet.title,
                    "start_row": row_index,
                    "start_col": ordered_cols[0],
                    "group_key": (worksheet.title, row_index, ordered_cols[0], tuple(headers)),
                }
            )

    return candidate_regions


def _plan_rectangular_merges(rows: list[list[str]]) -> tuple[list[tuple[int, int, int, int]], int]:
    if not rows:
        return [], 0

    column_count = max((len(row) for row in rows), default=0)
    normalized_rows = [row + [""] * (column_count - len(row)) for row in rows]
    merge_regions: list[tuple[int, int, int, int]] = []
    warnings = 0

    for row_index, row in enumerate(normalized_rows):
        for col_index, raw_text in enumerate(row):
            marker_type = _classify_merge_marker(raw_text)
            display_text = _restore_escaped_merge_marker_literals(raw_text)
            if marker_type is not None or not display_text.strip():
                continue

            width = 1
            while col_index + width < column_count:
                if _classify_merge_marker(normalized_rows[row_index][col_index + width]) != "left":
                    break
                width += 1

            height = 1
            invalid_row = False
            scan_row = row_index + 1
            while scan_row < len(normalized_rows):
                segment = normalized_rows[scan_row][col_index : col_index + width]
                segment_markers = [_classify_merge_marker(value) for value in segment]
                if all(marker == "up" for marker in segment_markers):
                    height += 1
                    scan_row += 1
                    continue
                if any(marker is not None for marker in segment_markers):
                    invalid_row = True
                break

            if invalid_row:
                warnings += 1
                continue
            if width > 1 or height > 1:
                merge_regions.append((row_index, col_index, row_index + height - 1, col_index + width - 1))

    return merge_regions, warnings


def _classify_merge_marker(raw_text: str) -> str | None:
    stripped = raw_text.strip()
    if stripped == "<":
        return "left"
    if stripped == "^":
        return "up"
    return None


def _restore_escaped_merge_marker_literals(text: str) -> str:
    restored: list[str] = []
    index = 0
    while index < len(text):
        if text[index] == "\\" and index + 1 < len(text) and text[index + 1] in {"<", "^"}:
            restored.append(text[index + 1])
            index += 2
            continue
        restored.append(text[index])
        index += 1
    return "".join(restored)


def _apply_markdown_table_merge_plans(
    workbook: Any,
    merge_plans: list[dict[str, int]],
    template_merged_ranges: dict[str, list[str]],
) -> tuple[int, int]:
    template_bounds_by_sheet: dict[str, list[tuple[int, int, int, int]]] = defaultdict(list)
    applied_bounds_by_sheet: dict[str, list[tuple[int, int, int, int]]] = defaultdict(list)

    for sheet_name, merged_ranges in template_merged_ranges.items():
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = _required_range_boundaries(merged_range)
            template_bounds_by_sheet[sheet_name].append((min_row, min_col, max_row, max_col))

    applied = 0
    warnings = 0
    for merge_plan in merge_plans:
        sheet_name = str(merge_plan["sheet_name"])
        min_row = int(merge_plan["start_row"])
        min_col = int(merge_plan["start_col"])
        max_row = int(merge_plan["end_row"])
        max_col = int(merge_plan["end_col"])
        bounds = (min_row, min_col, max_row, max_col)

        if any(_ranges_overlap(bounds, existing) for existing in template_bounds_by_sheet.get(sheet_name, [])):
            warnings += 1
            continue
        if any(_ranges_overlap(bounds, existing) for existing in applied_bounds_by_sheet.get(sheet_name, [])):
            warnings += 1
            continue

        try:
            worksheet = workbook[sheet_name]
            worksheet.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        except Exception:
            warnings += 1
            continue
        applied_bounds_by_sheet[sheet_name].append(bounds)
        applied += 1

    return applied, warnings


def _required_range_boundaries(merged_range: str) -> tuple[int, int, int, int]:
    """Return complete bounds for an openpyxl-validated merged-cell range."""
    min_col, min_row, max_col, max_row = range_boundaries(merged_range)
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise ValueError(f"Merged-cell range has incomplete bounds: {merged_range!r}")
    return min_col, min_row, max_col, max_row


def _ranges_overlap(left_bounds: tuple[int, int, int, int], right_bounds: tuple[int, int, int, int]) -> bool:
    left_min_row, left_min_col, left_max_row, left_max_col = left_bounds
    right_min_row, right_min_col, right_max_row, right_max_col = right_bounds
    return not (
        left_max_row < right_min_row
        or right_max_row < left_min_row
        or left_max_col < right_min_col
        or right_max_col < left_min_col
    )


def _whole_cell_placeholder_name(value: str) -> str | None:
    stripped = value.strip()
    if not stripped.startswith("{{") or not stripped.endswith("}}"):
        return None
    if _PLACEHOLDER_RE.fullmatch(stripped) is None:
        return None
    return stripped[2:-2]


def _image_placeholder_re(image_scope: str | None) -> re.Pattern[str]:
    if image_scope is None:
        return _IMAGE_PLACEHOLDER_RE
    return re.compile(rf"\{{\{{IMAGE@{re.escape(image_scope)}:([^{{}}\r\n]+)\}}\}}")


def _promote_template_image_placeholders(workbook: Any, image_scope: str) -> None:
    """Trust only image markers that existed in the loaded template."""
    marker = f"IMAGE@{image_scope}"

    def promote(match: re.Match[str]) -> str:
        payload = match.group(1)
        image_path, separator, dimensions = payload.partition("|")
        encoded_path = quote(image_path, safe="/:\\._-~")
        return f"{{{{{marker}:{encoded_path}{separator}{dimensions}}}}}"

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{IMAGE:" in cell.value:
                    cell.value = _IMAGE_PLACEHOLDER_RE.sub(promote, cell.value)


def process_image_placeholders(
    workbook: Any,
    reference_path: str | Path,
    *,
    image_scope: str | None = None,
) -> int:
    """Replace canonical image placeholders with drawings anchored to their cells.

    Relative image paths are resolved beside *reference_path*.  This makes the
    same implementation usable for template workbooks and ordinary Markdown
    table exports.
    """
    from openpyxl.drawing.image import Image as XlsxImage

    inserted = 0
    reference_dir = Path(reference_path).resolve().parent
    placeholder_re = _image_placeholder_re(image_scope)
    marker = "{{IMAGE:" if image_scope is None else f"{{{{IMAGE@{image_scope}:"

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or marker not in cell.value:
                    continue

                next_value = cell.value
                for match in placeholder_re.finditer(cell.value):
                    payload = match.group(1)
                    placeholder = match.group(0)
                    image_path, width, height = _parse_image_payload(
                        payload,
                        decode_path=image_scope is not None,
                    )
                    resolved_path = _resolve_image_path(image_path, reference_dir)

                    if resolved_path is None or not _is_image_content(resolved_path):
                        next_value = next_value.replace(placeholder, "")
                        continue

                    try:
                        image = XlsxImage(str(resolved_path))
                    except Exception:
                        next_value = next_value.replace(placeholder, "")
                        continue

                    if width is not None or height is not None:
                        image.width, image.height = _scale_image(image.width, image.height, width, height)
                    elif image.width > 200:
                        scale = 200 / image.width
                        image.width = 200
                        image.height = int(image.height * scale)

                    image.anchor = cell.coordinate
                    worksheet.add_image(image)
                    _fit_cell_to_image(worksheet, cell, image.width, image.height)
                    next_value = next_value.replace(placeholder, "")
                    inserted += 1

                cell.value = next_value

    return inserted


def _parse_image_payload(
    payload: str,
    *,
    decode_path: bool = False,
) -> tuple[str, int | None, int | None]:
    normalized = payload.replace("\\|", "|")
    parts = normalized.split("|")
    image_path = parts[0].strip()
    if decode_path:
        image_path = unquote(image_path)
    width = _parse_optional_int(parts[1]) if len(parts) > 1 else None
    height = _parse_optional_int(parts[2]) if len(parts) > 2 else None
    return image_path, width, height


def _parse_optional_int(value: str) -> int | None:
    stripped = value.strip()
    return int(stripped) if stripped.isdigit() else None


def _resolve_image_path(image_path: str, template_dir: Path) -> Path | None:
    candidate = Path(image_path)
    if candidate.is_absolute():
        return candidate if candidate.is_file() else None

    for base in (Path.cwd(), template_dir):
        resolved = (base / candidate).resolve()
        if resolved.is_file():
            return resolved
    return None


def _is_image_content(path: Path) -> bool:
    try:
        detected_format = detect_content_format(str(path)).format
    except OSError:
        return False
    if get_category(detected_format) != CATEGORY_IMAGE:
        return False
    try:
        with Image.open(path) as image:
            image.load()
    except Exception:
        return False
    return True


def _scale_image(
    original_width: int, original_height: int, target_width: int | None, target_height: int | None
) -> tuple[int, int]:
    if target_width is not None and target_height is not None:
        return target_width, target_height
    if target_width is not None and original_width:
        scale = target_width / original_width
        return target_width, int(original_height * scale)
    if target_height is not None and original_height:
        scale = target_height / original_height
        return int(original_width * scale), target_height
    return original_width, original_height


def _fit_cell_to_image(worksheet: Any, cell: Any, width: int, height: int) -> None:
    row_height = height * 0.75
    current_height = worksheet.row_dimensions[cell.row].height
    if current_height is None or current_height < row_height:
        worksheet.row_dimensions[cell.row].height = row_height

    column_letter = get_column_letter(cell.column)
    column_width = width / 7.0
    current_width = worksheet.column_dimensions[column_letter].width
    if current_width is None or current_width < column_width:
        worksheet.column_dimensions[column_letter].width = column_width


def _clean_remaining_placeholders(workbook: Any) -> int:
    cleaned = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or "{{" not in cell.value:
                    continue

                def clean(match: re.Match[str]) -> str:
                    token = match.group(1)
                    if token.startswith(("IMAGE:", "IMAGE@")):
                        return match.group(0)
                    return ""

                next_value = _PLACEHOLDER_RE.sub(clean, cell.value)
                if next_value != cell.value:
                    cell.value = next_value
                    cleaned += 1
    return cleaned


def _plain_text_from_markup(value: str) -> str:
    """Project YAML/table text into an XLSX-safe plain-text cell value.

    Markdown table autolinks remain readable, ``<br>`` keeps its line break,
    and legacy HTML/entities are reduced to text. XML-forbidden control
    characters and invisible compatibility markers must never make a
    completed conversion fail while the workbook is being serialized.
    """
    cleaned = _ANGLE_URL_RE.sub(r"\1", value)
    cleaned = _ANGLE_EMAIL_RE.sub(r"\1", cleaned)
    cleaned = _HTML_BREAK_RE.sub("\n", cleaned)
    cleaned = _HTML_COMMENT_RE.sub("", cleaned)
    cleaned = _HTML_TAG_RE.sub("", cleaned)
    cleaned = html.unescape(cleaned)
    cleaned = cleaned.replace("\xa0", " ").replace("\u2002", " ").replace("\u2003", "　")
    cleaned = cleaned.translate(_ZERO_WIDTH_TRANSLATION)
    return "".join(
        character for character in cleaned if character in "\t\n\r" or unicodedata.category(character) != "Cc"
    )


def _coerce_cell_value(raw_value: Any, *, list_separator: str = "、") -> tuple[Any, str | None]:
    if raw_value is None:
        return "", None
    if isinstance(raw_value, list):
        return (
            list_separator.join(
                _plain_text_from_markup(str(item)) for item in raw_value if item not in (None, "", "null", "None")
            ),
            None,
        )
    if isinstance(raw_value, (int, float)):
        text = str(raw_value).replace(".", "").replace("-", "").replace("+", "")
        return raw_value, numbers.FORMAT_TEXT if len(text) > 15 else None

    text = _plain_text_from_markup(str(raw_value).strip())
    if _is_fraction(text):
        return f"={text}", numbers.FORMAT_NUMBER_00
    if _is_numeric_string(text):
        digit_count = len(text.replace(".", "").replace("-", "").replace("+", ""))
        return text, numbers.FORMAT_TEXT if digit_count > 15 or text.isdigit() else None
    return text, None


def _is_numeric_string(value: str) -> bool:
    if not value:
        return False
    try:
        float(value)
        return True
    except ValueError:
        return False


def _is_fraction(value: str) -> bool:
    candidate = value[1:] if value.startswith(("+", "-")) else value
    parts = candidate.split("/")
    return len(parts) == 2 and all(part.isdigit() for part in parts)
