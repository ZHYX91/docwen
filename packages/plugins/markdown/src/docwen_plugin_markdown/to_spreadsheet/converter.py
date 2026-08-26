"""MD → XLSX / CSV converters — extract Markdown tables to spreadsheet formats."""

from __future__ import annotations

import csv
import re
import secrets
import time
from pathlib import Path
from typing import TYPE_CHECKING
from urllib.parse import unquote

from openpyxl import Workbook

from docwen_core.export_semantics import LinkRuntimeConfig
from docwen_core.links import process_markdown_links, split_yaml_front_matter_source
from docwen_core.models.artifact import (
    ARTIFACT_KIND_PRIMARY,
    ArtifactManifest,
)
from docwen_core.models.result import (
    ConversionDiagnostic,
    ConversionErrorInfo,
    ConversionMetrics,
    ConversionResult,
)
from docwen_plugin_markdown.common_utils import (
    parse_md_tables,
    read_input_markdown,
    write_table_to_csv,
)
from docwen_plugin_markdown.to_spreadsheet.template_xlsx import (
    build_template_workbook,
    process_image_placeholders,
)

if TYPE_CHECKING:
    from docwen_core.protocols.execution_context import ConverterContext

MEDIA_TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MEDIA_TYPE_CSV = "text/csv"
_IMAGE_PLACEHOLDER_RE = re.compile(r"\{\{IMAGE:([^}]+)\}\}")


def _request_link_config(context: ConverterContext) -> LinkRuntimeConfig:
    """Project the active request's read-only ``link`` snapshot."""
    raw_config = context.config.get("link", {})
    if not isinstance(raw_config, dict):
        raw_config = {}
    return LinkRuntimeConfig.from_config(raw_config)


def _request_yaml_list_separator(context: ConverterContext) -> str:
    """Resolve the exact YAML list separator from this request snapshot."""
    raw = context.config.get("conversion.md_to_docx.list_separator", None)
    if raw is None:
        return "、"
    return str(raw)


def _image_placeholder_re(image_scope: str | None) -> re.Pattern[str]:
    if image_scope is None:
        return _IMAGE_PLACEHOLDER_RE
    return re.compile(rf"\{{\{{IMAGE@{re.escape(image_scope)}:([^{{}}\r\n]+)\}}\}}")


def _csv_image_fallbacks(
    content: str,
    *,
    image_scope: str | None = None,
) -> str:
    """Downgrade binary image placeholders to portable filename text for CSV."""

    def _filename(match: re.Match[str]) -> str:
        payload = match.group(1).replace(r"\|", "|")
        image_path = payload.split("|", 1)[0].strip().replace("\\", "/")
        if image_scope is not None:
            image_path = unquote(image_path)
        return image_path.rsplit("/", 1)[-1]

    return _image_placeholder_re(image_scope).sub(_filename, content)


def _csv_safe_sheet_name(sheet_name: str) -> str:
    """Return a stable, filesystem-safe sheet-name fragment for CSV filenames."""
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", sheet_name.strip().replace(" ", "_"))
    cleaned = cleaned.strip("._")
    return cleaned or "Sheet"


def _write_worksheet_to_csv(worksheet, output_path: str | Path) -> None:
    """Write an openpyxl worksheet as UTF-8 CSV with a spreadsheet-friendly BOM."""
    with Path(output_path).open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])


class MdToXlsxConverter:
    """Convert Markdown tables to an XLSX spreadsheet.

    Each Markdown table becomes a separate sheet in the workbook.
    Table headings become the header row in each sheet.
    """

    def convert(self, context: ConverterContext) -> ConversionResult:
        t_start = time.monotonic()
        task_id = context.request.request_id
        cancellable = context.cancellation
        logger = context.logger
        progress = context.progress
        workspace = context.workspace

        try:
            cancellable.check()

            # ── Read input ──────────────────────────────────────────
            input_path = workspace.input_path
            progress.report_progress(5.0, "Reading Markdown input")
            content, input_bytes = read_input_markdown(input_path)
            yaml_front, markdown_body = split_yaml_front_matter_source(content)
            image_scope = secrets.token_urlsafe(24)
            markdown_body = process_markdown_links(
                markdown_body,
                input_path,
                link_config=_request_link_config(context),
                target_format="xlsx",
                table_safe=True,
                temp_dir=str(workspace.staging_dir),
                image_scope=image_scope,
            )
            content = yaml_front + markdown_body

            # ── Build XLSX workbook ─────────────────────────────────
            cancellable.check()
            progress.report_progress(50.0, "Building XLSX workbook")

            template_name = str(getattr(context.request, "options", {}).get("template_name", "") or "").strip()
            template_stats: dict[str, int] = {}
            if template_name:
                template_path = Path(template_name)
                if not template_path.is_file():
                    message = f"MD→XLSX template file not found: {template_name}"
                    logger.warning(message)
                    return ConversionResult(
                        task_id=task_id,
                        success=False,
                        error=ConversionErrorInfo(
                            error_type="invalid_input",
                            message=message,
                            diagnostic_code="MD2XLSX-TEMPLATE-NOT-FOUND",
                        ),
                        metrics=ConversionMetrics(duration_ms=(time.monotonic() - t_start) * 1000.0),
                        diagnostics=[
                            ConversionDiagnostic(
                                level="error",
                                message=message,
                                code="MD2XLSX-TEMPLATE-NOT-FOUND",
                            )
                        ],
                    )
                wb, template_stats = build_template_workbook(
                    content,
                    template_path,
                    source_stem=Path(input_path).stem,
                    image_scope=image_scope,
                    list_separator=_request_yaml_list_separator(context),
                )
                actual_table_count = template_stats.get("table_count", 0)
            else:
                # ── Parse tables ────────────────────────────────────
                cancellable.check()
                progress.report_progress(30.0, "Parsing Markdown tables")
                tables = parse_md_tables(markdown_body)

                if not tables:
                    # Create an empty workbook for MDs with no tables
                    tables = [{"headers": [], "rows": []}]
                    actual_table_count = 0
                else:
                    actual_table_count = len(tables)

                wb = Workbook()
                # Remove default sheet
                wb.remove(wb.active)  # pyright: ignore[reportArgumentType]

                for idx, table in enumerate(tables):
                    sheet_name = f"Table_{idx + 1}" if idx > 0 else "Sheet1"
                    ws = wb.create_sheet(title=sheet_name[:31])

                    # Headers
                    if table["headers"]:
                        for col_idx, header in enumerate(table["headers"], 1):
                            ws.cell(row=1, column=col_idx, value=header)

                    # Data rows
                    start_row = 2 if table["headers"] else 1
                    for row_idx, row in enumerate(table["rows"], start_row):
                        cancellable.check()
                        for col_idx, val in enumerate(row, 1):
                            ws.cell(row=row_idx, column=col_idx, value=val)

                process_image_placeholders(
                    wb,
                    input_path,
                    image_scope=image_scope,
                )

            # ── Write to staging ────────────────────────────────────
            cancellable.check()
            progress.report_progress(80.0, "Writing XLSX to staging")

            output_path = workspace.create_artifact_path(ARTIFACT_KIND_PRIMARY, ".xlsx")
            input_stem = Path(input_path).stem
            suggested_name = f"{input_stem}.xlsx"

            wb.save(output_path)

            output_bytes = Path(output_path).stat().st_size

            # ── Register artifact ───────────────────────────────────
            artifact = ArtifactManifest(
                artifact_id=f"{task_id}-xlsx",
                kind=ARTIFACT_KIND_PRIMARY,
                staging_path=output_path,
                suggested_name=suggested_name,
                media_type=MEDIA_TYPE_XLSX,
                is_primary=True,
                metadata={
                    "source_format": "markdown",
                    "target_format": "xlsx",
                    "sheet_count": len(wb.sheetnames),
                    "table_count": actual_table_count,
                    **({"template_name": template_name, **template_stats} if template_name else {}),
                },
            )
            workspace.add_artifact(artifact)

            elapsed_ms = (time.monotonic() - t_start) * 1000.0
            progress.report_progress(100.0, "Done")

            logger.info(f"MD→XLSX complete: {input_path} → {suggested_name} ({actual_table_count} table(s))")

            return ConversionResult(
                task_id=task_id,
                success=True,
                artifacts=[artifact],
                metrics=ConversionMetrics(
                    duration_ms=elapsed_ms,
                    input_bytes=input_bytes,
                    output_bytes=output_bytes,
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="info",
                        message=(f"MD→XLSX conversion successful ({actual_table_count} table(s))"),
                        code="MD2XLSX-OK",
                    )
                ],
            )

        except Exception as exc:
            elapsed_ms = (time.monotonic() - t_start) * 1000.0
            logger.error(f"MD→XLSX failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=str(exc),
                    diagnostic_code="MD2XLSX-ERROR",
                ),
                metrics=ConversionMetrics(duration_ms=elapsed_ms),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"MD→XLSX conversion failed: {exc}",
                        code="MD2XLSX-ERROR",
                    )
                ],
            )


class MdToCsvConverter:
    """Convert Markdown tables to CSV files.

    Each Markdown table becomes a separate CSV file.
    Multiple tables produce multiple artifacts.
    """

    def convert(self, context: ConverterContext) -> ConversionResult:
        t_start = time.monotonic()
        task_id = context.request.request_id
        cancellable = context.cancellation
        logger = context.logger
        progress = context.progress
        workspace = context.workspace

        try:
            cancellable.check()

            # ── Read input ──────────────────────────────────────────
            input_path = workspace.input_path
            progress.report_progress(5.0, "Reading Markdown input")
            content, input_bytes = read_input_markdown(input_path)
            yaml_front, markdown_body = split_yaml_front_matter_source(content)
            image_scope = secrets.token_urlsafe(24)
            markdown_body = process_markdown_links(
                markdown_body,
                input_path,
                link_config=_request_link_config(context),
                target_format="csv",
                table_safe=True,
                temp_dir=str(workspace.staging_dir),
                image_scope=image_scope,
            )
            markdown_body = _csv_image_fallbacks(
                markdown_body,
                image_scope=image_scope,
            )
            content = yaml_front + markdown_body

            template_name = str(getattr(context.request, "options", {}).get("template_name", "") or "").strip()
            if template_name:
                template_path = Path(template_name)
                if not template_path.is_file():
                    message = f"MD→CSV template file not found: {template_name}"
                    logger.warning(message)
                    return ConversionResult(
                        task_id=task_id,
                        success=False,
                        error=ConversionErrorInfo(
                            error_type="invalid_input",
                            message=message,
                            diagnostic_code="MD2CSV-TEMPLATE-NOT-FOUND",
                        ),
                        metrics=ConversionMetrics(duration_ms=(time.monotonic() - t_start) * 1000.0),
                        diagnostics=[
                            ConversionDiagnostic(
                                level="error",
                                message=message,
                                code="MD2CSV-TEMPLATE-NOT-FOUND",
                            )
                        ],
                    )

                cancellable.check()
                progress.report_progress(45.0, "Building XLSX template workbook")
                workbook, template_stats = build_template_workbook(
                    content,
                    template_path,
                    source_stem=Path(input_path).stem,
                    image_scope=image_scope,
                    list_separator=_request_yaml_list_separator(context),
                )
                return self._convert_template_workbook_to_csv_artifacts(
                    workbook=workbook,
                    template_stats=template_stats,
                    context=context,
                    input_path=input_path,
                    input_bytes=input_bytes,
                    t_start=t_start,
                    template_name=template_name,
                )

            # ── Parse tables ────────────────────────────────────────
            cancellable.check()
            progress.report_progress(30.0, "Parsing Markdown tables")
            tables = parse_md_tables(markdown_body)

            if not tables:
                return ConversionResult(
                    task_id=task_id,
                    success=False,
                    error=ConversionErrorInfo(
                        error_type="conversion_failed",
                        message="No tables found in Markdown input",
                        diagnostic_code="MD2CSV-NO-TABLES",
                    ),
                    metrics=ConversionMetrics(duration_ms=(time.monotonic() - t_start) * 1000.0),
                    diagnostics=[
                        ConversionDiagnostic(
                            level="error",
                            message="No tables found in Markdown input",
                            code="MD2CSV-NO-TABLES",
                        )
                    ],
                )

            # ── Write each table to a CSV in staging ─────────────────
            cancellable.check()
            progress.report_progress(50.0, "Writing CSV files to staging")

            input_stem = Path(input_path).stem
            artifacts: list[ArtifactManifest] = []
            total_output_bytes = 0

            for idx, table in enumerate(tables):
                cancellable.check()
                suffix = f"_{idx + 1}.csv" if idx > 0 else ".csv"
                csv_path = workspace.create_artifact_path(ARTIFACT_KIND_PRIMARY, suffix)
                suggested = f"{input_stem}_{idx + 1}.csv" if idx > 0 else f"{input_stem}.csv"

                write_table_to_csv(table, csv_path)

                file_bytes = Path(csv_path).stat().st_size
                total_output_bytes += file_bytes

                artifact = ArtifactManifest(
                    artifact_id=f"{task_id}-csv-{idx}",
                    kind=ARTIFACT_KIND_PRIMARY,
                    staging_path=csv_path,
                    suggested_name=suggested,
                    media_type=MEDIA_TYPE_CSV,
                    is_primary=(idx == 0),
                    metadata={
                        "source_format": "markdown",
                        "target_format": "csv",
                        "table_index": idx,
                        "row_count": len(table["rows"]),
                        "header_count": len(table["headers"]),
                    },
                )
                workspace.add_artifact(artifact)
                artifacts.append(artifact)

            elapsed_ms = (time.monotonic() - t_start) * 1000.0
            progress.report_progress(100.0, "Done")

            logger.info(f"MD→CSV complete: {input_path} → {len(artifacts)} file(s), {len(tables)} table(s)")

            return ConversionResult(
                task_id=task_id,
                success=True,
                artifacts=artifacts,
                metrics=ConversionMetrics(
                    duration_ms=elapsed_ms,
                    input_bytes=input_bytes,
                    output_bytes=total_output_bytes,
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="info",
                        message=(f"MD→CSV conversion successful ({len(tables)} table(s))"),
                        code="MD2CSV-OK",
                    )
                ],
            )

        except Exception as exc:
            elapsed_ms = (time.monotonic() - t_start) * 1000.0
            logger.error(f"MD→CSV failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=str(exc),
                    diagnostic_code="MD2CSV-ERROR",
                ),
                metrics=ConversionMetrics(duration_ms=elapsed_ms),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"MD→CSV conversion failed: {exc}",
                        code="MD2CSV-ERROR",
                    )
                ],
            )

    def _convert_template_workbook_to_csv_artifacts(
        self,
        *,
        workbook,
        template_stats: dict[str, int],
        context: ConverterContext,
        input_path: str,
        input_bytes: int,
        t_start: float,
        template_name: str,
    ) -> ConversionResult:
        task_id = context.request.request_id
        cancellable = context.cancellation
        logger = context.logger
        progress = context.progress
        workspace = context.workspace

        input_stem = Path(input_path).stem
        folder_name = f"{input_stem}_fromMd"
        artifacts: list[ArtifactManifest] = []
        total_output_bytes = 0

        cancellable.check()
        progress.report_progress(70.0, "Writing template workbook sheets to CSV")

        for idx, sheet_name in enumerate(workbook.sheetnames):
            cancellable.check()
            worksheet = workbook[sheet_name]
            csv_path = workspace.create_artifact_path(ARTIFACT_KIND_PRIMARY, f"_{idx}.csv")
            safe_sheet = _csv_safe_sheet_name(sheet_name)
            suggested = f"{folder_name}/{input_stem}_{safe_sheet}_fromMd.csv"

            _write_worksheet_to_csv(worksheet, csv_path)

            file_bytes = Path(csv_path).stat().st_size
            total_output_bytes += file_bytes

            artifact = ArtifactManifest(
                artifact_id=f"{task_id}-csv-template-{idx}",
                kind=ARTIFACT_KIND_PRIMARY,
                staging_path=csv_path,
                suggested_name=suggested,
                media_type=MEDIA_TYPE_CSV,
                is_primary=(idx == 0),
                metadata={
                    "source_format": "markdown",
                    "target_format": "csv",
                    "template_name": template_name,
                    "csv_output_folder": folder_name,
                    "sheet_name": sheet_name,
                    "sheet_index": idx,
                    **template_stats,
                },
            )
            workspace.add_artifact(artifact)
            artifacts.append(artifact)

        elapsed_ms = (time.monotonic() - t_start) * 1000.0
        progress.report_progress(100.0, "Done")
        logger.info(f"MD→CSV template chain complete: {input_path} → {folder_name} ({len(artifacts)} sheet(s))")

        return ConversionResult(
            task_id=task_id,
            success=True,
            artifacts=artifacts,
            metrics=ConversionMetrics(
                duration_ms=elapsed_ms,
                input_bytes=input_bytes,
                output_bytes=total_output_bytes,
            ),
            diagnostics=[
                ConversionDiagnostic(
                    level="info",
                    message=(f"MD→CSV template conversion successful ({len(artifacts)} sheet(s))"),
                    code="MD2CSV-TEMPLATE-OK",
                )
            ],
        )
