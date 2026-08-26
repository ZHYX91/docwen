"""Focused tests split from test_md_to_spreadsheet.py."""

from __future__ import annotations

from ._md_to_spreadsheet_support import (
    SAMPLE_MD_TABLES,
    MdToCsvConverter,
    MdToXlsxConverter,
    Path,
    _active_worksheet,
    _cell_value,
    _old_system_spreadsheet_release_gate_fixture,
    _png_bytes,
    make_context,
    pytest,
    write_temp_md,
)


class TestMdToXlsx:
    """Tests for MD → XLSX conversion."""

    @pytest.mark.contract
    def test_template_name_embeds_markdown_file_as_table_safe_cell_text(self, tmp_path: Path):
        """Old Tk/PySide6 table-safe embeds become one XLSX cell with literal pipes."""
        from openpyxl import Workbook, load_workbook

        case = _old_system_spreadsheet_release_gate_fixture()["cases"]["table_safe_embedded_md_file"]
        embedded = tmp_path / "embed.md"
        embedded.write_text(case["embedded_markdown"], encoding="utf-8")

        template = tmp_path / "embedded-cell-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Sheet1"
        ws["A1"] = "{{↓Name}}"
        ws["B1"] = "{{↓Age}}"
        wb.save(template)
        wb.close()

        md_path = tmp_path / "source.md"
        md_path.write_text(case["input_markdown"], encoding="utf-8")
        ctx, _workspace = make_context(
            str(md_path),
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        result = MdToXlsxConverter().convert(ctx)

        assert result.success
        out_wb = load_workbook(Path(result.artifacts[0].staging_path))
        out_ws = _active_worksheet(out_wb)
        assert {coord: _cell_value(out_ws, coord) for coord in case["expected_cells"]} == case["expected_cells"]
        out_wb.close()

    @pytest.mark.contract
    def test_template_name_ignores_table_shape_inside_fenced_code(self, tmp_path: Path):
        """Old PySide6 release-gate guards ignored table-looking text in code fences."""
        from openpyxl import Workbook, load_workbook

        case = _old_system_spreadsheet_release_gate_fixture()["cases"]["fenced_code_table_ignored"]
        template = tmp_path / "fenced-code-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Sheet1"
        ws["A1"] = "{{↓Name}}"
        ws["B1"] = "{{↓Age}}"
        wb.save(template)
        wb.close()

        md_path = write_temp_md(case["input_markdown"])
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        out_wb = load_workbook(Path(result.artifacts[0].staging_path))
        out_ws = _active_worksheet(out_wb)
        assert {coord: _cell_value(out_ws, coord) for coord in case["expected_cells"]} == case["expected_cells"]
        for key, value in case["expected_metadata"].items():
            assert result.artifacts[0].metadata[key] == value
        out_wb.close()

    @pytest.mark.contract
    def test_template_image_placeholders_use_content_not_suffix(self, tmp_path: Path):
        """Real wrong-suffix images work; fake/corrupt image names fail closed."""
        from openpyxl import Workbook, load_workbook

        image_path = tmp_path / "logo.resource"
        image_path.write_bytes(_png_bytes())
        (tmp_path / "fake.png").write_bytes(b"\x00\x01\x02\x03" * 8)
        (tmp_path / "corrupt.png").write_bytes(b"\x89PNG\r\n\x1a\ntruncated-image")

        template = tmp_path / "image-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Report"
        ws["A1"] = "{{title}}"
        ws["A3"] = "{{IMAGE:logo.resource|32|16}}"
        ws["A4"] = "{{IMAGE:fake.png}}"
        ws["A5"] = "{{IMAGE:corrupt.png}}"
        ws["A6"] = "{{IMAGE:missing.png}}"
        wb.save(template)
        wb.close()

        md_content = """---
title: Image Report
---

| Name |
| --- |
| Alice |
"""
        md_path = write_temp_md(md_content)
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        output = Path(result.artifacts[0].staging_path)
        out_wb = load_workbook(output)
        out_ws = out_wb["Report"]
        assert out_ws["A1"].value == "Image Report"
        assert out_ws["A3"].value is None
        assert out_ws["A4"].value is None
        assert out_ws["A5"].value is None
        assert out_ws["A6"].value is None
        images = vars(out_ws).get("_images")  # openpyxl exposes loaded images here.
        assert isinstance(images, list)
        assert len(images) == 1
        assert result.artifacts[0].metadata["image_placeholders"] == 1
        out_wb.close()

    @pytest.mark.contract
    def test_scoped_template_image_placeholder_decodes_quoted_space_path(self, tmp_path: Path) -> None:
        """Trusted template image paths survive request scoping and URL quoting."""
        from openpyxl import Workbook, load_workbook

        image_path = tmp_path / "space name.png"
        image_path.write_bytes(_png_bytes())
        template = tmp_path / "spaced-image-template.xlsx"
        workbook = Workbook()
        worksheet = _active_worksheet(workbook)
        worksheet["A1"] = "{{IMAGE:space name.png|8|4}}"
        workbook.save(template)
        workbook.close()

        source = write_temp_md("| Name |\n| --- |\n| Alice |\n")
        context, _workspace = make_context(
            source,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        result = MdToXlsxConverter().convert(context)

        assert result.success
        output = load_workbook(Path(result.artifacts[0].staging_path))
        output_sheet = _active_worksheet(output)
        assert output_sheet["A1"].value is None
        images = vars(output_sheet).get("_images")
        assert isinstance(images, list)
        assert len(images) == 1
        assert result.artifacts[0].metadata["image_placeholders"] == 1
        output.close()

    @pytest.mark.contract
    def test_template_name_exports_csv_folder_chain_from_xlsx_template(self, tmp_path: Path):
        """CSV targets can use the old MD->XLSX template chain and folder-style sheet names."""
        import csv

        from openpyxl import Workbook

        template = tmp_path / "csv-template.xlsx"
        down_arrow = chr(0x2193)
        right_arrow = chr(0x2192)
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Report Sheet"
        ws["A1"] = "{{title}}"
        ws["A3"] = "{{" + down_arrow + "Name}}"
        ws["B3"] = "{{" + down_arrow + "Age}}"
        summary = wb.create_sheet("Summary")
        summary["A1"] = "{{owner}}"
        summary["A3"] = "{{" + right_arrow + "Name}}"
        wb.save(template)
        wb.close()

        md_content = """---
title: CSV Report
owner: Finance
---

| Name | Age |
| --- | --- |
| Alice | 30 |
| Bob | 25 |
"""
        md_path = write_temp_md(md_content)
        ctx, _workspace = make_context(
            md_path,
            target_format="csv",
            options={"template_name": str(template)},
        )

        converter = MdToCsvConverter()
        result = converter.convert(ctx)

        assert result.success
        assert len(result.artifacts) == 2
        first, second = result.artifacts
        input_stem = Path(md_path).stem
        assert first.is_primary is True
        assert first.suggested_name == f"{input_stem}_fromMd/{input_stem}_Report_Sheet_fromMd.csv"
        assert second.suggested_name == f"{input_stem}_fromMd/{input_stem}_Summary_fromMd.csv"
        assert first.metadata["template_name"] == str(template)
        assert first.metadata["csv_output_folder"] == f"{input_stem}_fromMd"
        assert first.metadata["yaml_placeholders"] >= 1
        assert first.metadata["column_placeholders"] == 4
        assert second.metadata["row_placeholders"] == 2
        assert result.diagnostics[0].code == "MD2CSV-TEMPLATE-OK"

        with Path(first.staging_path).open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
        assert rows[0][0] == "CSV Report"
        assert rows[2][:2] == ["Alice", "30"]
        assert rows[3][:2] == ["Bob", "25"]

        with Path(second.staging_path).open(encoding="utf-8-sig", newline="") as handle:
            summary_rows = list(csv.reader(handle))
        assert summary_rows[0][0] == "Finance"
        assert summary_rows[2][:2] == ["Alice", "Bob"]

    @pytest.mark.contract
    def test_template_name_skips_merged_non_anchor_and_protected_cells(self, tmp_path: Path):
        """XLSX template fills preserve old merged/protected-cell skip semantics."""
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Protection

        template = tmp_path / "merged-protected-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Report"
        ws["A3"] = "{{↓Name}}"
        ws["A3"].protection = Protection(locked=False)
        ws.merge_cells("A3:A4")
        ws["A5"].protection = Protection(locked=False)
        ws["A8"] = "{{→Age}}"
        ws["A8"].protection = Protection(locked=False)
        ws.merge_cells("A8:B8")
        ws["C8"].protection = Protection(locked=False)
        ws["C3"] = "{{↓City}}"
        ws["C3"].protection = Protection(locked=False)
        ws["C5"].protection = Protection(locked=False)
        ws.protection.sheet = True
        wb.save(template)
        wb.close()

        md_content = """| Name | Age | City |
| --- | --- | --- |
| Alice | 30 | Paris |
| Bob | 25 | Berlin |
"""
        md_path = write_temp_md(md_content)
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        output = Path(result.artifacts[0].staging_path)
        out_wb = load_workbook(output)
        out_ws = out_wb["Report"]
        assert "A3:A4" in {str(merged_range) for merged_range in out_ws.merged_cells.ranges}
        assert "A8:B8" in {str(merged_range) for merged_range in out_ws.merged_cells.ranges}
        assert out_ws["A3"].value == "Alice"
        assert out_ws["A5"].value == "Bob"
        assert out_ws["A8"].value == "30"
        assert out_ws["C8"].value == "25"
        assert out_ws["C3"].value == "Paris"
        assert out_ws["C4"].value is None
        assert out_ws["C4"].comment is not None
        assert "Berlin" in out_ws["C4"].comment.text
        assert result.artifacts[0].metadata["merged_cells_skipped"] == 2
        assert result.artifacts[0].metadata["protected_cells_skipped"] == 1
        out_wb.close()

    @pytest.mark.contract
    def test_template_name_restores_markdown_table_merge_markers(self, tmp_path: Path):
        """XLSX templates restore Markdown table < and ^ merge markers."""
        from openpyxl import Workbook, load_workbook

        template = tmp_path / "merge-marker-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Report"
        ws["A3"] = "{{↓Name}}"
        ws["B3"] = "{{↓Age}}"
        ws["C3"] = "{{↓Literal}}"
        wb.save(template)
        wb.close()

        md_content = r"""| Name | Age | Literal |
| --- | --- | --- |
| Alice | < | \< |
| ^ | ^ | \^ |
| Bob | 25 | Plain |
"""
        md_path = write_temp_md(md_content)
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        output = Path(result.artifacts[0].staging_path)
        out_wb = load_workbook(output)
        out_ws = out_wb["Report"]
        merged_ranges = {str(merged_range) for merged_range in out_ws.merged_cells.ranges}
        assert "A3:B4" in merged_ranges
        assert out_ws["A3"].value == "Alice"
        assert out_ws["A5"].value == "Bob"
        assert out_ws["B5"].value == "25"
        assert out_ws["C3"].value == "<"
        assert out_ws["C4"].value == "^"
        assert result.artifacts[0].metadata["markdown_table_merges"] == 1
        assert result.artifacts[0].metadata["markdown_table_merge_warnings"] == 0
        out_wb.close()

    @pytest.mark.contract
    def test_template_name_degrades_invalid_markdown_merge_markers_to_plain_text(self, tmp_path: Path):
        """Invalid old-system merge-marker rectangles should not create partial XLSX merges."""
        from openpyxl import Workbook, load_workbook

        case = _old_system_spreadsheet_release_gate_fixture()["cases"]["invalid_merge_degrades_to_plain_text"]
        template = tmp_path / "invalid-merge-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Sheet1"
        ws["A1"] = "{{↓A}}"
        ws["B1"] = "{{↓B}}"
        ws["C1"] = "{{↓C}}"
        wb.save(template)
        wb.close()

        md_path = write_temp_md(case["input_markdown"])
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        result = MdToXlsxConverter().convert(ctx)

        assert result.success
        out_wb = load_workbook(Path(result.artifacts[0].staging_path))
        out_ws = _active_worksheet(out_wb)
        assert {str(merged_range) for merged_range in out_ws.merged_cells.ranges} == set(case["expected_merged_ranges"])
        assert {coord: _cell_value(out_ws, coord) for coord in case["expected_cells"]} == case["expected_cells"]
        for key, value in case["expected_metadata"].items():
            assert result.artifacts[0].metadata[key] == value
        out_wb.close()

    @pytest.mark.contract
    def test_template_name_keeps_template_merge_when_markdown_merge_conflicts(self, tmp_path: Path):
        """Markdown merge plans should not overwrite existing template merge geometry."""
        from openpyxl import Workbook, load_workbook

        case = _old_system_spreadsheet_release_gate_fixture()["cases"]["template_merge_conflict_keeps_template_merge"]
        template = tmp_path / "template-merge-conflict.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Sheet1"
        ws["A1"] = "{{↓A}}"
        ws["B1"] = "{{↓B}}"
        ws.merge_cells("A1:A2")
        wb.save(template)
        wb.close()

        md_path = write_temp_md(case["input_markdown"])
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        result = MdToXlsxConverter().convert(ctx)

        assert result.success
        out_wb = load_workbook(Path(result.artifacts[0].staging_path))
        out_ws = _active_worksheet(out_wb)
        assert {str(merged_range) for merged_range in out_ws.merged_cells.ranges} == set(case["expected_merged_ranges"])
        assert {coord: _cell_value(out_ws, coord) for coord in case["expected_cells"]} == case["expected_cells"]
        for key, value in case["expected_metadata"].items():
            assert result.artifacts[0].metadata[key] == value
        out_wb.close()

    @pytest.mark.contract
    def test_missing_template_name_returns_structured_error(self):
        """Missing template paths fail clearly instead of falling back silently."""
        md_path = write_temp_md(SAMPLE_MD_TABLES)
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": "/missing/report-template.xlsx"},
        )

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert not result.success
        assert result.artifacts == []
        assert result.error is not None
        assert result.error.error_type == "invalid_input"
        assert result.error.diagnostic_code == "MD2XLSX-TEMPLATE-NOT-FOUND"
        assert result.diagnostics[0].code == "MD2XLSX-TEMPLATE-NOT-FOUND"
