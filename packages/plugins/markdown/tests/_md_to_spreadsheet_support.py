"""Semantic parity tests for MD → XLSX and MD → CSV conversions."""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image

from docwen_core.export_semantics import LinkRuntimeConfig
from docwen_core.links import make_table_safe
from docwen_plugin_markdown.common_utils import parse_raw_md_tables
from docwen_plugin_markdown.to_spreadsheet import converter as spreadsheet_converter
from docwen_plugin_markdown.to_spreadsheet.converter import (
    MdToCsvConverter,
    MdToXlsxConverter,
)
from docwen_plugin_markdown.to_spreadsheet.template_xlsx import (
    _plain_text_from_markup,
    _required_range_boundaries,
)

from .conftest import (
    SAMPLE_MD_TABLES,
    make_context,
    write_temp_md,
)


def _old_system_spreadsheet_smoke_fixture() -> dict:
    project_root = Path(__file__).resolve().parents[4]
    fixture_path = project_root / "tests" / "fixtures" / "golden" / "markdown_xlsx_old_system_smoke_semantics.json"
    return json.loads(fixture_path.read_text(encoding="utf-8"))


def _png_bytes() -> bytes:
    with BytesIO() as buffer:
        Image.new("RGB", (2, 2), (40, 80, 120)).save(buffer, format="PNG")
        return buffer.getvalue()


def _old_system_spreadsheet_release_gate_fixture() -> dict:
    project_root = Path(__file__).resolve().parents[4]
    fixture_path = project_root / "tests" / "fixtures" / "golden" / "old_system_md_xlsx_release_gate_semantics.json"
    return json.loads(fixture_path.read_text(encoding="utf-8"))


def _bundled_spreadsheet_template_id() -> str:
    from docwen_runtime.templates import TemplateRegistry

    project_root = Path(__file__).resolve().parents[4]
    expected_path = (project_root / "templates" / "English Sample Sheet Template.xlsx").resolve()
    matches = [
        template.id
        for template in TemplateRegistry.default().list_templates("xlsx")
        if template.path.resolve() == expected_path
    ]
    assert len(matches) == 1
    return matches[0]


def _active_worksheet(workbook: Workbook) -> Worksheet:
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    return worksheet


def _cell_value(worksheet: Worksheet, coordinate: str) -> object:
    row, column = coordinate_to_tuple(coordinate)
    return worksheet.cell(row=row, column=column).value


def _execute_markdown_runtime(
    input_file: Path,
    target_format: str,
    output_dir: Path,
    options: dict | None = None,
):
    from docwen_core.models.file_ref import FileRef
    from docwen_core.models.request import ConversionRequest, OutputPolicy
    from docwen_plugin_markdown.plugin import MarkdownPlugin
    from docwen_runtime.engine.route_resolver import RouteResolver
    from docwen_runtime.engine.task_manager import TaskManager
    from docwen_runtime.output.finalizer import OutputFinalizer
    from docwen_runtime.plugin_registry.registry import PluginRegistry
    from docwen_runtime.workspace.manager import WorkspaceManager

    registry = PluginRegistry()
    registry.register(MarkdownPlugin())
    task_manager = TaskManager(
        registry,
        RouteResolver(registry),
        WorkspaceManager(root_dir=str(output_dir.parent / "workspace")),
        OutputFinalizer(),
    )
    request = ConversionRequest(
        request_id=f"md-{target_format}-old-system-runtime-smoke",
        input_refs=[
            FileRef(
                path=str(input_file),
                format="markdown",
                category="markdown",
                size_bytes=input_file.stat().st_size,
            )
        ],
        target_format=target_format,
        options=options or {},
        output_policy=OutputPolicy(output_dir=str(output_dir)),
    )
    return task_manager.execute_single(request)


__all__ = (
    "SAMPLE_MD_TABLES",
    "LinkRuntimeConfig",
    "MdToCsvConverter",
    "MdToXlsxConverter",
    "Path",
    "Workbook",
    "_active_worksheet",
    "_bundled_spreadsheet_template_id",
    "_cell_value",
    "_execute_markdown_runtime",
    "_old_system_spreadsheet_release_gate_fixture",
    "_old_system_spreadsheet_smoke_fixture",
    "_plain_text_from_markup",
    "_png_bytes",
    "_required_range_boundaries",
    "load_workbook",
    "make_context",
    "make_table_safe",
    "parse_raw_md_tables",
    "pytest",
    "spreadsheet_converter",
    "write_temp_md",
)
