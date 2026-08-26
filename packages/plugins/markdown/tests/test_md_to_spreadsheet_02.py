"""Focused tests split from test_md_to_spreadsheet.py."""

from __future__ import annotations

from ._md_to_spreadsheet_support import (
    SAMPLE_MD_TABLES,
    MdToCsvConverter,
    MdToXlsxConverter,
    Path,
    Workbook,
    _active_worksheet,
    _bundled_spreadsheet_template_id,
    _cell_value,
    _execute_markdown_runtime,
    _old_system_spreadsheet_release_gate_fixture,
    _old_system_spreadsheet_smoke_fixture,
    _plain_text_from_markup,
    _png_bytes,
    load_workbook,
    make_context,
    pytest,
    write_temp_md,
)


class TestMdToXlsx:
    """Tests for MD → XLSX conversion."""

    @pytest.mark.parametrize(
        ("source", "expected"),
        [
            ("<b>A</b>&nbsp;B\u200b", "A B"),
            ("A<br/>B<!--hidden-->C", "A\nBC"),
            ("X&#20013;&#x6587;Y\x00\x1f\x7fZ", "X中文YZ"),
            ("<https://example.com> <user@example.com>", "https://example.com user@example.com"),
        ],
    )
    @pytest.mark.contract
    def test_xlsx_plain_text_projection_preserves_old_cleanup_contract(self, source: str, expected: str) -> None:
        assert _plain_text_from_markup(source) == expected

    @pytest.mark.contract
    def test_template_text_cells_clean_markup_and_illegal_xml_controls(self, tmp_path: Path) -> None:
        template = tmp_path / "plain-text-template.xlsx"
        workbook = Workbook()
        worksheet = _active_worksheet(workbook)
        worksheet["A1"] = "{{title}}"
        worksheet["B1"] = "Title: {{title}}"
        worksheet["A3"] = "{{↓Name}}"
        workbook.save(template)
        workbook.close()

        markdown = tmp_path / "plain-text-input.md"
        markdown.write_text(
            '---\ntitle: "<b>A</b>&nbsp;B\\u200b<br/>C<!--hidden-->&#x4E2D;\\u0001"\n---\n\n'
            "| Name |\n| --- |\n| <i>Table</i>&nbsp;Value\u200b\x01 |\n",
            encoding="utf-8",
        )
        context, _workspace = make_context(
            str(markdown),
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        result = MdToXlsxConverter().convert(context)

        assert result.success, result.error.message if result.error else "conversion failed"
        output = load_workbook(result.artifacts[0].staging_path, data_only=False)
        output_sheet = _active_worksheet(output)
        assert output_sheet["A1"].value == "A B\nC中"
        assert output_sheet["B1"].value == "Title: A B\nC中"
        assert output_sheet["A3"].value == "Table Value"
        output.close()

    @pytest.mark.contract
    def test_loads_content_validated_xlsx_template_with_wrong_suffix(self, tmp_path: Path) -> None:
        template_path = tmp_path / "renamed-template.docx"
        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet["A1"] = "{{title}}"
        workbook.save(template_path)
        md_path = tmp_path / "input.md"
        md_path.write_text("---\ntitle: Content first\n---\n", encoding="utf-8")
        context, _workspace = make_context(
            str(md_path),
            target_format="xlsx",
            options={"template_name": str(template_path)},
        )

        result = MdToXlsxConverter().convert(context)

        assert result.success is True
        with Path(result.artifacts[0].staging_path).open("rb") as stream:
            loaded = load_workbook(stream)
        loaded_worksheet = loaded.active
        assert loaded_worksheet is not None
        assert loaded_worksheet["A1"].value == "Content first"

    @pytest.mark.contract
    def test_converts_tables_to_xlsx(self):
        """Markdown tables are converted to XLSX with correct structure."""
        md_path = write_temp_md(SAMPLE_MD_TABLES)
        ctx, _workspace = make_context(md_path, target_format="xlsx")

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success, f"Conversion failed: {result.error.message if result.error else 'unknown'}"
        assert len(result.artifacts) == 1
        artifact = result.artifacts[0]
        assert artifact.kind == "primary"
        assert artifact.is_primary is True
        assert artifact.media_type == ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert artifact.suggested_name.endswith(".xlsx")

        # Verify the XLSX file exists and is valid
        output_path = Path(artifact.staging_path)
        assert output_path.exists()
        assert output_path.stat().st_size > 0

        # Verify XLSX content with openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(output_path)
        sheets = wb.sheetnames
        assert len(sheets) >= 1, "No sheets in workbook"

        # First sheet should have the first table's data
        ws = wb[sheets[0]]
        # Check headers
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(1, 2).value == "Age"
        assert ws.cell(1, 3).value == "City"

        # Check first data row
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == "30"
        assert ws.cell(2, 3).value == "Beijing"

        # Second sheet should have the second table
        if len(sheets) >= 2:
            ws2 = wb[sheets[1]]
            assert ws2.cell(1, 1).value == "Product"
            assert ws2.cell(1, 2).value == "Price"
            assert ws2.cell(1, 3).value == "Qty"

        # Metrics
        assert result.metrics.input_bytes > 0
        assert result.metrics.output_bytes > 0

    @pytest.mark.contract
    def test_single_table_md(self):
        """Single table Markdown produces correct XLSX."""
        md_content = "# Tables\n\n| X | Y |\n|---|---|\n| a | b |\n"
        md_path = write_temp_md(md_content)
        ctx, _workspace = make_context(md_path, target_format="xlsx")

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        from openpyxl import load_workbook

        wb = load_workbook(Path(result.artifacts[0].staging_path))
        ws = wb[wb.sheetnames[0]]
        assert ws.cell(1, 1).value == "X"
        assert ws.cell(1, 2).value == "Y"
        assert ws.cell(2, 1).value == "a"
        assert ws.cell(2, 2).value == "b"

    @pytest.mark.parametrize("image_syntax", ["wiki", "markdown"])
    @pytest.mark.contract
    def test_non_template_xlsx_embeds_images_and_clears_placeholders(
        self,
        tmp_path: Path,
        image_syntax: str,
    ) -> None:
        """Both Markdown image syntaxes become real worksheet drawings."""
        from openpyxl import load_workbook

        image_path = tmp_path / "pixel.png"
        image_path.write_bytes(_png_bytes())
        syntax = "![[pixel.png]]" if image_syntax == "wiki" else "![pixel](pixel.png)"
        source = tmp_path / f"image-{image_syntax}.md"
        source.write_text(
            f"| Kind | Image |\n| --- | --- |\n| {image_syntax} | {syntax} |\n",
            encoding="utf-8",
        )
        context, _workspace = make_context(
            str(source),
            target_format="xlsx",
            config_values={
                "link": {
                    "embed_links": {
                        "wiki_image_mode": "embed",
                        "markdown_image_mode": "embed",
                        "md_file_mode": "embed",
                    }
                }
            },
        )

        result = MdToXlsxConverter().convert(context)

        assert result.success is True, result.error
        workbook = load_workbook(Path(result.artifacts[0].staging_path))
        try:
            worksheet = _active_worksheet(workbook)
            images = vars(worksheet).get("_images")
            assert isinstance(images, list)
            assert len(images) == 1
            assert worksheet["B2"].value in (None, "")
        finally:
            workbook.close()

    @pytest.mark.contract
    def test_no_table_md_produces_empty_workbook(self):
        """MD with no tables produces a workbook with one empty sheet."""
        md_content = "# Just a heading\n\nNo tables here."
        md_path = write_temp_md(md_content)
        ctx, _workspace = make_context(md_path, target_format="xlsx")

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        from openpyxl import load_workbook

        wb = load_workbook(Path(result.artifacts[0].staging_path))
        assert len(wb.sheetnames) == 1

    @pytest.mark.contract
    def test_xlsx_diagnostics(self):
        """Conversion result includes success diagnostics."""
        md_path = write_temp_md("| A |\n|---|\n| 1 |\n")
        ctx, _workspace = make_context(md_path, target_format="xlsx")

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        assert len(result.diagnostics) >= 1
        assert result.diagnostics[0].code == "MD2XLSX-OK"

    @pytest.mark.contract
    def test_artifact_metadata(self):
        """Artifact metadata includes table count."""
        md_path = write_temp_md(SAMPLE_MD_TABLES)
        ctx, _workspace = make_context(md_path, target_format="xlsx")

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        metadata = result.artifacts[0].metadata
        assert metadata.get("source_format") == "markdown"
        assert metadata.get("target_format") == "xlsx"
        assert metadata.get("table_count", 0) >= 2

    @pytest.mark.contract
    def test_template_name_fills_xlsx_template_yaml_column_and_row_placeholders(self, tmp_path: Path):
        """XLSX templates consume YAML fields plus vertical and horizontal table placeholders."""
        from openpyxl import Workbook, load_workbook

        template = tmp_path / "report-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Report"
        ws["A1"] = "{{title}}"
        ws["B1"] = "Report for {{owner}}"
        ws["A3"] = "{{↓Name}}"
        ws["B3"] = "{{↓Age}}"
        ws["C3"] = "{{missing}}"
        ws["A7"] = "{{→Name}}"
        ws["A8"] = "{{→Age}}"
        wb.save(template)
        wb.close()

        md_content = """---
title: Quarterly Report
owner: Finance
---

| Name | Age |
| --- | --- |
| Alice | 30 |
| Bob | 25 |
"""
        md_path = write_temp_md(md_content)
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        output = Path(result.artifacts[0].staging_path)
        out_wb = load_workbook(output)
        out_ws = out_wb["Report"]
        assert out_ws["A1"].value == "Quarterly Report"
        assert out_ws["B1"].value == "Report for Finance"
        assert out_ws["A3"].value == "Alice"
        assert out_ws["A4"].value == "Bob"
        assert out_ws["B3"].value == "30"
        assert out_ws["B4"].value == "25"
        assert out_ws["C3"].value is None
        assert out_ws["A7"].value == "Alice"
        assert out_ws["B7"].value == "Bob"
        assert out_ws["A8"].value == "30"
        assert out_ws["B8"].value == "25"
        assert result.artifacts[0].metadata["template_name"] == str(template)
        assert result.artifacts[0].metadata["yaml_placeholders"] == 2
        assert result.artifacts[0].metadata["column_placeholders"] == 4
        assert result.artifacts[0].metadata["row_placeholders"] == 4
        out_wb.close()

    @pytest.mark.contract
    def test_template_fraction_keeps_two_decimal_display_in_both_fill_directions(self, tmp_path: Path) -> None:
        """Fraction formulas preserve the old workbook's stable two-decimal display."""
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import numbers

        template = tmp_path / "fraction-template.xlsx"
        workbook = Workbook()
        worksheet = _active_worksheet(workbook)
        worksheet["A1"] = "{{↓Ratio}}"
        worksheet["C1"] = "{{→Ratio}}"
        workbook.save(template)
        workbook.close()

        markdown = "| Ratio |\n| --- |\n| 1/2 |\n| 3/4 |\n"
        source = write_temp_md(markdown)
        context, _workspace = make_context(
            source,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        result = MdToXlsxConverter().convert(context)

        assert result.success
        output = load_workbook(Path(result.artifacts[0].staging_path), data_only=False)
        output_sheet = _active_worksheet(output)
        for coordinate, formula in {
            "A1": "=1/2",
            "A2": "=3/4",
            "C1": "=1/2",
            "D1": "=3/4",
        }.items():
            assert output_sheet[coordinate].value == formula
            assert output_sheet[coordinate].number_format == numbers.FORMAT_NUMBER_00
        output.close()

    @pytest.mark.parametrize(
        ("separator", "expected"),
        [(None, "甲、乙"), (" / ", "甲 / 乙"), ("", "甲乙")],
    )
    @pytest.mark.contract
    def test_template_name_joins_yaml_lists_with_request_config_separator(
        self,
        tmp_path: Path,
        separator: str | None,
        expected: str,
    ) -> None:
        from openpyxl import Workbook, load_workbook

        template = tmp_path / "yaml-list-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws["A1"] = "{{authors}}"
        ws["B1"] = "Authors: {{authors}}"
        ws["C1"] = "{{nested}}"
        wb.save(template)
        wb.close()

        md_path = write_temp_md("---\nauthors:\n  - 甲\n  - 乙\nnested:\n  - 甲\n  - [乙, 丙]\n---\n")
        config_values = {} if separator is None else {"conversion": {"md_to_docx": {"list_separator": separator}}}
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": str(template)},
            config_values=config_values,
        )

        result = MdToXlsxConverter().convert(ctx)

        assert result.success
        output = load_workbook(Path(result.artifacts[0].staging_path))
        output_ws = _active_worksheet(output)
        assert output_ws["A1"].value == expected
        assert output_ws["B1"].value == f"Authors: {expected}"
        effective_separator = "、" if separator is None else separator
        assert output_ws["C1"].value == f"甲{effective_separator}['乙', '丙']"
        output.close()

    @pytest.mark.contract
    def test_csv_template_chain_joins_yaml_lists_with_request_config_separator(self, tmp_path: Path) -> None:
        import csv

        template = tmp_path / "yaml-list-csv-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws["A1"] = "{{authors}}"
        wb.save(template)
        wb.close()

        md_path = write_temp_md("---\nauthors:\n  - 甲\n  - 乙\n---\n")
        ctx, _workspace = make_context(
            md_path,
            target_format="csv",
            options={"template_name": str(template)},
            config_values={
                "conversion": {"md_to_docx": {"list_separator": " / "}},
            },
        )

        result = MdToCsvConverter().convert(ctx)

        assert result.success
        with Path(result.artifacts[0].staging_path).open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
        assert rows[0][0] == "甲 / 乙"

    @pytest.mark.contract
    def test_canonical_english_sheet_template_fills_old_project_smoke_cells(self, tmp_path: Path):
        """The bundled XLSX template keeps the old Tk/PySide6 smoke semantics."""
        from openpyxl import load_workbook

        project_root = Path(__file__).resolve().parents[4]
        fixture = _old_system_spreadsheet_smoke_fixture()
        template = project_root / "templates" / "English Sample Sheet Template.xlsx"
        md_path = tmp_path / "canonical-template-smoke.md"
        md_path.write_text(fixture["input_markdown"], encoding="utf-8")
        ctx, _workspace = make_context(
            str(md_path),
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        output = Path(result.artifacts[0].staging_path)
        out_wb = load_workbook(output)
        out_ws = _active_worksheet(out_wb)
        expected_cells = fixture["xlsx_cells"]
        assert {coord: _cell_value(out_ws, coord) for coord in expected_cells} == expected_cells
        out_wb.close()

    @pytest.mark.integration
    def test_markdown_xlsx_old_system_fixture_finalizes_through_runtime(self, tmp_path: Path):
        """Runtime finalizer places the old-system MD->XLSX smoke output."""
        from openpyxl import load_workbook

        fixture = _old_system_spreadsheet_smoke_fixture()
        md_path = tmp_path / "canonical-template-smoke.md"
        md_path.write_text(fixture["input_markdown"], encoding="utf-8")
        output_dir = tmp_path / "final-output"

        result = _execute_markdown_runtime(
            md_path,
            "xlsx",
            output_dir,
            options={"template_name": _bundled_spreadsheet_template_id()},
        )

        assert result.success
        assert len(result.artifacts) == 1
        assert any(diagnostic.code == "FINALIZER_DONE" for diagnostic in result.diagnostics)
        output = Path(result.artifacts[0].staging_path)
        assert output.parent == output_dir
        assert output.suffix == ".xlsx"
        assert output.exists()
        assert not output.is_relative_to(tmp_path / "workspace")

        out_wb = load_workbook(output)
        out_ws = _active_worksheet(out_wb)
        expected_cells = fixture["xlsx_cells"]
        assert {coord: _cell_value(out_ws, coord) for coord in expected_cells} == expected_cells
        out_wb.close()

    @pytest.mark.contract
    def test_template_name_preserves_old_inline_table_cell_text_semantics(self, tmp_path: Path):
        """Old PySide6 workbook guards keep links and code-span pipes as cell text."""
        from openpyxl import Workbook, load_workbook

        case = _old_system_spreadsheet_release_gate_fixture()["cases"]["inline_source_text"]
        template = tmp_path / "inline-cell-template.xlsx"
        wb = Workbook()
        ws = _active_worksheet(wb)
        ws.title = "Sheet1"
        ws["A1"] = "{{↓Name}}"
        ws["B1"] = "{{↓Age}}"
        wb.save(template)
        wb.close()

        md_path = write_temp_md(case["input_markdown"])
        ctx, _workspace = make_context(
            md_path,
            target_format="xlsx",
            options={"template_name": str(template)},
        )

        converter = MdToXlsxConverter()
        result = converter.convert(ctx)

        assert result.success
        out_wb = load_workbook(Path(result.artifacts[0].staging_path))
        out_ws = _active_worksheet(out_wb)
        assert {coord: _cell_value(out_ws, coord) for coord in case["expected_cells"]} == case["expected_cells"]
        out_wb.close()
