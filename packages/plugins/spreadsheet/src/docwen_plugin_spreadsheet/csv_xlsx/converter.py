"""CSV/TSV ↔ XLSX interconversion converters.

Implements:
- ROUTE-CSV-XLSX-001: csv → xlsx
- ROUTE-XLSX-CSV-001: xlsx → csv
- ROUTE-TSV-XLSX-001: tsv → xlsx
- ROUTE-XLSX-TSV-001: xlsx → tsv

All converters are pure Python (openpyxl + csv module).
"""

from __future__ import annotations

import csv
import os
import re
import uuid
from collections.abc import Callable
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from docwen_core.protocols.execution_context import ConverterContext


def _load_admitted_xlsx(file_path: str, *, data_only: bool = True) -> Any:
    """Load admitted XLSX content without consulting its user-facing suffix.

    ``openpyxl.load_workbook()`` rejects misleading filenames before it reads
    the OOXML package when passed a path.  Core admission already established
    the concrete source format, so plugin parsers consume the bytes directly.
    """
    import openpyxl

    with Path(file_path).open("rb") as workbook_stream:
        return openpyxl.load_workbook(workbook_stream, data_only=data_only)


def _maybe_number(v: str) -> Any:
    """Try to convert a string to int or float, preserving leading zeros."""
    s = (v or "").strip()
    if s == "":
        return ""
    # Integer
    if re.fullmatch(r"-?[0-9]+", s):
        if len(s) > 1 and s.startswith("0"):
            return v
        if len(s) > 2 and s.startswith("-0"):
            return v
        try:
            return int(s)
        except Exception:
            return v
    # Float
    if re.fullmatch(r"-?[0-9]+\.[0-9]+", s):
        try:
            return float(s)
        except Exception:
            return v
    return v


def _detect_csv_encoding(file_path: str) -> str:
    """Detect the encoding of a CSV/TSV file by trying a sample of the file.

    Tries encodings in the same order as ``_read_csv_flexible()``
    (utf-8-sig, utf-8, gbk, utf-16) to ensure consistency between
    CSV/TSV→XLSX and CSV→MD routes.  If all attempts fail, falls back
    to utf-8-sig.
    """
    candidates = ("utf-8-sig", "utf-8", "gbk", "utf-16")

    try:
        with open(file_path, "rb") as f:
            sample_bytes = f.read(65536)
    except Exception:
        return "utf-8-sig"

    for encoding in candidates:
        try:
            sample_bytes.decode(encoding)
            return encoding
        except (UnicodeDecodeError, UnicodeError):
            continue

    return "utf-8-sig"


def _build_delimited_workbook(
    input_path: str,
    *,
    sep: str,
    cancel_check: Callable[[], None] | None = None,
) -> tuple[Any, int]:
    """Build the canonical one-sheet workbook used by delimited input routes."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Sheet1"

    row_count = 0
    detected_enc = _detect_csv_encoding(input_path)
    try:
        with open(input_path, encoding=detected_enc, newline="") as f:
            reader = csv.reader(f, delimiter=sep)
            for r_idx, row in enumerate(reader, 1):
                if cancel_check is not None and r_idx % 1000 == 0:
                    cancel_check()
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=_maybe_number(value))
                row_count = r_idx
    except Exception:
        wb.close()
        raise
    return wb, row_count


class CsvToXlsxConverter:
    """Convert a CSV file to XLSX (ROUTE-CSV-XLSX-001)."""

    def convert(self, context: ConverterContext) -> Any:
        from docwen_core.models.artifact import ArtifactManifest
        from docwen_core.models.result import (
            ConversionDiagnostic,
            ConversionErrorInfo,
            ConversionMetrics,
            ConversionResult,
        )

        task_id = context.request.request_id
        input_path = context.workspace.input_path

        context.cancellation.check()
        context.progress.report_progress(0.0, "Starting CSV → XLSX conversion")
        context.logger.info(f"CSV→XLSX: reading {input_path}")

        # ── Phase 1: Parse CSV ───────────────────────────────────────
        try:
            wb, row_count = _build_delimited_workbook(
                input_path,
                sep=",",
                cancel_check=context.cancellation.check,
            )

            context.progress.report_progress(50.0, "Writing XLSX...")
        except Exception as exc:
            context.logger.error(f"CSV→XLSX parse failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=str(exc),
                    diagnostic_code="CSV2XLSX-PARSE-ERROR",
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"Failed to parse CSV: {exc}",
                        code="CSV2XLSX-PARSE-ERROR",
                    ),
                ],
            )

        # ── Phase 2: Write XLSX ───────────────────────────────────────
        output_path = context.workspace.create_artifact_path("primary", ".xlsx")
        try:
            wb.save(output_path)
        except Exception as exc:
            context.logger.error(f"CSV→XLSX write failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=f"Failed to write XLSX file: {exc}",
                    diagnostic_code="CSV2XLSX-WRITE-ERROR",
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"File write error at {output_path}: {exc}",
                        code="CSV2XLSX-WRITE-ERROR",
                    ),
                ],
            )
        finally:
            wb.close()

        input_basename = os.path.basename(input_path)
        suggested_name = input_basename.rsplit(".", 1)[0] + ".xlsx"

        artifact = ArtifactManifest(
            artifact_id=str(uuid.uuid4()),
            kind="primary",
            staging_path=output_path,
            suggested_name=suggested_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata={"row_count": row_count, "sheet_count": 1},
            is_primary=True,
        )
        context.workspace.add_artifact(artifact)

        context.progress.report_artifact_ready(artifact.artifact_id, suggested_name)
        context.progress.report_progress(100.0, "CSV → XLSX complete")
        context.logger.info(f"CSV→XLSX complete: {row_count} rows")

        return ConversionResult(
            task_id=task_id,
            success=True,
            artifacts=[artifact],
            diagnostics=[
                ConversionDiagnostic(
                    level="info",
                    message=f"Converted CSV to XLSX: {row_count} rows",
                    code="CSV2XLSX-OK",
                ),
            ],
            error=None,
            metrics=ConversionMetrics(
                duration_ms=0.0,
                input_bytes=os.path.getsize(input_path) if os.path.isfile(input_path) else 0,
                output_bytes=os.path.getsize(output_path) if os.path.isfile(output_path) else 0,
                extra={"row_count": row_count},
            ),
        )


class XlsxToCsvConverter:
    """Convert an XLSX file to CSV files — one per sheet (ROUTE-XLSX-CSV-001)."""

    def convert(
        self,
        context: ConverterContext,
        *,
        suggested_stem: str | None = None,
        metadata_base: dict[str, Any] | None = None,
        progress_start: float = 0.0,
    ) -> Any:
        from docwen_core.models.artifact import ArtifactManifest
        from docwen_core.models.result import (
            ConversionDiagnostic,
            ConversionErrorInfo,
            ConversionMetrics,
            ConversionResult,
        )

        task_id = context.request.request_id
        input_path = context.workspace.input_path
        input_stem = suggested_stem or Path(input_path).stem

        context.cancellation.check()
        context.progress.report_progress(progress_start, "Starting XLSX → CSV conversion")
        context.logger.info(f"XLSX→CSV: reading {input_path}")

        # ── Phase 1: Parse XLSX ───────────────────────────────────────
        wb = None
        try:
            wb = _load_admitted_xlsx(input_path, data_only=True)
        except Exception as exc:
            context.logger.error(f"XLSX→CSV parse failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=str(exc),
                    diagnostic_code="XLSX2CSV-PARSE-ERROR",
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"Failed to parse XLSX workbook: {exc}",
                        code="XLSX2CSV-PARSE-ERROR",
                    ),
                ],
            )

        # ── Phase 2: Write CSV per sheet ───────────────────────────────
        artifacts: list[ArtifactManifest] = []
        total_sheets = len(wb.sheetnames)
        total_rows = 0

        try:
            for idx, sheet_name in enumerate(wb.sheetnames):
                context.cancellation.check()
                progress = progress_start + (100.0 - progress_start) * (idx / max(total_sheets, 1))
                context.progress.report_progress(progress, f"Converting sheet: {sheet_name}")

                ws = wb[sheet_name]
                clean_name = sheet_name.replace(" ", "_")

                output_path = context.workspace.create_artifact_path("primary" if idx == 0 else "auxiliary", ".csv")

                row_count = 0
                with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(["" if v is None else v for v in row])
                        row_count += 1

                total_rows += row_count

                artifact = ArtifactManifest(
                    artifact_id=str(uuid.uuid4()),
                    kind="primary" if idx == 0 else "auxiliary",
                    staging_path=output_path,
                    suggested_name=f"{input_stem}_{clean_name}.csv",
                    media_type="text/csv",
                    metadata={
                        **(metadata_base or {}),
                        "sheet_name": sheet_name,
                        "row_count": row_count,
                        "sheet_index": idx,
                        "sheet_count": total_sheets,
                    },
                    is_primary=(idx == 0),
                )
                context.workspace.add_artifact(artifact)
                artifacts.append(artifact)

        except Exception as exc:
            context.logger.error(f"XLSX→CSV write failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=str(exc),
                    diagnostic_code="XLSX2CSV-WRITE-ERROR",
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"Failed to write CSV output: {exc}",
                        code="XLSX2CSV-WRITE-ERROR",
                    ),
                ],
            )
        finally:
            if wb is not None:
                wb.close()

        context.progress.report_progress(100.0, "XLSX → CSV complete")
        context.logger.info(f"XLSX→CSV complete: {len(artifacts)} sheets, {total_rows} rows")

        return ConversionResult(
            task_id=task_id,
            success=True,
            artifacts=artifacts,
            diagnostics=[
                ConversionDiagnostic(
                    level="info",
                    message=f"Converted XLSX to CSV: {len(artifacts)} sheets",
                    code="XLSX2CSV-OK",
                ),
            ],
            error=None,
            metrics=ConversionMetrics(
                duration_ms=0.0,
                input_bytes=os.path.getsize(input_path) if os.path.isfile(input_path) else 0,
                output_bytes=sum(os.path.getsize(a.staging_path) for a in artifacts if os.path.isfile(a.staging_path)),
                extra={"sheet_count": len(artifacts), "total_rows": total_rows},
            ),
        )


class TsvToXlsxConverter:
    """Convert a TSV file to XLSX (ROUTE-TSV-XLSX-001)."""

    def convert(self, context: ConverterContext) -> Any:
        from docwen_core.models.artifact import ArtifactManifest
        from docwen_core.models.result import (
            ConversionDiagnostic,
            ConversionErrorInfo,
            ConversionMetrics,
            ConversionResult,
        )

        task_id = context.request.request_id
        input_path = context.workspace.input_path

        context.cancellation.check()
        context.progress.report_progress(0.0, "Starting TSV → XLSX conversion")
        context.logger.info(f"TSV→XLSX: reading {input_path}")

        # ── Phase 1: Parse TSV ───────────────────────────────────────
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet()
            ws.title = "Sheet1"

            row_count = 0
            detected_enc = _detect_csv_encoding(input_path)
            with open(input_path, encoding=detected_enc, newline="") as f:
                reader = csv.reader(f, delimiter="\t")
                for r_idx, row in enumerate(reader, 1):
                    # Allow cancellation mid-conversion for large files
                    if r_idx % 1000 == 0:
                        context.cancellation.check()
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=_maybe_number(value))
                    row_count = r_idx

            context.progress.report_progress(50.0, "Writing XLSX...")
        except Exception as exc:
            context.logger.error(f"TSV→XLSX parse failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=str(exc),
                    diagnostic_code="TSV2XLSX-PARSE-ERROR",
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"Failed to parse TSV: {exc}",
                        code="TSV2XLSX-PARSE-ERROR",
                    ),
                ],
            )

        # ── Phase 2: Write XLSX ───────────────────────────────────────
        output_path = context.workspace.create_artifact_path("primary", ".xlsx")
        try:
            wb.save(output_path)
        except Exception as exc:
            context.logger.error(f"TSV→XLSX write failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=f"Failed to write XLSX file: {exc}",
                    diagnostic_code="TSV2XLSX-WRITE-ERROR",
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"File write error at {output_path}: {exc}",
                        code="TSV2XLSX-WRITE-ERROR",
                    ),
                ],
            )
        finally:
            wb.close()

        input_basename = os.path.basename(input_path)
        suggested_name = input_basename.rsplit(".", 1)[0] + ".xlsx"

        artifact = ArtifactManifest(
            artifact_id=str(uuid.uuid4()),
            kind="primary",
            staging_path=output_path,
            suggested_name=suggested_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata={"row_count": row_count, "sheet_count": 1},
            is_primary=True,
        )
        context.workspace.add_artifact(artifact)

        context.progress.report_artifact_ready(artifact.artifact_id, suggested_name)
        context.progress.report_progress(100.0, "TSV → XLSX complete")

        return ConversionResult(
            task_id=task_id,
            success=True,
            artifacts=[artifact],
            diagnostics=[
                ConversionDiagnostic(
                    level="info",
                    message=f"Converted TSV to XLSX: {row_count} rows",
                    code="TSV2XLSX-OK",
                ),
            ],
            error=None,
            metrics=ConversionMetrics(
                duration_ms=0.0,
                input_bytes=os.path.getsize(input_path) if os.path.isfile(input_path) else 0,
                output_bytes=os.path.getsize(output_path) if os.path.isfile(output_path) else 0,
                extra={"row_count": row_count},
            ),
        )


class XlsxToTsvConverter:
    """Convert an XLSX file to TSV files — one per sheet (ROUTE-XLSX-TSV-001)."""

    def convert(self, context: ConverterContext) -> Any:
        from docwen_core.models.artifact import ArtifactManifest
        from docwen_core.models.result import (
            ConversionDiagnostic,
            ConversionErrorInfo,
            ConversionMetrics,
            ConversionResult,
        )

        task_id = context.request.request_id
        input_path = context.workspace.input_path
        input_stem = Path(input_path).stem

        context.cancellation.check()
        context.progress.report_progress(0.0, "Starting XLSX → TSV conversion")
        context.logger.info(f"XLSX→TSV: reading {input_path}")

        # ── Phase 1: Parse XLSX ───────────────────────────────────────
        wb = None
        try:
            wb = _load_admitted_xlsx(input_path, data_only=True)
        except Exception as exc:
            context.logger.error(f"XLSX→TSV parse failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=str(exc),
                    diagnostic_code="XLSX2TSV-PARSE-ERROR",
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"Failed to parse XLSX workbook: {exc}",
                        code="XLSX2TSV-PARSE-ERROR",
                    ),
                ],
            )

        # ── Phase 2: Write TSV per sheet ───────────────────────────────
        artifacts: list[ArtifactManifest] = []
        total_sheets = len(wb.sheetnames)
        total_rows = 0

        try:
            for idx, sheet_name in enumerate(wb.sheetnames):
                context.cancellation.check()
                progress = 50.0 * (idx / max(total_sheets, 1))
                context.progress.report_progress(progress, f"Converting sheet: {sheet_name}")

                ws = wb[sheet_name]
                clean_name = sheet_name.replace(" ", "_")

                output_path = context.workspace.create_artifact_path("primary" if idx == 0 else "auxiliary", ".tsv")

                row_count = 0
                with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f, delimiter="\t")
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(["" if v is None else v for v in row])
                        row_count += 1

                total_rows += row_count

                artifact = ArtifactManifest(
                    artifact_id=str(uuid.uuid4()),
                    kind="primary" if idx == 0 else "auxiliary",
                    staging_path=output_path,
                    suggested_name=f"{input_stem}_{clean_name}.tsv",
                    media_type="text/tab-separated-values",
                    metadata={
                        "sheet_name": sheet_name,
                        "row_count": row_count,
                        "sheet_index": idx,
                    },
                    is_primary=(idx == 0),
                )
                context.workspace.add_artifact(artifact)
                artifacts.append(artifact)

        except Exception as exc:
            context.logger.error(f"XLSX→TSV write failed: {exc}")
            return ConversionResult(
                task_id=task_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=str(exc),
                    diagnostic_code="XLSX2TSV-WRITE-ERROR",
                ),
                diagnostics=[
                    ConversionDiagnostic(
                        level="error",
                        message=f"Failed to write TSV output: {exc}",
                        code="XLSX2TSV-WRITE-ERROR",
                    ),
                ],
            )
        finally:
            if wb is not None:
                wb.close()

        context.progress.report_progress(100.0, "XLSX → TSV complete")
        context.logger.info(f"XLSX→TSV complete: {len(artifacts)} sheets, {total_rows} rows")

        return ConversionResult(
            task_id=task_id,
            success=True,
            artifacts=artifacts,
            diagnostics=[
                ConversionDiagnostic(
                    level="info",
                    message=f"Converted XLSX to TSV: {len(artifacts)} sheets",
                    code="XLSX2TSV-OK",
                ),
            ],
            error=None,
            metrics=ConversionMetrics(
                duration_ms=0.0,
                input_bytes=os.path.getsize(input_path) if os.path.isfile(input_path) else 0,
                output_bytes=sum(os.path.getsize(a.staging_path) for a in artifacts if os.path.isfile(a.staging_path)),
                extra={"sheet_count": len(artifacts), "total_rows": total_rows},
            ),
        )
