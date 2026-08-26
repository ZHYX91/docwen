"""Focused tests split from test_xlsx_to_md_golden.py."""

from __future__ import annotations

from ._xlsx_to_md_golden_support import (
    Any,
    Generator,
    Path,
    _build_fake_context,
    _deliverable_artifacts,
    _document_node_root,
    _load_xlsx_to_md_old_system_fixture,
    pytest,
    tempfile,
)

pytestmark = pytest.mark.golden


@pytest.mark.contract
class TestSpreadsheetToMdDirect:
    """Test SpreadsheetToMarkdownConverter directly with fake context."""

    def test_convert_cancellation_checked(self, sample_xlsx_path: Path) -> None:
        """Conversion should respect cancellation."""
        from docwen_core.errors import CancellationRequested
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging)
            # Cancel before starting
            context.cancellation.cancel("test cancel")

            with pytest.raises(CancellationRequested):
                SpreadsheetToMarkdownConverter().convert(context)


@pytest.mark.integration
class TestSpreadsheetToMdPipeline:
    """Test spreadsheet→MD through the full runtime pipeline."""

    @pytest.fixture
    def pipeline(self) -> Generator[dict[str, Any], None, None]:
        """Build a full runtime pipeline for spreadsheet→md conversion."""
        from docwen_plugin_spreadsheet.plugin import SpreadsheetPlugin
        from docwen_runtime.engine.route_resolver import RouteResolver
        from docwen_runtime.engine.task_manager import TaskManager
        from docwen_runtime.output.finalizer import OutputFinalizer
        from docwen_runtime.plugin_registry.registry import PluginRegistry
        from docwen_runtime.workspace.manager import WorkspaceManager

        plugin = SpreadsheetPlugin()
        plugin_registry = PluginRegistry()
        plugin_registry.register(plugin)

        route_resolver = RouteResolver(plugin_registry)

        ws_root = tempfile.mkdtemp(prefix="docwen_test_ws_")
        ws_manager = WorkspaceManager(ws_root)
        finalizer = OutputFinalizer()
        task_manager = TaskManager(
            plugin_registry=plugin_registry,
            route_resolver=route_resolver,
            workspace_manager=ws_manager,
            output_finalizer=finalizer,
        )

        yield {
            "task_manager": task_manager,
            "workspace_manager": ws_manager,
            "ws_root": ws_root,
        }

        # Cleanup
        import shutil

        shutil.rmtree(ws_root, ignore_errors=True)

    def test_pipeline_xlsx_to_md(self, pipeline: dict[str, Any], sample_xlsx_path: Path) -> None:
        """Full pipeline: XLSX→MD through runtime."""
        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        file_ref = FileRef(
            path=str(sample_xlsx_path),
            format="xlsx",
            category="spreadsheet",
        )
        request = ConversionRequest(
            request_id="pipe-001",
            input_refs=[file_ref],
            target_format="md",
            output_policy=OutputPolicy(),
        )

        result = pipeline["task_manager"].execute_single(request)
        assert result.success is True
        assert len(result.artifacts) >= 1
        assert result.artifacts[0].kind == "primary"

    def test_pipeline_xlsx_old_system_fixture_finalizes_primary_markdown(
        self,
        pipeline: dict[str, Any],
        sample_xlsx_path: Path,
        tmp_path: Path,
    ) -> None:
        """Current XLSX→MD old-system fixture should finalize primary Markdown into output_dir."""
        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        fixture = _load_xlsx_to_md_old_system_fixture()
        output_dir = tmp_path / "output_xlsx_old_system_fixture"
        output_dir.mkdir()
        request = ConversionRequest(
            request_id="pipe-xlsx-old-system-finalize",
            input_refs=[
                FileRef(
                    path=str(sample_xlsx_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=sample_xlsx_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={
                "to_md_keep_images": False,
                "yaml_key_labels": {"title": "Titel"},
            },
        )

        result = pipeline["task_manager"].execute_single(request)

        assert result.success is True
        assert len(_deliverable_artifacts(result)) == 1
        artifact = _deliverable_artifacts(result)[0]
        current = fixture["projects"]["docwen-current"]
        assert artifact.media_type == current["artifact_media_type"]
        assert artifact.metadata["source_suggested_name"] == current["suggested_name"]
        for key, value in current["metadata"].items():
            assert artifact.metadata[key] == value
        assert any(d.code == "FINALIZER_DONE" for d in result.diagnostics)

        artifact_path = Path(artifact.staging_path)
        node_root = _document_node_root(artifact_path, output_dir)
        assert artifact_path.name == f"{node_root.name}.md"
        assert artifact_path.exists()

        content = artifact_path.read_text(encoding="utf-8")
        assert str(Path(pipeline["ws_root"])) not in content
        assert f"Titel: {sample_xlsx_path.stem}" in content
        assert f"title: {sample_xlsx_path.stem}" not in content
        for expected in fixture["expected_markdown_semantics"]["contains"]:
            assert expected in content
        for heading in fixture["expected_markdown_semantics"]["sheet_headings"]:
            assert heading in content

    def test_pipeline_xlsx_images_are_finalized_as_relative_markdown_artifacts(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
    ) -> None:
        """XLSX images must survive runtime finalization, not only staging."""
        import io

        import openpyxl
        import openpyxl.drawing.image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        image_bytes = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
            b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc```\x00\x00\x00\x04\x00\x01\xf6\x178U"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Data"
        sheet.cell(row=1, column=1, value="Header")
        sheet.cell(row=2, column=1, value="Value")
        sheet.add_image(openpyxl.drawing.image.Image(io.BytesIO(image_bytes)), "B2")

        input_path = tmp_path / "xlsx_image_artifact_probe.xlsx"
        workbook.save(input_path)
        workbook.close()

        output_dir = tmp_path / "output_xlsx_images"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-image-finalize",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={"to_md_keep_images": True, "image_link_style": "markdown_embed"},
        )

        result = pipeline["task_manager"].execute_single(request)

        assert result.success is True
        markdown_artifact = next(artifact for artifact in result.artifacts if artifact.is_primary)
        image_artifacts = [artifact for artifact in result.artifacts if artifact.kind == "image"]
        assert markdown_artifact.metadata["image_count"] == 1
        assert len(image_artifacts) == 1
        node_root = _document_node_root(Path(markdown_artifact.staging_path), output_dir)
        assert _document_node_root(Path(image_artifacts[0].staging_path), output_dir) == node_root
        assert Path(image_artifacts[0].staging_path).is_file()

        content = Path(markdown_artifact.staging_path).read_text(encoding="utf-8")
        assert image_artifacts[0].suggested_name in content
        assert str(output_dir) not in content
        assert "docwen_test_ws_" not in content

    def test_pipeline_xlsx_embedded_image_matches_old_system_projection(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
    ) -> None:
        """Focused XLSX embedded-image projection preserves old-system reachability."""
        import io

        import openpyxl
        import openpyxl.drawing.image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["image_ocr_scope"]["embedded_image_probe"]
        image_bytes = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
            b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc```\x00\x00\x00\x04\x00\x01\xf6\x178U"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = scope["input"]["worksheet"]
        sheet.cell(row=1, column=1, value="Label")
        sheet.cell(row=1, column=2, value="Value")
        sheet.cell(row=2, column=1, value="Before image")
        sheet.cell(row=2, column=2, value=42)
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(image_bytes)),
            scope["input"]["image_anchor"],
        )
        input_path = tmp_path / scope["input"]["name"]
        workbook.save(input_path)
        workbook.close()

        output_dir = tmp_path / "embedded_image_output"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-embedded-image-projection",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={"to_md_keep_images": True, "image_link_style": "markdown_embed"},
        )

        result = pipeline["task_manager"].execute_single(request)

        expected = scope["current_projection"]
        assert result.success is True
        assert [diagnostic.code for diagnostic in result.diagnostics] == expected["diagnostic_codes"]
        assert len(_deliverable_artifacts(result)) == expected["artifact_count"]
        markdown_artifact = next(artifact for artifact in result.artifacts if artifact.is_primary)
        image_artifacts = [artifact for artifact in result.artifacts if artifact.kind == "image"]
        assert len(image_artifacts) == expected["image_artifact_count"]
        assert markdown_artifact.metadata["image_count"] == expected["markdown_image_count"]
        node_root = _document_node_root(Path(markdown_artifact.staging_path), output_dir)
        assert _document_node_root(Path(image_artifacts[0].staging_path), output_dir) == node_root
        assert Path(image_artifacts[0].staging_path).is_file()
        assert image_artifacts[0].suggested_name == expected["image_suggested_name"]

        content = Path(markdown_artifact.staging_path).read_text(encoding="utf-8")
        assert str(Path(pipeline["ws_root"])) not in content
        assert str(output_dir) not in content
        for token in expected["required_markdown_tokens"]:
            assert token in content

    def test_pipeline_xlsx_multi_image_matches_old_system_projection(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
    ) -> None:
        """Focused XLSX multi-image projection preserves old-system order."""
        import io

        import openpyxl
        import openpyxl.drawing.image
        from PIL import Image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["image_ocr_scope"]["multi_image_probe"]

        def png_bytes(rgb: tuple[int, int, int]) -> bytes:
            buffer = io.BytesIO()
            Image.new("RGB", (1, 1), rgb).save(buffer, format="PNG")
            return buffer.getvalue()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = scope["input"]["worksheet"]
        sheet.cell(row=1, column=1, value="Label")
        sheet.cell(row=1, column=2, value="Value")
        sheet.cell(row=2, column=1, value="Before image")
        sheet.cell(row=2, column=2, value=42)
        sheet.cell(row=4, column=1, value="Second image")
        sheet.cell(row=4, column=2, value=84)
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(png_bytes((255, 0, 0)))),
            scope["input"]["image_anchors"][0],
        )
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(png_bytes((0, 0, 255)))),
            scope["input"]["image_anchors"][1],
        )
        input_path = tmp_path / scope["input"]["name"]
        workbook.save(input_path)
        workbook.close()

        output_dir = tmp_path / "multi_image_output"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-multi-image-projection",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={
                "to_md_keep_images": True,
                "to_md_enable_ocr": False,
                "image_mode": "file",
                "image_link_style": "markdown_embed",
            },
        )

        result = pipeline["task_manager"].execute_single(request)

        expected = scope["current_projection"]
        assert result.success is True
        assert [diagnostic.code for diagnostic in result.diagnostics] == expected["diagnostic_codes"]
        assert len(_deliverable_artifacts(result)) == expected["artifact_count"]
        markdown_artifact = next(artifact for artifact in result.artifacts if artifact.is_primary)
        image_artifacts = [artifact for artifact in result.artifacts if artifact.kind == "image"]
        assert len(image_artifacts) == expected["image_artifact_count"]
        assert [artifact.suggested_name for artifact in image_artifacts] == expected["image_suggested_names"]
        for key, value in expected["artifact_metadata"].items():
            assert markdown_artifact.metadata[key] == value
        node_root = _document_node_root(Path(markdown_artifact.staging_path), output_dir)
        for artifact in image_artifacts:
            assert _document_node_root(Path(artifact.staging_path), output_dir) == node_root
            assert Path(artifact.staging_path).is_file()

        content = Path(markdown_artifact.staging_path).read_text(encoding="utf-8")
        assert str(Path(pipeline["ws_root"])) not in content
        assert str(output_dir) not in content
        assert content.count("![") == expected["markdown_image_count"]
        first_index = content.index("ImageSheet_image1.png")
        second_index = content.index("ImageSheet_image2.png")
        assert first_index < second_index
        for token in expected["required_markdown_tokens"]:
            assert token in content

    def test_pipeline_xlsx_multi_sheet_image_matches_old_system_projection(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
    ) -> None:
        """Focused XLSX multi-sheet image projection preserves old-system order."""
        import io

        import openpyxl
        import openpyxl.drawing.image
        from PIL import Image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["image_ocr_scope"]["multi_sheet_image_probe"]

        def png_bytes(rgb: tuple[int, int, int]) -> bytes:
            buffer = io.BytesIO()
            Image.new("RGB", (1, 1), rgb).save(buffer, format="PNG")
            return buffer.getvalue()

        workbook = openpyxl.Workbook()
        first_sheet = workbook.active
        assert first_sheet is not None
        first_sheet.title = scope["input"]["worksheets"][0]
        first_sheet.cell(row=1, column=1, value="Label")
        first_sheet.cell(row=1, column=2, value="Value")
        first_sheet.cell(row=2, column=1, value="First image")
        first_sheet.cell(row=2, column=2, value=11)
        first_sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(png_bytes((255, 0, 0)))),
            scope["input"]["image_anchors"]["FirstSheet"],
        )

        second_sheet = workbook.create_sheet(scope["input"]["worksheets"][1])
        second_sheet.cell(row=1, column=1, value="Label")
        second_sheet.cell(row=1, column=2, value="Value")
        second_sheet.cell(row=2, column=1, value="Second image")
        second_sheet.cell(row=2, column=2, value=22)
        second_sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(png_bytes((0, 0, 255)))),
            scope["input"]["image_anchors"]["SecondSheet"],
        )

        input_path = tmp_path / scope["input"]["name"]
        workbook.save(input_path)
        workbook.close()

        output_dir = tmp_path / "multi_sheet_image_output"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-multi-sheet-image-projection",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={
                "to_md_keep_images": True,
                "to_md_enable_ocr": False,
                "image_mode": "file",
                "image_link_style": "markdown_embed",
            },
        )

        result = pipeline["task_manager"].execute_single(request)

        expected = scope["current_projection"]
        assert result.success is True
        assert [diagnostic.code for diagnostic in result.diagnostics] == expected["diagnostic_codes"]
        assert len(_deliverable_artifacts(result)) == expected["artifact_count"]
        markdown_artifact = next(artifact for artifact in result.artifacts if artifact.is_primary)
        image_artifacts = [artifact for artifact in result.artifacts if artifact.kind == "image"]
        assert len(image_artifacts) == expected["image_artifact_count"]
        assert [artifact.suggested_name for artifact in image_artifacts] == expected["image_suggested_names"]
        for key, value in expected["artifact_metadata"].items():
            assert markdown_artifact.metadata[key] == value
        node_root = _document_node_root(Path(markdown_artifact.staging_path), output_dir)
        for artifact in image_artifacts:
            assert _document_node_root(Path(artifact.staging_path), output_dir) == node_root
            assert Path(artifact.staging_path).is_file()

        content = Path(markdown_artifact.staging_path).read_text(encoding="utf-8")
        assert str(Path(pipeline["ws_root"])) not in content
        assert str(output_dir) not in content
        assert content.count("![") == expected["markdown_image_count"]
        first_index = content.index("FirstSheet_image1.png")
        second_index = content.index("SecondSheet_image2.png")
        assert first_index < second_index
        assert expected["images_in_order"] is True
        for token in expected["required_markdown_tokens"]:
            assert token in content
