"""Focused tests split from test_xlsx_to_md_golden.py."""

from __future__ import annotations

from ._xlsx_to_md_golden_support import (
    Path,
    _build_fake_context,
    _load_xlsx_to_md_old_system_fixture,
    os,
    pytest,
    tempfile,
)

pytestmark = pytest.mark.golden


@pytest.mark.contract
class TestSpreadsheetToMdDirect:
    """Test SpreadsheetToMarkdownConverter directly with fake context."""

    def test_convert_xlsx_basic(self, sample_xlsx_path: Path) -> None:
        """Direct conversion of a known XLSX should produce valid Markdown."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging)
            converter = SpreadsheetToMarkdownConverter()
            result = converter.convert(context)

            assert result.success is True
            assert len(result.artifacts) >= 1
            assert result.artifacts[0].kind == "primary"
            assert result.artifacts[0].media_type == "text/markdown"

            # Read the output
            output_path = result.artifacts[0].staging_path
            assert os.path.isfile(output_path)
            content = Path(output_path).read_text(encoding="utf-8")

            # Check content
            assert "Sales" in content
            assert "Summary" in content
            assert "Product" in content
            assert "Alpha" in content
            assert "Beta" in content
            assert "Gamma" in content

            # Architecture guard: no legacy two-stage marker pipeline
            assert "{{IMAGE:" not in content, "XLSX→MD output must not contain legacy {{IMAGE:...}} markers"

    @pytest.mark.parametrize(
        ("fixture_name", "admitted_format", "misleading_suffix", "expected_text"),
        [
            ("sample_csv_path", "csv", ".xlsx", "Alice"),
            ("sample_tsv_path", "tsv", ".csv", "95"),
            ("sample_xlsx_path", "xlsx", ".csv", "Sales"),
        ],
    )
    def test_parser_uses_admitted_format_when_suffix_is_wrong(
        self,
        request: pytest.FixtureRequest,
        tmp_path: Path,
        fixture_name: str,
        admitted_format: str,
        misleading_suffix: str,
        expected_text: str,
    ) -> None:
        """CSV, TSV, and XLSX parser selection is content-admission owned."""
        from docwen_plugin_spreadsheet.to_markdown.converter import SpreadsheetToMarkdownConverter

        source_path = request.getfixturevalue(fixture_name)
        misleading_path = tmp_path / f"admitted-{admitted_format}{misleading_suffix}"
        misleading_path.write_bytes(Path(source_path).read_bytes())

        staging = tmp_path / f"staging-{admitted_format}"
        staging.mkdir()
        context = _build_fake_context(
            str(misleading_path),
            str(staging),
            source_format=admitted_format,
        )

        result = SpreadsheetToMarkdownConverter().convert(context)

        assert result.success is True
        markdown = Path(result.artifacts[0].staging_path).read_text(encoding="utf-8")
        assert expected_text in markdown
        assert result.metrics.extra["sheets"] >= 1

    def test_xlsx_to_md_matches_old_system_semantic_fixture(self, sample_xlsx_path: Path) -> None:
        """GOLDEN-003: Current output preserves old-system XLSX→MD semantics."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        fixture = _load_xlsx_to_md_old_system_fixture()
        assert fixture["golden_id"] == "GOLDEN-003"

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(
                str(sample_xlsx_path),
                staging,
                options={"to_md_keep_images": False},
            )
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success is True
            assert len(result.artifacts) == 1

            artifact = result.artifacts[0]
            content = Path(artifact.staging_path).read_text(encoding="utf-8")

            for expected in fixture["expected_markdown_semantics"]["contains"]:
                assert expected in content
            for heading in fixture["expected_markdown_semantics"]["sheet_headings"]:
                assert heading in content

            current = fixture["projects"]["docwen-current"]
            assert artifact.media_type == current["artifact_media_type"]
            assert artifact.suggested_name == current["suggested_name"]
            for key, value in current["metadata"].items():
                assert artifact.metadata[key] == value

            for project_name in ("docwen-ref-tk", "docwen-ref-pyside6", "docwen-current"):
                project = fixture["projects"][project_name]
                assert project["success"] is True
                assert project["markdown_contains_all_expected_semantics"] is True

    def test_convert_xlsx_has_sections(self, sample_xlsx_path: Path) -> None:
        """Each sheet should appear as a level-1 heading."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success
            content = Path(result.artifacts[0].staging_path).read_text("utf-8")

            assert "# Sales" in content
            assert "# Summary" in content

    def test_convert_xlsx_has_yaml_frontmatter(self, sample_xlsx_path: Path) -> None:
        """Output should contain YAML frontmatter with title."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success
            content = Path(result.artifacts[0].staging_path).read_text("utf-8")

            assert content.startswith("---")
            assert "title:" in content

    def test_convert_xlsx_uses_locale_yaml_title_label(self, sample_xlsx_path: Path) -> None:
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(
                str(sample_xlsx_path),
                staging,
                options={"yaml_key_labels": {"title": "Titel"}},
            )
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success
            content = Path(result.artifacts[0].staging_path).read_text("utf-8")

            assert f"Titel: {sample_xlsx_path.stem}" in content
            assert f"title: {sample_xlsx_path.stem}" not in content

    def test_convert_xlsx_has_table(self, sample_xlsx_path: Path) -> None:
        """Output should contain markdown table with pipe syntax."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success
            content = Path(result.artifacts[0].staging_path).read_text("utf-8")

            # Should contain pipe table markers
            assert "|" in content
            # Should contain data values
            assert "Alpha" in content
            assert "99.90" in content or "99.9" in content

    def test_convert_xlsx_artifact_metadata(self, sample_xlsx_path: Path) -> None:
        """Artifact should contain metadata about the conversion."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success
            artifact = result.artifacts[0]
            assert artifact.metadata["sheet_count"] == 2  # Sales + Summary
            assert artifact.metadata["row_count"] > 0
            assert artifact.metadata["block_count"] > 0

    def test_convert_xlsx_metrics_present(self, sample_xlsx_path: Path) -> None:
        """Result should include valid metrics."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success
            assert result.metrics is not None
            assert result.metrics.input_bytes > 0
            assert result.metrics.output_bytes > 0

    def test_convert_csv_basic(self, sample_csv_path: Path) -> None:
        """Direct conversion of CSV to Markdown."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_csv_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success is True
            content = Path(result.artifacts[0].staging_path).read_text("utf-8")
            assert "Name" in content
            assert "Alice" in content
            assert "Bob" in content

    def test_convert_csv_has_title(self, sample_csv_path: Path) -> None:
        """CSV output should have filename as title heading."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_csv_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success
            content = Path(result.artifacts[0].staging_path).read_text("utf-8")
            assert "# test_data" in content

    def test_convert_csv_quoted_multiline_is_current_enhancement_boundary(self, tmp_path: Path) -> None:
        """Quoted multiline CSV succeeds in current direct route despite old detection failure."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["quoted_multiline_csv_scope"]
        input_path = tmp_path / scope["input"]["name"]
        input_path.write_text(scope["input"]["text"], encoding=scope["input"]["encoding"])

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(input_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success is True, f"unexpected error: {result.error}"
            assert len(result.artifacts) == 1
            artifact = result.artifacts[0]
            expected = scope["current_projection"]
            content = Path(artifact.staging_path).read_text(encoding="utf-8")

            assert artifact.media_type == expected["artifact_media_type"]
            assert artifact.suggested_name == expected["suggested_name"]
            assert artifact.metadata == expected["artifact_metadata"]
            assert result.metrics.extra == expected["metrics"]
            for token in expected["required_markdown_tokens"]:
                assert token in content
            table_lines = [line for line in content.splitlines() if line.startswith("|")]
            assert table_lines == expected["table_lines"]

    @pytest.mark.parametrize(
        ("case_name", "expected_suggested_name"),
        [
            ("gbk_simple.csv", "gbk_simple.md"),
            ("utf16_simple.csv", "utf16_simple.md"),
            ("semicolon_simple.csv", "semicolon_simple.md"),
        ],
    )
    def test_convert_csv_encoding_and_delimiter_variants_match_old_current_projection(
        self, tmp_path: Path, case_name: str, expected_suggested_name: str
    ) -> None:
        """Focused CSV encoding and delimiter variants remain old/current compatible."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["csv_encoding_delimiter_scope"]
        case = next(item for item in scope["probe_inputs"] if item["name"] == case_name)
        input_path = tmp_path / case["name"]
        input_path.write_text(case["text"], encoding=case["encoding"])

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(input_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success is True, f"unexpected error: {result.error}"
            assert len(result.artifacts) == 1
            artifact = result.artifacts[0]
            expected = scope["current_projection"]
            content = Path(artifact.staging_path).read_text(encoding="utf-8")

            assert artifact.media_type == expected["artifact_media_type"]
            assert artifact.suggested_name == expected_suggested_name
            assert artifact.metadata == expected["artifact_metadata"]
            assert result.metrics.extra == expected["metrics"]
            for token in expected["required_markdown_tokens"]:
                assert token in content
            table_lines = [line for line in content.splitlines() if line.startswith("|")]
            assert table_lines == expected["table_lines"]

    def test_convert_tsv_preserves_old_preconversion_blank_row_blocks(self, tmp_path: Path) -> None:
        """TSV blank rows remain table-block boundaries after direct-route migration."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        input_path = tmp_path / "delimited_probe.tsv"
        input_path.write_text(
            "\n".join(
                [
                    "Region\tProduct\tNote\tAmount",
                    "North\tAlpha\tline1 line2\t10",
                    "South\tBeta\tpipe | value\t20",
                    "",
                    "Region\tProduct\tNote\tAmount",
                    "East\tGamma\tsemi;colon\t5",
                    "West\tDelta\tplain\t7",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(input_path), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success is True
            artifact = result.artifacts[0]
            content = Path(artifact.staging_path).read_text(encoding="utf-8")

            assert artifact.metadata["block_count"] == 2
            separator_lines = [line for line in content.splitlines() if line.startswith("|:---------|")]
            assert len(separator_lines) == 2
            assert "pipe \\| value" in content
            assert "| East     | Gamma" in content

    def test_convert_xlsx_with_merge_strategy_option(self, sample_xlsx_path: Path) -> None:
        """Should accept and use table_merge_strategy option."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(
                str(sample_xlsx_path),
                staging,
                options={"table_merge_strategy": "empty"},
            )
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success is True
            content = Path(result.artifacts[0].staging_path).read_text("utf-8")
            assert "Q1 Sales Summary" in content

    def test_convert_xlsx_uses_data_only_formula_values_not_formula_text(self, tmp_path: Path) -> None:
        """GOLDEN-003 boundary: standard XLSX→MD exports values, not formula objects."""
        import openpyxl

        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Formula"
        sheet.cell(row=1, column=1, value="Item")
        sheet.cell(row=1, column=2, value="Qty")
        sheet.cell(row=2, column=1, value="Alpha")
        sheet.cell(row=2, column=2, value=1)
        sheet.cell(row=3, column=1, value="Beta")
        sheet.cell(row=3, column=2, value=2)
        sheet.cell(row=4, column=1, value="Total")
        sheet.cell(row=4, column=2, value="=SUM(B2:B3)")
        input_path = tmp_path / "formula_probe.xlsx"
        workbook.save(input_path)
        workbook.close()

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(input_path), staging, options={"to_md_keep_images": False})
            result = SpreadsheetToMarkdownConverter().convert(context)
            assert result.success is True
            content = Path(result.artifacts[0].staging_path).read_text("utf-8")
        assert "# Formula" in content
        assert "Alpha" in content
        assert "Beta" in content
        assert "Total" in content
        assert "=SUM(B2:B3)" not in content
        assert "SUM(B2:B3)" not in content

    def test_convert_xlsx_cached_formula_value_matches_old_system_projection(self, tmp_path: Path) -> None:
        """Cached XLSX formula values render as values across old/current systems."""
        import re
        import zipfile

        import openpyxl

        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["formula_style_protection_scope"]["cached_formula_probe"]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = scope["input"]["worksheet"]
        sheet.cell(row=1, column=1, value="Item")
        sheet.cell(row=1, column=2, value="Value")
        sheet.cell(row=2, column=1, value="Alpha")
        sheet.cell(row=2, column=2, value=10)
        sheet.cell(row=3, column=1, value="Beta")
        sheet.cell(row=3, column=2, value=5)
        sheet.cell(row=4, column=1, value="Total")
        sheet.cell(row=4, column=2, value=scope["input"]["formula_text"])
        input_path = tmp_path / scope["input"]["name"]
        workbook.save(input_path)
        workbook.close()

        extracted = tmp_path / "xlsx_parts"
        with zipfile.ZipFile(input_path, "r") as zip_in:
            zip_in.extractall(extracted)
        sheet_xml_path = extracted / "xl" / "worksheets" / "sheet1.xml"
        sheet_xml = sheet_xml_path.read_text(encoding="utf-8")
        sheet_xml = re.sub(
            r'<c r="B4"([^>]*)><f>SUM\(B2:B3\)</f><v></v></c>',
            r'<c r="B4"\1><f>SUM(B2:B3)</f><v>15</v></c>',
            sheet_xml,
        )
        assert "<v>15</v>" in sheet_xml
        sheet_xml_path.write_text(sheet_xml, encoding="utf-8")
        with zipfile.ZipFile(input_path, "w", zipfile.ZIP_DEFLATED) as zip_out:
            for file in extracted.rglob("*"):
                if file.is_file():
                    zip_out.write(file, file.relative_to(extracted).as_posix())

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(input_path), staging, options={"to_md_keep_images": False})
            result = SpreadsheetToMarkdownConverter().convert(context)
            assert result.success is True, f"unexpected error: {result.error}"
            artifact = result.artifacts[0]
            content = Path(artifact.staging_path).read_text("utf-8")

        expected = scope["current_projection"]
        assert artifact.media_type == expected["artifact_media_type"]
        assert artifact.suggested_name == expected["suggested_name"]
        assert artifact.metadata == expected["artifact_metadata"]
        assert result.metrics.extra == expected["metrics"]
        assert scope["input"]["formula_text"] not in content
        assert "SUM(B2:B3)" not in content
        assert "| Total  |      15 |" in content
        table_lines = [line for line in content.splitlines() if line.startswith("|")]
        assert table_lines == expected["table_lines"]

    def test_convert_error_on_invalid_file(self, tmp_path: Path) -> None:
        """Should return failure result for invalid files."""
        from docwen_plugin_spreadsheet.to_markdown.converter import (
            SpreadsheetToMarkdownConverter,
        )

        bad_file = tmp_path / "not_a_spreadsheet.txt"
        bad_file.write_text("This is not a spreadsheet.")

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(bad_file), staging)
            result = SpreadsheetToMarkdownConverter().convert(context)

            assert result.success is False
            assert result.error is not None
