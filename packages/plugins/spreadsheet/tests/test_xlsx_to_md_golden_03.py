"""Focused tests split from test_xlsx_to_md_golden.py."""

from __future__ import annotations

from ._xlsx_to_md_golden_support import (
    Any,
    Generator,
    Path,
    _deliverable_artifacts,
    _document_node_root,
    _io_path,
    _legacy_markdown_projection,
    _load_xlsx_to_md_old_system_fixture,
    pytest,
    tempfile,
)

pytestmark = pytest.mark.golden


@pytest.mark.integration
class TestSpreadsheetToMdPipeline:
    """Test spreadsheet→MD through the full runtime pipeline."""

    @pytest.fixture
    def pipeline(self) -> Generator[dict[str, Any], None, None]:
        """Build a full runtime pipeline for spreadsheet→md conversion."""
        from docwen_plugin_spreadsheet.plugin import SpreadsheetPlugin
        from docwen_runtime.engine.route_resolver import RouteResolver
        from docwen_runtime.engine.task_manager import TaskManager
        from docwen_runtime.output.finalizer import OutputFinalizer
        from docwen_runtime.plugin_registry.registry import PluginRegistry
        from docwen_runtime.workspace.manager import WorkspaceManager

        plugin = SpreadsheetPlugin()
        plugin_registry = PluginRegistry()
        plugin_registry.register(plugin)

        route_resolver = RouteResolver(plugin_registry)

        ws_root = tempfile.mkdtemp(prefix="docwen_test_ws_")
        ws_manager = WorkspaceManager(ws_root)
        finalizer = OutputFinalizer()
        task_manager = TaskManager(
            plugin_registry=plugin_registry,
            route_resolver=route_resolver,
            workspace_manager=ws_manager,
            output_finalizer=finalizer,
        )

        yield {
            "task_manager": task_manager,
            "workspace_manager": ws_manager,
            "ws_root": ws_root,
        }

        # Cleanup
        import shutil

        shutil.rmtree(ws_root, ignore_errors=True)

    def test_pipeline_xlsx_base64_image_mode_matches_old_system_projection(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
    ) -> None:
        """Focused XLSX base64 image-mode projection matches the old systems."""
        import io

        import openpyxl
        import openpyxl.drawing.image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["image_ocr_scope"]["base64_image_mode_probe"]
        image_bytes = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
            b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc```\x00\x00\x00\x04\x00\x01\xf6\x178U"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = scope["input"]["worksheet"]
        sheet.cell(row=1, column=1, value="Label")
        sheet.cell(row=1, column=2, value="Value")
        sheet.cell(row=2, column=1, value="Before image")
        sheet.cell(row=2, column=2, value=42)
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(image_bytes)),
            scope["input"]["image_anchor"],
        )
        input_path = tmp_path / scope["input"]["name"]
        workbook.save(input_path)
        workbook.close()

        output_dir = tmp_path / "base64_image_output"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-base64-image-projection",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={
                "to_md_keep_images": True,
                "to_md_enable_ocr": False,
                "image_mode": "base64",
                "image_link_style": "wiki_embed",
            },
        )

        result = pipeline["task_manager"].execute_single(request)

        expected = scope["current_projection"]
        assert result.success is True
        assert [diagnostic.code for diagnostic in result.diagnostics] == expected["diagnostic_codes"]
        assert len(_deliverable_artifacts(result)) == expected["artifact_count"]
        assert not any(artifact.kind == "image" for artifact in result.artifacts)
        markdown_artifact = next(artifact for artifact in result.artifacts if artifact.is_primary)
        assert markdown_artifact.metadata["image_count"] == expected["markdown_image_count"]
        _document_node_root(Path(markdown_artifact.staging_path), output_dir)
        assert Path(markdown_artifact.staging_path).is_file()

        content = Path(markdown_artifact.staging_path).read_text(encoding="utf-8")
        assert str(Path(pipeline["ws_root"])) not in content
        assert str(output_dir) not in content
        assert ".png" not in content
        for token in expected["required_markdown_tokens"]:
            assert token in content

    def test_pipeline_xlsx_suppresses_images_matches_old_system_projection(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
    ) -> None:
        """Focused XLSX no-image projection matches the old systems."""
        import io

        import openpyxl
        import openpyxl.drawing.image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["image_ocr_scope"]["suppress_image_probe"]
        image_bytes = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
            b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc```\x00\x00\x00\x04\x00\x01\xf6\x178U"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = scope["input"]["worksheet"]
        sheet.cell(row=1, column=1, value="Label")
        sheet.cell(row=1, column=2, value="Value")
        sheet.cell(row=2, column=1, value="Before image")
        sheet.cell(row=2, column=2, value=42)
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(image_bytes)),
            scope["input"]["image_anchor"],
        )
        input_path = tmp_path / scope["input"]["name"]
        workbook.save(input_path)
        workbook.close()

        output_dir = tmp_path / "suppress_image_output"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-suppress-image-projection",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={
                "to_md_keep_images": False,
                "to_md_enable_ocr": False,
                "image_mode": "file",
                "image_link_style": "wiki_embed",
            },
        )

        result = pipeline["task_manager"].execute_single(request)

        expected = scope["current_projection"]
        assert result.success is True
        assert [diagnostic.code for diagnostic in result.diagnostics] == expected["diagnostic_codes"]
        assert len(_deliverable_artifacts(result)) == expected["artifact_count"]
        assert not any(artifact.kind == "image" for artifact in result.artifacts)
        markdown_artifact = next(artifact for artifact in result.artifacts if artifact.is_primary)
        assert markdown_artifact.metadata["image_count"] == expected["markdown_image_count"]
        _document_node_root(Path(markdown_artifact.staging_path), output_dir)
        assert Path(markdown_artifact.staging_path).is_file()

        content = Path(markdown_artifact.staging_path).read_text(encoding="utf-8")
        assert str(Path(pipeline["ws_root"])) not in content
        assert str(output_dir) not in content
        assert ("data:image/png;base64," in content) is expected["contains_data_uri"]
        assert ("![[" in content) is expected["contains_wiki_image"]
        assert ("![" in content) is expected["contains_markdown_image"]
        assert (".png" in content) is expected["contains_png_token"]
        for token in expected["required_markdown_tokens"]:
            assert token in content

    def test_pipeline_xlsx_ocr_sidecar_without_images_matches_old_system_projection(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Focused XLSX no-image OCR sidecar projection matches the old systems."""
        import io

        import openpyxl
        import openpyxl.drawing.image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["image_ocr_scope"]["ocr_no_image_sidecar_probe"]
        image_bytes = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
            b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc```\x00\x00\x00\x04\x00\x01\xf6\x178U"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = scope["input"]["worksheet"]
        sheet.cell(row=1, column=1, value="Label")
        sheet.cell(row=1, column=2, value="Value")
        sheet.cell(row=2, column=1, value="Before image")
        sheet.cell(row=2, column=2, value=42)
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(image_bytes)),
            scope["input"]["image_anchor"],
        )
        input_path = tmp_path / scope["input"]["name"]
        workbook.save(input_path)
        workbook.close()

        import docwen_core.text.ocr as ocr

        monkeypatch.setattr(
            ocr,
            "run_ocr_outcome",
            lambda _path, *, source_format, ocr_language="auto", current_locale="zh_CN": ocr.OcrOutcome(
                ocr.OcrStatus.SUCCESS,
                text=scope["input"]["stubbed_ocr_text"],
            ),
        )

        output_dir = tmp_path / "ocr_no_image_output"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-ocr-no-image-sidecar-projection",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={
                "to_md_keep_images": False,
                "to_md_enable_ocr": True,
                "image_mode": "file",
                "ocr_placement": "image_md",
                "image_link_style": "wiki_embed",
            },
        )

        result = pipeline["task_manager"].execute_single(request)

        expected = scope["current_projection"]
        assert result.success is True
        assert [diagnostic.code for diagnostic in result.diagnostics] == expected["diagnostic_codes"]
        assert len(_deliverable_artifacts(result)) == expected["artifact_count"]
        primary_artifacts = [artifact for artifact in result.artifacts if artifact.is_primary]
        auxiliary_artifacts = [artifact for artifact in result.artifacts if artifact.kind == "auxiliary"]
        assert len(primary_artifacts) == expected["primary_artifact_count"]
        assert len(auxiliary_artifacts) == expected["auxiliary_artifact_count"]
        assert not any(artifact.kind == "image" for artifact in result.artifacts)
        markdown_artifact = primary_artifacts[0]
        sidecar_artifact = auxiliary_artifacts[0]
        assert markdown_artifact.metadata["image_count"] == expected["primary_metadata_image_count"]
        node_root = _document_node_root(Path(markdown_artifact.staging_path), output_dir)
        assert _document_node_root(Path(sidecar_artifact.staging_path), output_dir) == node_root
        assert Path(markdown_artifact.staging_path).is_file()
        assert _io_path(sidecar_artifact.staging_path).is_file()

        primary_content = Path(markdown_artifact.staging_path).read_text(encoding="utf-8")
        primary_content = _legacy_markdown_projection(primary_content, auxiliary_artifacts)
        sidecar_content = _legacy_markdown_projection(
            _io_path(sidecar_artifact.staging_path).read_text(encoding="utf-8"), auxiliary_artifacts
        )
        assert str(Path(pipeline["ws_root"])) not in primary_content
        assert str(Path(pipeline["ws_root"])) not in sidecar_content
        assert str(output_dir) not in primary_content
        assert str(output_dir) not in sidecar_content
        assert (scope["input"]["stubbed_ocr_text"] in primary_content) is expected["primary_markdown_contains_ocr_text"]
        assert ("![[" in primary_content) is expected["primary_markdown_contains_md_sidecar_link"]
        assert (".png" in primary_content) is expected["primary_markdown_contains_png_token"]
        assert (scope["input"]["stubbed_ocr_text"] in sidecar_content) is expected["sidecar_contains_ocr_text"]
        assert ("![" in sidecar_content) is expected["sidecar_contains_image_markdown"]
        for token in expected["required_primary_markdown_tokens"]:
            assert token in primary_content
        for token in expected["required_sidecar_markdown_tokens"]:
            assert token in sidecar_content

    def test_pipeline_xlsx_multi_ocr_sidecar_without_images_matches_old_system_projection(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Focused XLSX multi-image OCR sidecar projection matches the old systems."""
        import io

        import openpyxl
        import openpyxl.drawing.image
        from PIL import Image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["image_ocr_scope"]["multi_ocr_no_image_sidecar_probe"]

        def png_bytes(rgb: tuple[int, int, int]) -> bytes:
            buffer = io.BytesIO()
            Image.new("RGB", (1, 1), rgb).save(buffer, format="PNG")
            return buffer.getvalue()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = scope["input"]["worksheet"]
        sheet.cell(row=1, column=1, value="Label")
        sheet.cell(row=1, column=2, value="Value")
        sheet.cell(row=2, column=1, value="First OCR")
        sheet.cell(row=2, column=2, value=101)
        sheet.cell(row=4, column=1, value="Second OCR")
        sheet.cell(row=4, column=2, value=202)
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(png_bytes((255, 0, 0)))),
            scope["input"]["image_anchors"][0],
        )
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(png_bytes((0, 0, 255)))),
            scope["input"]["image_anchors"][1],
        )
        input_path = tmp_path / scope["input"]["name"]
        workbook.save(input_path)
        workbook.close()

        import docwen_core.text.ocr as ocr

        monkeypatch.setattr(
            ocr,
            "run_ocr_outcome",
            lambda _path, *, source_format, ocr_language="auto", current_locale="zh_CN": ocr.OcrOutcome(
                ocr.OcrStatus.SUCCESS,
                text=scope["input"]["stubbed_ocr_text"],
            ),
        )

        output_dir = tmp_path / "multi_ocr_no_image_output"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-multi-ocr-no-image-sidecar-projection",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={
                "to_md_keep_images": False,
                "to_md_enable_ocr": True,
                "image_mode": "file",
                "ocr_placement": "image_md",
                "image_link_style": "wiki_embed",
            },
        )

        result = pipeline["task_manager"].execute_single(request)

        expected = scope["current_projection"]
        assert result.success is True
        assert [diagnostic.code for diagnostic in result.diagnostics] == expected["diagnostic_codes"]
        assert len(_deliverable_artifacts(result)) == expected["artifact_count"]
        primary_artifacts = [artifact for artifact in result.artifacts if artifact.is_primary]
        auxiliary_artifacts = [artifact for artifact in result.artifacts if artifact.kind == "auxiliary"]
        assert len(primary_artifacts) == expected["primary_artifact_count"]
        assert len(auxiliary_artifacts) == expected["auxiliary_artifact_count"]
        assert not any(artifact.kind == "image" for artifact in result.artifacts)
        assert [artifact.metadata["source_suggested_name"] for artifact in auxiliary_artifacts] == expected[
            "sidecar_suggested_names"
        ]

        markdown_artifact = primary_artifacts[0]
        for key, value in expected["primary_metadata"].items():
            assert markdown_artifact.metadata[key] == value
        node_root = _document_node_root(Path(markdown_artifact.staging_path), output_dir)
        for artifact in auxiliary_artifacts:
            assert _document_node_root(Path(artifact.staging_path), output_dir) == node_root
            assert _io_path(artifact.staging_path).is_file()

        primary_content = _legacy_markdown_projection(
            Path(markdown_artifact.staging_path).read_text(encoding="utf-8"), auxiliary_artifacts
        )
        sidecar_contents = [
            _legacy_markdown_projection(
                _io_path(artifact.staging_path).read_text(encoding="utf-8"), auxiliary_artifacts
            )
            for artifact in auxiliary_artifacts
        ]
        assert str(Path(pipeline["ws_root"])) not in primary_content
        assert str(output_dir) not in primary_content
        assert all(str(Path(pipeline["ws_root"])) not in content for content in sidecar_contents)
        assert all(str(output_dir) not in content for content in sidecar_contents)
        assert (scope["input"]["stubbed_ocr_text"] in primary_content) is expected["primary_markdown_contains_ocr_text"]
        assert ("![[" in primary_content) is expected["primary_markdown_contains_md_sidecar_link"]
        assert (".png" in primary_content) is expected["primary_markdown_contains_png_token"]
        assert (
            all(scope["input"]["stubbed_ocr_text"] in content for content in sidecar_contents)
            is expected["sidecars_contain_ocr_text"]
        )
        assert any("![" in content for content in sidecar_contents) is expected["sidecars_contain_image_markdown"]
        first_index = primary_content.index("multi_ocr_no_image_probe__img_001_ocr.md")
        second_index = primary_content.index("multi_ocr_no_image_probe__img_002_ocr.md")
        assert (first_index < second_index) is expected["sidecar_links_in_order"]
        for token in expected["required_primary_markdown_tokens"]:
            assert token in primary_content
        for content in sidecar_contents:
            for token in expected["required_sidecar_markdown_tokens"]:
                assert token in content
