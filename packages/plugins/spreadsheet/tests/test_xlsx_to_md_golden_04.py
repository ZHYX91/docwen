"""Focused tests split from test_xlsx_to_md_golden.py."""

from __future__ import annotations

from ._xlsx_to_md_golden_support import (
    Any,
    Generator,
    Path,
    _deliverable_artifacts,
    _document_node_root,
    _io_path,
    _legacy_markdown_projection,
    _load_xlsx_to_md_old_system_fixture,
    pytest,
    tempfile,
)

pytestmark = pytest.mark.golden


@pytest.mark.integration
class TestSpreadsheetToMdPipeline:
    """Test spreadsheet→MD through the full runtime pipeline."""

    @pytest.fixture
    def pipeline(self) -> Generator[dict[str, Any], None, None]:
        """Build a full runtime pipeline for spreadsheet→md conversion."""
        from docwen_plugin_spreadsheet.plugin import SpreadsheetPlugin
        from docwen_runtime.engine.route_resolver import RouteResolver
        from docwen_runtime.engine.task_manager import TaskManager
        from docwen_runtime.output.finalizer import OutputFinalizer
        from docwen_runtime.plugin_registry.registry import PluginRegistry
        from docwen_runtime.workspace.manager import WorkspaceManager

        plugin = SpreadsheetPlugin()
        plugin_registry = PluginRegistry()
        plugin_registry.register(plugin)

        route_resolver = RouteResolver(plugin_registry)

        ws_root = tempfile.mkdtemp(prefix="docwen_test_ws_")
        ws_manager = WorkspaceManager(ws_root)
        finalizer = OutputFinalizer()
        task_manager = TaskManager(
            plugin_registry=plugin_registry,
            route_resolver=route_resolver,
            workspace_manager=ws_manager,
            output_finalizer=finalizer,
        )

        yield {
            "task_manager": task_manager,
            "workspace_manager": ws_manager,
            "ws_root": ws_root,
        }

        # Cleanup
        import shutil

        shutil.rmtree(ws_root, ignore_errors=True)

    def test_pipeline_xlsx_multi_ocr_sidecar_with_images_matches_old_system_projection(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Focused XLSX multi-image OCR sidecar with retained images matches old systems."""
        import io

        import openpyxl
        import openpyxl.drawing.image
        from PIL import Image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        fixture = _load_xlsx_to_md_old_system_fixture()
        scope = fixture["shared_behavior_evidence"]["image_ocr_scope"]["multi_ocr_with_images_sidecar_probe"]

        def png_bytes(rgb: tuple[int, int, int]) -> bytes:
            buffer = io.BytesIO()
            Image.new("RGB", (1, 1), rgb).save(buffer, format="PNG")
            return buffer.getvalue()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = scope["input"]["worksheet"]
        sheet.cell(row=1, column=1, value="Label")
        sheet.cell(row=1, column=2, value="Value")
        sheet.cell(row=2, column=1, value="First OCR image")
        sheet.cell(row=2, column=2, value=301)
        sheet.cell(row=4, column=1, value="Second OCR image")
        sheet.cell(row=4, column=2, value=302)
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(png_bytes((255, 0, 0)))),
            scope["input"]["image_anchors"][0],
        )
        sheet.add_image(
            openpyxl.drawing.image.Image(io.BytesIO(png_bytes((0, 0, 255)))),
            scope["input"]["image_anchors"][1],
        )
        input_path = tmp_path / scope["input"]["name"]
        workbook.save(input_path)
        workbook.close()

        import docwen_core.text.ocr as ocr

        monkeypatch.setattr(
            ocr,
            "run_ocr_outcome",
            lambda _path, *, source_format, ocr_language="auto", current_locale="zh_CN": ocr.OcrOutcome(
                ocr.OcrStatus.SUCCESS,
                text=scope["input"]["stubbed_ocr_text"],
            ),
        )

        output_dir = tmp_path / "multi_ocr_with_images_output"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-multi-ocr-with-images-sidecar-projection",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={
                "to_md_keep_images": True,
                "to_md_enable_ocr": True,
                "image_mode": "file",
                "ocr_placement": "image_md",
                "image_link_style": "wiki_embed",
            },
        )

        result = pipeline["task_manager"].execute_single(request)

        expected = scope["current_projection"]
        assert result.success is True
        assert [diagnostic.code for diagnostic in result.diagnostics] == expected["diagnostic_codes"]
        assert len(_deliverable_artifacts(result)) == expected["artifact_count"]
        primary_artifacts = [artifact for artifact in result.artifacts if artifact.is_primary]
        auxiliary_artifacts = [artifact for artifact in result.artifacts if artifact.kind == "auxiliary"]
        image_artifacts = [artifact for artifact in result.artifacts if artifact.kind == "image"]
        assert len(primary_artifacts) == expected["primary_artifact_count"]
        assert len(auxiliary_artifacts) == expected["auxiliary_artifact_count"]
        assert len(image_artifacts) == expected["image_artifact_count"]
        assert [artifact.suggested_name for artifact in image_artifacts] == expected["image_suggested_names"]
        assert [artifact.metadata["source_suggested_name"] for artifact in auxiliary_artifacts] == expected[
            "sidecar_suggested_names"
        ]

        markdown_artifact = primary_artifacts[0]
        for key, value in expected["primary_metadata"].items():
            assert markdown_artifact.metadata[key] == value
        node_root = _document_node_root(Path(markdown_artifact.staging_path), output_dir)
        for artifact in auxiliary_artifacts + image_artifacts:
            assert _document_node_root(Path(artifact.staging_path), output_dir) == node_root
            assert _io_path(artifact.staging_path).is_file()

        primary_content = _legacy_markdown_projection(
            Path(markdown_artifact.staging_path).read_text(encoding="utf-8"), auxiliary_artifacts
        )
        sidecar_contents = [
            _legacy_markdown_projection(
                _io_path(artifact.staging_path).read_text(encoding="utf-8"), auxiliary_artifacts
            )
            for artifact in auxiliary_artifacts
        ]
        assert str(Path(pipeline["ws_root"])) not in primary_content
        assert str(output_dir) not in primary_content
        assert all(str(Path(pipeline["ws_root"])) not in content for content in sidecar_contents)
        assert all(str(output_dir) not in content for content in sidecar_contents)
        assert (scope["input"]["stubbed_ocr_text"] in primary_content) is expected["primary_markdown_contains_ocr_text"]
        assert ("![[" in primary_content) is expected["primary_markdown_contains_md_sidecar_link"]
        assert (".png" in primary_content) is expected["primary_markdown_contains_png_token"]
        assert (
            all(scope["input"]["stubbed_ocr_text"] in content for content in sidecar_contents)
            is expected["sidecars_contain_ocr_text"]
        )
        assert (
            any("![[" in content and ".png]]" in content for content in sidecar_contents)
            is expected["sidecars_contain_image_markdown"]
        )
        first_index = primary_content.index("multi_ocr_with_images_probe__img_001_ocr.md")
        second_index = primary_content.index("multi_ocr_with_images_probe__img_002_ocr.md")
        assert (first_index < second_index) is expected["sidecar_links_in_order"]
        for token in expected["required_primary_markdown_tokens"]:
            assert token in primary_content
        for content in sidecar_contents:
            for token in expected["required_sidecar_markdown_tokens"]:
                assert token in content

    def test_pipeline_xlsx_all_ocr_outcomes_warn_and_continue_later_images(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Typed OCR failures warn while the workbook and later images continue."""
        import io

        import openpyxl
        import openpyxl.drawing.image
        from PIL import Image

        import docwen_core.text.ocr as ocr
        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Best effort"
        sheet.cell(row=1, column=1, value="spreadsheet base content")
        for index in range(6):
            image_buffer = io.BytesIO()
            Image.new("RGB", (2, 2), color=(index * 20, 0, 0)).save(image_buffer, format="PNG")
            sheet.add_image(
                openpyxl.drawing.image.Image(io.BytesIO(image_buffer.getvalue())),
                f"A{index * 2 + 2}",
            )
        input_path = tmp_path / "spreadsheet-ocr-best-effort.xlsx"
        workbook.save(input_path)
        workbook.close()

        outcomes = iter(
            [
                ocr.OcrOutcome(ocr.OcrStatus.UNAVAILABLE, message="private unavailable detail"),
                ocr.OcrOutcome(ocr.OcrStatus.MODEL_MISSING, message="private model path"),
                ocr.OcrOutcome(ocr.OcrStatus.INITIALIZATION_FAILED, message="private init detail"),
                ocr.OcrOutcome(ocr.OcrStatus.RECOGNITION_FAILED, message="private recognition detail"),
                ocr.OcrOutcome(ocr.OcrStatus.NO_TEXT),
                ocr.OcrOutcome(ocr.OcrStatus.SUCCESS, text="later spreadsheet OCR"),
            ]
        )
        calls: list[str] = []

        def _run_ocr_outcome(path: str, **_kwargs: Any) -> ocr.OcrOutcome:
            calls.append(str(path))
            return next(outcomes)

        monkeypatch.setattr(ocr, "run_ocr_outcome", _run_ocr_outcome)

        output_dir = tmp_path / "spreadsheet-ocr-best-effort-output"
        output_dir.mkdir()
        request = ConversionRequest(
            request_id="spreadsheet-ocr-best-effort",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={
                "to_md_keep_images": False,
                "to_md_enable_ocr": True,
                "ocr_placement": "main_md",
            },
        )
        events: list[Any] = []

        result = pipeline["task_manager"].execute_single(request, on_event=events.append)

        assert result.success is True
        markdown = Path(next(artifact for artifact in result.artifacts if artifact.is_primary).staging_path).read_text(
            encoding="utf-8"
        )
        assert "spreadsheet base content" in markdown
        assert "later spreadsheet OCR" in markdown
        assert len(calls) == 6
        warnings = [
            event.payload
            for event in events
            if event.event_type == "diagnostic" and event.payload.get("code") == "OCR-BEST-EFFORT"
        ]
        assert len(warnings) == 6
        assert [warning["level"] for warning in warnings] == ["warning"] * 6
        assert [warning["message"].split("status=", 1)[1].split(";", 1)[0] for warning in warnings] == [
            "unavailable",
            "model_missing",
            "initialization_failed",
            "recognition_failed",
            "no_text",
            "success",
        ]
        assert all(warning["location"].startswith("Best effort:") for warning in warnings)
        assert all("private" not in warning["message"] for warning in warnings)

    def test_pipeline_xlsx_image_mode_falls_back_to_export_semantics(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
    ) -> None:
        """XLSX→MD must inherit Export image mode when the request omits image_mode."""
        import io

        import openpyxl
        import openpyxl.drawing.image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        image_bytes = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
            b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc```\x00\x00\x00\x04\x00\x01\xf6\x178U"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Data"
        sheet.cell(row=1, column=1, value="Header")
        sheet.add_image(openpyxl.drawing.image.Image(io.BytesIO(image_bytes)), "B2")

        input_path = tmp_path / "xlsx_export_fallback_probe.xlsx"
        workbook.save(input_path)
        workbook.close()

        output_dir = tmp_path / "output_xlsx_export_fallback"
        output_dir.mkdir()

        request = ConversionRequest(
            request_id="pipe-xlsx-export-fallback",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={"to_md_keep_images": True, "image_link_style": "markdown_embed"},
            config_snapshot={"export": {"to_md_image_extraction_mode": "base64"}},
        )

        result = pipeline["task_manager"].execute_single(request)

        assert result.success is True
        markdown_artifact = next(artifact for artifact in result.artifacts if artifact.is_primary)
        content = Path(markdown_artifact.staging_path).read_text(encoding="utf-8")
        assert "data:image/png;base64," in content
        assert not any(artifact.kind == "image" for artifact in result.artifacts)

    def test_pipeline_xlsx_nonempty_snapshot_is_authoritative(
        self,
        pipeline: dict[str, Any],
        tmp_path: Path,
    ) -> None:
        import io

        import openpyxl
        import openpyxl.drawing.image

        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        image_bytes = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
            b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
            b"\x00\x00\x00\x0cIDATx\x9cc```\x00\x00\x00\x04\x00\x01\xf6\x178U"
            b"\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.cell(row=1, column=1, value="Header")
        sheet.add_image(openpyxl.drawing.image.Image(io.BytesIO(image_bytes)), "B2")
        input_path = tmp_path / "xlsx_request_snapshot_probe.xlsx"
        workbook.save(input_path)
        workbook.close()

        output_dir = tmp_path / "output_xlsx_request_snapshot"
        output_dir.mkdir()
        request = ConversionRequest(
            request_id="pipe-xlsx-request-snapshot",
            input_refs=[
                FileRef(
                    path=str(input_path),
                    format="xlsx",
                    category="spreadsheet",
                    size_bytes=input_path.stat().st_size,
                )
            ],
            target_format="md",
            output_policy=OutputPolicy(output_dir=str(output_dir)),
            options={"to_md_keep_images": True, "image_link_style": "markdown_embed"},
            config_snapshot={"export": {"to_md_image_extraction_mode": "base64"}},
        )

        result = pipeline["task_manager"].execute_single(request)

        assert result.success is True
        markdown_artifact = next(artifact for artifact in result.artifacts if artifact.is_primary)
        content = Path(markdown_artifact.staging_path).read_text(encoding="utf-8")
        assert "data:image/png;base64," in content
        assert not any(artifact.kind == "image" for artifact in result.artifacts)

    def test_pipeline_staging_only_write(self, pipeline: dict[str, Any], sample_xlsx_path: Path) -> None:
        """Plugin should only write to staging, not final output."""
        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        file_ref = FileRef(
            path=str(sample_xlsx_path),
            format="xlsx",
            category="spreadsheet",
        )
        request = ConversionRequest(
            request_id="pipe-002",
            input_refs=[file_ref],
            target_format="md",
            output_policy=OutputPolicy(),
        )

        result = pipeline["task_manager"].execute_single(request)
        assert result.success

        # Artifacts should have been finalized by output finalizer
        # (staging_path goes through the finalizer)
        for artifact in result.artifacts:
            assert artifact.staging_path, f"Artifact {artifact.artifact_id} has no staging_path"

    def test_pipeline_csv_to_md(self, pipeline: dict[str, Any], sample_csv_path: Path) -> None:
        """Full pipeline: CSV→MD through runtime (P0-5)."""
        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        file_ref = FileRef(
            path=str(sample_csv_path),
            format="csv",
            category="spreadsheet",
        )
        request = ConversionRequest(
            request_id="pipe-003",
            input_refs=[file_ref],
            target_format="md",
            output_policy=OutputPolicy(),
        )

        result = pipeline["task_manager"].execute_single(request)
        assert result.success is True
        assert len(result.artifacts) >= 1
        assert result.artifacts[0].kind == "primary"
        assert result.artifacts[0].media_type == "text/markdown"

        # Verify content — CSV data should be preserved in Markdown output
        content = Path(result.artifacts[0].staging_path).read_text("utf-8")
        assert "Name" in content
        assert "Alice" in content
        assert "New York" in content
        assert "# test_data" in content  # CSV filename as title

    def test_pipeline_tsv_to_md(self, pipeline: dict[str, Any], sample_tsv_path: Path) -> None:
        """Full pipeline: TSV→MD through runtime (M-4)."""
        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        file_ref = FileRef(
            path=str(sample_tsv_path),
            format="tsv",
            category="spreadsheet",
        )
        request = ConversionRequest(
            request_id="pipe-004",
            input_refs=[file_ref],
            target_format="md",
            output_policy=OutputPolicy(),
        )

        result = pipeline["task_manager"].execute_single(request)
        assert result.success is True
        assert len(result.artifacts) >= 1
        assert result.artifacts[0].kind == "primary"
        assert result.artifacts[0].media_type == "text/markdown"

        # Verify content — TSV data should be preserved in Markdown output
        content = Path(result.artifacts[0].staging_path).read_text("utf-8")
        assert "ID" in content
        assert "95" in content
        assert "A" in content
        assert "# test_data" in content  # TSV filename as title
