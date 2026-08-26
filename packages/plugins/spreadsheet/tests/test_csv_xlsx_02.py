"""Focused tests split from test_csv_xlsx.py."""

from __future__ import annotations

from ._csv_xlsx_support import (
    Any,
    Path,
    _load_merge_tables_broader_old_system_fixture,
    _load_merge_tables_old_system_fixture,
    _read_broader_merge_projection,
    _read_cached_formula_merge_projection,
    _read_used_values,
    _write_broader_merge_workbook,
    _write_cached_formula_merge_workbook,
    _write_workbook,
    os,
    pytest,
    tempfile,
)

pytestmark = [pytest.mark.golden, pytest.mark.contract]


class TestTableMerge:
    """ACT-MERGE-TABLES: Table merging tests."""

    def _create_fixture_inputs(self, tmp_path: Path, mode: str) -> list[str]:
        fixture = _load_merge_tables_old_system_fixture()
        base_path = tmp_path / f"base_{mode}.xlsx"
        collect_path = tmp_path / f"collect_{mode}.xlsx"
        _write_workbook(base_path, fixture["input_workbooks"]["base.xlsx"])
        _write_workbook(collect_path, fixture["input_workbooks"]["collect.xlsx"])
        return [str(base_path), str(collect_path)]

    def _build_merge_context(
        self,
        input_paths: list[str],
        staging_dir: str,
        options: dict[str, Any] | None = None,
        input_formats: list[str] | None = None,
    ) -> Any:
        from tests.support.config import FakeConfigView
        from tests.support.execution import FakeExecutionContext
        from tests.support.logging import FakePluginLogger
        from tests.support.progress import FakeProgressSink
        from tests.support.workspace import FakeWorkspaceHandle

        from docwen_core.cancellation import CancellationToken
        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy

        token = CancellationToken()
        formats = input_formats or [Path(p).suffix.lstrip(".") for p in input_paths]
        file_refs = [
            FileRef(path=path, format=source_format, category="spreadsheet")
            for path, source_format in zip(input_paths, formats, strict=True)
        ]
        request = ConversionRequest(
            request_id="test-merge-001",
            input_refs=file_refs,
            target_format="xlsx",
            action_name="merge_tables",
            options=options or {},
            output_policy=OutputPolicy(),
        )
        workspace = FakeWorkspaceHandle(input_paths[0], staging_dir)
        progress = FakeProgressSink()
        config = FakeConfigView()

        return FakeExecutionContext(
            request=request,
            workspace=workspace,
            config=config,
            progress=progress,
            cancellation=token,
            logger=FakePluginLogger(),
        )

    def test_merge_by_cell_basic(self, tmp_path: Path) -> None:
        """Cell-mode merge should combine two XLSX files with matching headers."""
        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        fixture = _load_merge_tables_old_system_fixture()
        expected = fixture["projects"]["docwen-current"]["cell"]
        input_paths = self._create_fixture_inputs(tmp_path, "cell")

        with tempfile.TemporaryDirectory() as staging:
            context = self._build_merge_context(
                input_paths,
                staging,
                options={"merge_mode": "cell"},
            )
            result = TableMergerConverter().convert(context)

            assert result.success is True
            assert len(result.artifacts) == 1
            artifact = result.artifacts[0]
            assert os.path.isfile(artifact.staging_path)
            assert artifact.suggested_name == expected["suggested_name"]
            assert artifact.media_type == expected["media_type"]
            assert artifact.metadata == expected["metadata"]
            assert result.metrics.extra == expected["metrics"]
            assert _read_used_values(artifact.staging_path) == expected["values"]

    def test_merge_uses_each_admitted_format_when_xlsx_suffixes_are_wrong(self, tmp_path: Path) -> None:
        """Every merge input keeps its own Core-admitted concrete format."""
        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        base_path = tmp_path / "base.txt"
        collect_path = tmp_path / "collect.csv"
        canonical_base = tmp_path / "base.xlsx"
        canonical_collect = tmp_path / "collect.xlsx"
        _write_workbook(canonical_base, [["Item", "Value"], ["Alpha", 10]])
        _write_workbook(canonical_collect, [["Item", "Value"], ["Alpha", 2]])
        base_path.write_bytes(canonical_base.read_bytes())
        collect_path.write_bytes(canonical_collect.read_bytes())

        staging = tmp_path / "merge-wrong-suffix-staging"
        staging.mkdir()
        context = self._build_merge_context(
            [str(base_path), str(collect_path)],
            str(staging),
            options={"merge_mode": "cell"},
            input_formats=["xlsx", "xlsx"],
        )

        result = TableMergerConverter().convert(context)

        assert result.success is True
        assert _read_used_values(result.artifacts[0].staging_path) == [
            ["Item", "Value"],
            ["Alpha", 12],
        ]

    def test_merge_by_row_basic(self, tmp_path: Path) -> None:
        """Row-mode merge should insert collect rows into base at similar positions."""
        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        fixture = _load_merge_tables_old_system_fixture()
        expected = fixture["projects"]["docwen-current"]["row"]
        input_paths = self._create_fixture_inputs(tmp_path, "row")

        with tempfile.TemporaryDirectory() as staging:
            context = self._build_merge_context(
                input_paths,
                staging,
                options={"merge_mode": "row"},
            )
            result = TableMergerConverter().convert(context)

            assert result.success is True
            assert len(result.artifacts) == 1
            artifact = result.artifacts[0]
            assert os.path.isfile(artifact.staging_path)
            assert artifact.suggested_name == expected["suggested_name"]
            assert artifact.media_type == expected["media_type"]
            assert artifact.metadata == expected["metadata"]
            assert result.metrics.extra == expected["metrics"]
            assert _read_used_values(artifact.staging_path) == expected["values"]

    def test_merge_by_column_basic_matches_old_system_fixture(self, tmp_path: Path) -> None:
        """Column-mode merge should interleave collect columns like the old systems."""
        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        fixture = _load_merge_tables_old_system_fixture()
        expected = fixture["projects"]["docwen-current"]["col"]
        input_paths = self._create_fixture_inputs(tmp_path, "col")

        with tempfile.TemporaryDirectory() as staging:
            context = self._build_merge_context(
                input_paths,
                staging,
                options={"merge_mode": "col"},
            )
            result = TableMergerConverter().convert(context)

            assert result.success is True
            assert len(result.artifacts) == 1
            artifact = result.artifacts[0]
            assert os.path.isfile(artifact.staging_path)
            assert artifact.suggested_name == expected["suggested_name"]
            assert artifact.media_type == expected["media_type"]
            assert artifact.metadata == expected["metadata"]
            assert result.metrics.extra == expected["metrics"]
            assert _read_used_values(artifact.staging_path) == expected["values"]

    def test_merge_preprocesses_merged_cells_like_old_systems(self, tmp_path: Path) -> None:
        """Merged cells are unmerged and filled before merge, matching old implementations."""
        import openpyxl

        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        fixture = _load_merge_tables_old_system_fixture()
        expected = fixture["shared_behavior_evidence"]["merged_cell_preprocessing"]

        base_path = tmp_path / "base_merged.xlsx"
        base_wb = openpyxl.Workbook()
        base_ws = base_wb.active
        assert base_ws is not None
        base_ws["A1"] = "Merged Header"
        base_ws["C1"] = "Score"
        base_ws["A2"] = "Alice"
        base_ws["B2"] = "Team A"
        base_ws["C2"] = 90
        base_ws.merge_cells(expected["input_merged_range"])
        base_wb.save(base_path)
        base_wb.close()

        collect_path = tmp_path / "collect_empty.xlsx"
        collect_wb = openpyxl.Workbook()
        collect_wb.save(collect_path)
        collect_wb.close()

        with tempfile.TemporaryDirectory() as staging:
            context = self._build_merge_context(
                [str(base_path), str(collect_path)],
                staging,
                options={"merge_mode": "cell"},
            )
            result = TableMergerConverter().convert(context)

            assert result.success is True
            artifact = result.artifacts[0]
            wb = openpyxl.load_workbook(artifact.staging_path, data_only=True)
            try:
                ws = wb.active
                assert ws is not None
                assert [str(merged_range) for merged_range in ws.merged_cells.ranges] == []
                assert _read_used_values(artifact.staging_path) == expected["filled_values"]
            finally:
                wb.close()

    def test_merge_by_cell_broader_workbook_projection_matches_old_systems(self, tmp_path: Path) -> None:
        """Cell-mode projection covers mixed sheets, protection, formulas, style, and merged cells."""
        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        fixture = _load_merge_tables_broader_old_system_fixture()
        base_path = tmp_path / "broader_base.xlsx"
        collect_path = tmp_path / "broader_collect.xlsx"
        _write_broader_merge_workbook(
            base_path,
            fixture["input_workbooks"]["broader_base.xlsx"],
        )
        _write_broader_merge_workbook(
            collect_path,
            fixture["input_workbooks"]["broader_collect.xlsx"],
        )
        expected = fixture["projects"]["docwen-current"]["cell"]

        with tempfile.TemporaryDirectory() as staging:
            context = self._build_merge_context(
                [str(base_path), str(collect_path)],
                staging,
                options={"merge_mode": "cell"},
            )
            result = TableMergerConverter().convert(context)

            assert result.success is True
            assert len(result.artifacts) == 1
            artifact = result.artifacts[0]
            assert artifact.suggested_name == expected["suggested_name"]
            assert artifact.media_type == expected["media_type"]
            assert artifact.metadata == expected["metadata"]
            assert result.metrics.extra == expected["metrics"]
            assert _read_broader_merge_projection(artifact.staging_path) == expected["projection"]

        expected_projection = expected["projection"]
        for project_name in ("docwen-ref-tk", "docwen-ref-pyside6"):
            assert fixture["projects"][project_name]["cell"]["projection"] == expected_projection

    def test_merge_by_cell_cached_formula_values_match_old_systems(self, tmp_path: Path) -> None:
        """Cell-mode merge uses cached/data-only formula values like the old systems."""
        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        fixture = _load_merge_tables_broader_old_system_fixture()
        scope = fixture["cached_formula_value_probe"]
        base_path = tmp_path / "cached_base.xlsx"
        collect_path = tmp_path / "cached_collect.xlsx"
        _write_cached_formula_merge_workbook(
            base_path,
            scope["input_workbooks"]["cached_base.xlsx"],
        )
        _write_cached_formula_merge_workbook(
            collect_path,
            scope["input_workbooks"]["cached_collect.xlsx"],
        )
        expected = scope["projects"]["docwen-current"]["cell"]

        with tempfile.TemporaryDirectory() as staging:
            context = self._build_merge_context(
                [str(base_path), str(collect_path)],
                staging,
                options={"merge_mode": "cell"},
            )
            result = TableMergerConverter().convert(context)

            assert result.success is True
            assert len(result.artifacts) == 1
            artifact = result.artifacts[0]
            assert artifact.suggested_name == expected["suggested_name"]
            assert artifact.media_type == expected["media_type"]
            assert artifact.metadata == expected["metadata"]
            assert result.metrics.extra == expected["metrics"]
            assert _read_cached_formula_merge_projection(artifact.staging_path) == expected["projection"]

        expected_projection = expected["projection"]
        for project_name in ("docwen-ref-tk", "docwen-ref-pyside6"):
            project = scope["projects"][project_name]["cell"]
            assert project["success"] is True
            assert project["projection"] == expected_projection

    def test_merge_preconverts_legacy_collect_workbook_through_spreadsheet_hub(
        self,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Binary collection inputs use the production spreadsheet-to-XLSX bridge before openpyxl."""
        from docwen_plugin_spreadsheet.format_conversion.converter import SmartSheetConverter
        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        base_path = tmp_path / "base.xlsx"
        hub_collect = tmp_path / "legacy-collect-converted.xlsx"
        # The admitted concrete format, not the misleading suffix, selects the
        # spreadsheet bridge.
        legacy_collect = tmp_path / "legacy-collect.xlsx"
        _write_workbook(base_path, [["Item", "Value"], ["Alpha", 10]])
        _write_workbook(hub_collect, [["Item", "Value"], ["Alpha", 2]])
        legacy_collect.write_bytes(b"legacy spreadsheet placeholder")
        calls: list[tuple[str, str]] = []

        def fake_prepare(_self, _context, path: str, source: str) -> tuple[str, str]:
            calls.append((Path(path).name, source))
            return str(hub_collect), "Fake Spreadsheet Bridge"

        monkeypatch.setattr(SmartSheetConverter, "_prepare_hub_xlsx", fake_prepare)

        with tempfile.TemporaryDirectory() as staging:
            context = self._build_merge_context(
                [str(base_path), str(legacy_collect)],
                staging,
                options={"merge_mode": "cell"},
                input_formats=["xlsx", "xls"],
            )
            result = TableMergerConverter().convert(context)

            assert result.success is True
            assert calls == [("legacy-collect.xlsx", "xls")]
            assert result.artifacts[0].suggested_name == "base_merged.xlsx"
            assert _read_used_values(result.artifacts[0].staging_path) == [
                ["Item", "Value"],
                ["Alpha", 12],
            ]

    def test_merge_needs_multiple_files(self, sample_base_xlsx_path: Path) -> None:
        """Should fail with only one input file."""
        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        with tempfile.TemporaryDirectory() as staging:
            context = self._build_merge_context(
                [str(sample_base_xlsx_path)],
                staging,
                options={"merge_mode": "cell"},
            )
            result = TableMergerConverter().convert(context)

            assert result.success is False
            assert result.error is not None
            assert "NEED-MORE-FILES" in (result.error.diagnostic_code or "")

    def test_merge_reports_bad_collect_file_name(self, sample_base_xlsx_path: Path, tmp_path: Path) -> None:
        """Invalid collection workbook errors should identify the offending file."""
        import zipfile

        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        broken = tmp_path / "broken.xlsx"
        with zipfile.ZipFile(broken, "w") as archive:
            archive.writestr("not-a-workbook.txt", "not an xlsx workbook")

        with tempfile.TemporaryDirectory() as staging:
            context = self._build_merge_context(
                [str(sample_base_xlsx_path), str(broken)],
                staging,
                options={"merge_mode": "row"},
            )
            result = TableMergerConverter().convert(context)

            assert result.success is False
            assert result.error is not None
            assert result.error.error_type == "conversion_failed"
            assert "broken.xlsx" in result.error.message
            assert "Failed to merge" in result.error.message

    def test_merge_tables_pipeline_finalizes_collision_with_legacy_suffix(self, tmp_path: Path) -> None:
        """Runtime finalizer gives repeated merge outputs the old-system _001 suffix."""
        from docwen_core.models.file_ref import FileRef
        from docwen_core.models.request import ConversionRequest, OutputPolicy
        from docwen_plugin_spreadsheet.plugin import SpreadsheetPlugin
        from docwen_runtime.engine.route_resolver import RouteResolver
        from docwen_runtime.engine.task_manager import TaskManager
        from docwen_runtime.output.finalizer import OutputFinalizer
        from docwen_runtime.plugin_registry.registry import PluginRegistry
        from docwen_runtime.workspace.manager import WorkspaceManager

        output_dir = tmp_path / "out"
        output_dir.mkdir()
        workspace_root = tmp_path / "workspace"

        registry = PluginRegistry()
        registry.register(SpreadsheetPlugin())
        task_manager = TaskManager(
            plugin_registry=registry,
            route_resolver=RouteResolver(registry),
            workspace_manager=WorkspaceManager(str(workspace_root)),
            output_finalizer=OutputFinalizer(),
        )

        fixture = _load_merge_tables_old_system_fixture()
        base_path = tmp_path / "base.xlsx"
        collect_path = tmp_path / "collect.xlsx"
        _write_workbook(base_path, fixture["input_workbooks"]["base.xlsx"])
        _write_workbook(collect_path, fixture["input_workbooks"]["collect.xlsx"])

        def _request(request_id: str) -> ConversionRequest:
            return ConversionRequest(
                request_id=request_id,
                input_refs=[
                    FileRef(path=str(base_path), format="xlsx", category="spreadsheet"),
                    FileRef(path=str(collect_path), format="xlsx", category="spreadsheet"),
                ],
                target_format="xlsx",
                action_name="merge_tables",
                options={"merge_mode": "row"},
                output_policy=OutputPolicy(output_dir=str(output_dir)),
            )

        first = task_manager.execute_single(_request("merge-finalize-1"))
        second = task_manager.execute_single(_request("merge-finalize-2"))

        assert first.success is True
        assert second.success is True
        first_artifact = first.artifacts[0]
        second_artifact = second.artifacts[0]
        assert Path(first_artifact.staging_path).parent == output_dir
        assert Path(second_artifact.staging_path).parent == output_dir
        assert Path(first_artifact.staging_path).name == "base_merged.xlsx"
        assert Path(second_artifact.staging_path).name == "base_merged_001.xlsx"
        assert first_artifact.suggested_name == "base_merged.xlsx"
        assert second_artifact.suggested_name == "base_merged.xlsx"
        assert _read_used_values(first_artifact.staging_path) == fixture["projects"]["docwen-current"]["row"]["values"]
        assert _read_used_values(second_artifact.staging_path) == fixture["projects"]["docwen-current"]["row"]["values"]
        assert not list(workspace_root.rglob("*.xlsx"))
