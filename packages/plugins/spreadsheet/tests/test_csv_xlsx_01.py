"""Focused tests split from test_csv_xlsx.py."""

from __future__ import annotations

import pytest

from ._csv_xlsx_support import (
    Path,
    _build_fake_context,
    os,
    tempfile,
)

pytestmark = [pytest.mark.golden, pytest.mark.contract]


class TestCsvToXlsx:
    """ROUTE-CSV-XLSX-001: CSV → XLSX conversion."""

    def test_csv_to_xlsx_basic(self, sample_csv_path: Path) -> None:
        """CSV should be converted to a valid XLSX file."""
        from docwen_plugin_spreadsheet.csv_xlsx.converter import CsvToXlsxConverter

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_csv_path), staging, target_format="xlsx")
            result = CsvToXlsxConverter().convert(context)

            assert result.success is True
            assert len(result.artifacts) == 1
            artifact = result.artifacts[0]
            assert artifact.media_type == ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            assert os.path.isfile(artifact.staging_path)

    def test_csv_to_xlsx_content_preserved(self, sample_csv_path: Path) -> None:
        """XLSX output should contain the original CSV data."""
        import openpyxl

        from docwen_plugin_spreadsheet.csv_xlsx.converter import CsvToXlsxConverter

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_csv_path), staging, target_format="xlsx")
            result = CsvToXlsxConverter().convert(context)

            assert result.success
            wb = openpyxl.load_workbook(result.artifacts[0].staging_path)
            ws = wb.active
            assert ws is not None

            # Header row
            assert ws.cell(1, 1).value == "Name"
            assert ws.cell(1, 2).value == "Age"
            assert ws.cell(1, 3).value == "City"

            # Data rows
            assert ws.cell(2, 1).value == "Alice"
            assert ws.cell(2, 2).value == 30  # auto-number conversion
            assert ws.cell(3, 1).value == "Bob"

            wb.close()

    def test_csv_to_xlsx_numeric_values(self, sample_csv_path: Path) -> None:
        """Numeric strings should be converted to numbers."""
        import openpyxl

        from docwen_plugin_spreadsheet.csv_xlsx.converter import CsvToXlsxConverter

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_csv_path), staging, target_format="xlsx")
            result = CsvToXlsxConverter().convert(context)

            assert result.success
            wb = openpyxl.load_workbook(result.artifacts[0].staging_path)
            ws = wb.active
            assert ws is not None

            # Age=30 should be int
            assert isinstance(ws.cell(2, 2).value, int)
            assert ws.cell(2, 2).value == 30
            wb.close()


class TestXlsxToCsv:
    """ROUTE-XLSX-CSV-001: XLSX → CSV conversion."""

    def test_xlsx_to_csv_basic(self, sample_xlsx_path: Path) -> None:
        """XLSX with 2 sheets should produce 2 CSV files."""
        from docwen_plugin_spreadsheet.csv_xlsx.converter import XlsxToCsvConverter

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging, target_format="csv")
            result = XlsxToCsvConverter().convert(context)

            assert result.success is True
            # 2 sheets = 2 CSV artifacts
            assert len(result.artifacts) == 2

            for artifact in result.artifacts:
                assert artifact.media_type == "text/csv"
                assert os.path.isfile(artifact.staging_path)

    def test_xlsx_to_csv_content(self, sample_xlsx_path: Path) -> None:
        """CSV output should contain the sheet data."""
        import csv

        from docwen_plugin_spreadsheet.csv_xlsx.converter import XlsxToCsvConverter

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging, target_format="csv")
            result = XlsxToCsvConverter().convert(context)

            assert result.success

            # First artifact should be from "Sales" sheet
            primary = result.artifacts[0]
            with open(primary.staging_path, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = list(reader)

            assert len(rows) >= 5  # header + 3 data + merged title row
            assert rows[0][0] == "Product"
            assert rows[1][0] == "Alpha"

    def test_xlsx_to_csv_uses_admitted_format_when_suffix_is_wrong(
        self,
        sample_xlsx_path: Path,
        tmp_path: Path,
    ) -> None:
        """The XLSX parser consumes admitted OOXML bytes, not the filename."""
        import csv

        from docwen_plugin_spreadsheet.csv_xlsx.converter import XlsxToCsvConverter

        misleading_path = tmp_path / "admitted-workbook.txt"
        misleading_path.write_bytes(sample_xlsx_path.read_bytes())
        staging = tmp_path / "xlsx-to-csv-staging"
        staging.mkdir()
        context = _build_fake_context(
            str(misleading_path),
            str(staging),
            target_format="csv",
            source_format="xlsx",
        )

        result = XlsxToCsvConverter().convert(context)

        assert result.success is True
        with Path(result.artifacts[0].staging_path).open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
        assert rows[0][0] == "Product"
        assert rows[1][0] == "Alpha"


class TestTsvToXlsx:
    """ROUTE-TSV-XLSX-001: TSV → XLSX conversion."""

    def test_tsv_to_xlsx_basic(self, sample_tsv_path: Path) -> None:
        """TSV should be converted to a valid XLSX."""
        import openpyxl

        from docwen_plugin_spreadsheet.csv_xlsx.converter import TsvToXlsxConverter

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_tsv_path), staging, target_format="xlsx")
            result = TsvToXlsxConverter().convert(context)

            assert result.success is True
            assert len(result.artifacts) == 1

            wb = openpyxl.load_workbook(result.artifacts[0].staging_path)
            ws = wb.active
            assert ws is not None
            assert ws.cell(1, 1).value == "ID"
            assert ws.cell(2, 2).value == 95
            wb.close()


class TestXlsxToTsv:
    """ROUTE-XLSX-TSV-001: XLSX → TSV conversion."""

    def test_xlsx_to_tsv_basic(self, sample_xlsx_path: Path) -> None:
        """XLSX should be converted to TSV files."""
        import csv

        from docwen_plugin_spreadsheet.csv_xlsx.converter import XlsxToTsvConverter

        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(sample_xlsx_path), staging, target_format="tsv")
            result = XlsxToTsvConverter().convert(context)

            assert result.success is True
            assert len(result.artifacts) == 2  # 2 sheets

            primary = result.artifacts[0]
            with open(primary.staging_path, encoding="utf-8-sig") as f:
                reader = csv.reader(f, delimiter="\t")
                rows = list(reader)

            assert rows[0][0] == "Product"
            assert rows[1][0] == "Alpha"

    def test_xlsx_to_tsv_uses_admitted_format_when_suffix_is_wrong(
        self,
        sample_xlsx_path: Path,
        tmp_path: Path,
    ) -> None:
        """A misleading suffix cannot make the TSV route reject admitted XLSX."""
        import csv

        from docwen_plugin_spreadsheet.csv_xlsx.converter import XlsxToTsvConverter

        misleading_path = tmp_path / "admitted-workbook.csv"
        misleading_path.write_bytes(sample_xlsx_path.read_bytes())
        staging = tmp_path / "xlsx-to-tsv-staging"
        staging.mkdir()
        context = _build_fake_context(
            str(misleading_path),
            str(staging),
            target_format="tsv",
            source_format="xlsx",
        )

        result = XlsxToTsvConverter().convert(context)

        assert result.success is True
        with Path(result.artifacts[0].staging_path).open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter="\t"))
        assert rows[0][0] == "Product"
        assert rows[1][0] == "Alpha"
