"""Tests for spreadsheet image extraction and table semantic grid consumption.

Covers the real export path that consumes ``build_table_semantic_grid``
and ``render_table_semantic_grid`` from ``docwen_core.export_semantics``
in the spreadsheet → Markdown converter.

Also covers:
- Embedded image extraction from XLSX workbooks (F-G6-012)
- YAML frontmatter generation (F-I2b-001)
- Image link formatting (F-I2b-004)
- Grid semantics (F-H3b-027, F-H3b-029)
"""

from __future__ import annotations

import io
import os
import struct
import tempfile
from pathlib import Path
from typing import Any

import openpyxl
import openpyxl.drawing.image
import pytest

from docwen_core.markdown_utils import format_sanitized_image_link
from docwen_core.yaml_tools import extract_yaml, generate_basic_yaml_frontmatter


def _make_minimal_png_bytes() -> bytes:
    """Return a valid 1×1 red PNG as raw bytes (68 bytes)."""

    # Adapted from https://www.w3.org/TR/PNG/
    def _chunk(chunk_type: bytes, data: bytes) -> bytes:
        c = chunk_type + data
        crc = struct.pack(">I", _crc32(c))
        return struct.pack(">I", len(data)) + c + crc

    ihdr = struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0)  # 1×1, RGB
    return (
        b"\x89PNG\r\n\x1a\n"
        + _chunk(b"IHDR", ihdr)
        + _chunk(
            b"IDAT",
            b"".join(
                [
                    b"\x78\x9c\x62\x60\x60\x60\x00\x00\x00\x04\x00\x01",
                ]
            ),
        )
        + _chunk(b"IEND", b"")
    )


def _make_png_image() -> openpyxl.drawing.image.Image:
    """Create an openpyxl ``Image`` backed by a minimal PNG."""
    return openpyxl.drawing.image.Image(io.BytesIO(_make_minimal_png_bytes()))


def _crc32(data: bytes) -> int:
    """Compute CRC-32 as used by PNG."""
    c = 0xFFFFFFFF
    for b in data:
        c ^= b
        for _ in range(8):
            if c & 1:
                c = (c >> 1) ^ 0xEDB88320
            else:
                c >>= 1
    return c ^ 0xFFFFFFFF


def _make_wb_with_merge(
    data: list[list[str | int | float | None]],
    merge_ranges: list[str] | None = None,
) -> openpyxl.Workbook:
    """Create a minimal openpyxl workbook with optional merge ranges.

    *data* is a 2‑D list of cell values.  *merge_ranges* are strings like
    ``"A1:B2"``.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for r, row in enumerate(data, start=1):
        for c, value in enumerate(row, start=1):
            if value is not None:
                ws.cell(row=r, column=c, value=value)
    if merge_ranges:
        for mr in merge_ranges:
            ws.merge_cells(mr)
    return wb


def _build_fake_context(
    input_path: str,
    staging_dir: str,
    options: dict[str, Any] | None = None,
) -> Any:
    """Build a fake PluginExecutionContext for direct converter testing."""
    from tests.support.config import FakeConfigView
    from tests.support.execution import FakeExecutionContext
    from tests.support.logging import FakePluginLogger
    from tests.support.progress import FakeProgressSink
    from tests.support.workspace import FakeWorkspaceHandle

    from docwen_core.cancellation import CancellationToken
    from docwen_core.models.file_ref import FileRef
    from docwen_core.models.request import ConversionRequest, OutputPolicy

    token = CancellationToken()
    file_ref = FileRef(
        path=input_path,
        format=Path(input_path).suffix.lstrip("."),
        category="spreadsheet",
    )
    request = ConversionRequest(
        request_id="test-img-001",
        input_refs=[file_ref],
        target_format="md",
        options=options or {},
        output_policy=OutputPolicy(),
    )
    workspace = FakeWorkspaceHandle(input_path, staging_dir)
    progress = FakeProgressSink()
    config = FakeConfigView()
    logger = FakePluginLogger()

    return FakeExecutionContext(
        request=request,
        workspace=workspace,
        config=config,
        progress=progress,
        cancellation=token,
        logger=logger,
    )


def _make_xlsx_with_images(
    tmp_path: Path,
    images_per_sheet: list[int] | None = None,
) -> Path:
    """Create an XLSX file with embedded images.

    *images_per_sheet* controls how many images go into each sheet.
    Default ``[1, 2]`` creates Sheet "Data" with 1 image and Sheet
    "Report" with 2 images.
    """
    if images_per_sheet is None:
        images_per_sheet = [1, 2]

    wb = openpyxl.Workbook()

    for s_idx, count in enumerate(images_per_sheet):
        if s_idx == 0:
            ws = wb.active
            assert ws is not None
            ws.title = "Data"
        else:
            ws = wb.create_sheet(f"Sheet{s_idx + 1}")

        # Add some cell data
        ws.cell(row=1, column=1, value="Header")
        for r in range(2, 5):
            ws.cell(row=r, column=1, value=f"Row{r}")

        # Embed images at different positions
        for i in range(count):
            img = _make_png_image()
            # Position: column B (2), rows spaced apart
            img.anchor = f"B{2 + i * 3}"
            ws.add_image(img)

    output_path = tmp_path / "with_images.xlsx"
    wb.save(str(output_path))
    wb.close()
    return output_path


__all__ = (
    "Path",
    "_build_fake_context",
    "_make_wb_with_merge",
    "_make_xlsx_with_images",
    "extract_yaml",
    "format_sanitized_image_link",
    "generate_basic_yaml_frontmatter",
    "openpyxl",
    "os",
    "pytest",
    "tempfile",
)
