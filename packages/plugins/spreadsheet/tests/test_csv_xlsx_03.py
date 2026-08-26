"""Focused tests split from test_csv_xlsx.py."""

from __future__ import annotations

from typing import ClassVar

from ._csv_xlsx_support import (
    Any,
    Path,
    _add_policy02_structured_formula,
    _build_fake_context,
    _rewrite_policy02_external_formula_as_defined_name,
    _write_policy02_complex_feature_workbook,
    _write_policy02_external_link_workbook,
    _write_policy02_ods,
    _write_policy02_protected_workbook,
    pytest,
    tempfile,
    zipfile,
)

pytestmark = [pytest.mark.golden, pytest.mark.contract]


class TestSmartSheetConverter:
    """ROUTE-SHEETFMT-* routes are backed by the external office bridge."""

    ALL_SMARTSHEET_PAIRS: ClassVar[list[tuple[str, str]]] = [
        ("xlsx", "xls"),
        ("xlsx", "ods"),
        ("xlsx", "et"),
        ("xls", "xlsx"),
        ("ods", "xlsx"),
        ("et", "xlsx"),
        ("xls", "ods"),
        ("xls", "et"),
        ("ods", "xls"),
        ("ods", "et"),
        ("et", "xls"),
        ("et", "ods"),
        ("csv", "xls"),
        ("csv", "ods"),
        ("xls", "csv"),
        ("ods", "csv"),
        ("et", "csv"),
    ]

    def _write_input_for_source(self, tmp_path: Path, source_format: str, sample_xlsx_path: Path) -> Path:
        if source_format == "xlsx":
            return sample_xlsx_path
        input_path = tmp_path / f"smart-source.{source_format}"
        if source_format == "csv":
            input_path.write_text("Name,Value\nAlpha,1\nBeta,2\n", encoding="utf-8")
        else:
            input_path.write_bytes(f"fake {source_format} workbook".encode("ascii"))
        return input_path

    def test_legacy_xls_limit_inspection_counts_cells_outside_biff8_grid(
        self,
        tmp_path: Path,
    ) -> None:
        import openpyxl

        from docwen_plugin_spreadsheet.format_conversion.legacy_xls_limits import (
            inspect_legacy_xls_limits,
        )

        source = tmp_path / "wide-and-tall.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Overflow"
        sheet["IW1"] = "wide"
        sheet["A65537"] = "tall"
        sheet["IW65537"] = "=1+1"
        workbook.save(source)
        workbook.close()

        inspection = inspect_legacy_xls_limits(source)

        assert inspection.out_of_bounds_cell_count == 3
        assert inspection.out_of_bounds_formula_count == 1
        assert inspection.out_of_bounds_row_cell_count == 2
        assert inspection.out_of_bounds_column_cell_count == 2
        assert [(sheet.name, sheet.cell_count, sheet.formula_count) for sheet in inspection.affected_sheets] == [
            ("Overflow", 3, 1)
        ]

    def test_xlsx_to_xls_warns_about_exact_legacy_grid_truncation(
        self,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        import openpyxl

        from docwen_core.office_bridge import BridgeResult
        from docwen_plugin_spreadsheet.format_conversion import converter as sheet_module
        from docwen_plugin_spreadsheet.format_conversion.converter import SmartSheetConverter

        source = tmp_path / "wide.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Wide"
        sheet["IW1"] = "=1+1"
        workbook.save(source)
        workbook.close()

        def fake_convert(input_path: str, output_path: str, **kwargs: Any) -> BridgeResult:
            Path(output_path).write_bytes(b"converted")
            return BridgeResult(True, output_path=output_path, backend="fake-office")

        monkeypatch.setattr(sheet_module, "convert_with_backend_priority", fake_convert)
        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(source), staging, target_format="xls")
            result = SmartSheetConverter().convert(context)

        assert result.success is True
        assert [diagnostic.code for diagnostic in result.diagnostics] == [
            "SHEETFMT-OK",
            "LEGACY_XLS_LIMIT_TRUNCATION",
        ]
        warning = result.diagnostics[1].message
        assert "65,536 rows" in warning
        assert "256 columns" in warning
        assert "1 populated cell" in warning
        assert "1 formula cell" in warning
        assert "Wide" in warning

    def test_xlsx_to_xls_honors_configured_spreadsheet_priority(
        self, sample_xlsx_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from docwen_core.office_bridge import BridgeResult
        from docwen_plugin_spreadsheet.format_conversion import converter as sheet_module
        from docwen_plugin_spreadsheet.format_conversion.converter import SmartSheetConverter

        calls: list[dict[str, Any]] = []

        def fake_convert(input_path: str, output_path: str, **kwargs: Any) -> BridgeResult:
            calls.append(kwargs)
            Path(output_path).write_bytes(b"converted")
            return BridgeResult(True, output_path=output_path, backend="Microsoft Excel")

        monkeypatch.setattr(sheet_module, "convert_with_backend_priority", fake_convert)

        config_snapshot = {
            "software": {
                "default_priority": {"spreadsheet_processors": ["msoffice_excel", "libreoffice", "wps_spreadsheets"]}
            }
        }
        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(
                str(sample_xlsx_path), staging, target_format="xls", config_snapshot=config_snapshot
            )
            result = SmartSheetConverter().convert(context)

        assert result.success is True
        assert calls[0]["backend_priority"] == ["msoffice_excel", "libreoffice", "wps_spreadsheets"]
        assert set(calls[0]["com_candidates"]) == {"wps_spreadsheets", "msoffice_excel"}
        assert calls[0]["com_timeout_s"] == 300.0
        assert calls[0]["libreoffice_timeout_s"] == 300.0

    def test_xlsx_to_ods_honors_configured_ods_priority_without_wps(
        self, sample_xlsx_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from docwen_core.office_bridge import BridgeResult
        from docwen_plugin_spreadsheet.format_conversion import converter as sheet_module
        from docwen_plugin_spreadsheet.format_conversion.converter import SmartSheetConverter

        calls: list[dict[str, Any]] = []

        def fake_convert(input_path: str, output_path: str, **kwargs: Any) -> BridgeResult:
            calls.append(kwargs)
            Path(output_path).write_bytes(b"converted")
            return BridgeResult(True, output_path=output_path, backend="LibreOffice")

        monkeypatch.setattr(sheet_module, "convert_with_backend_priority", fake_convert)

        config_snapshot = {"software": {"special_conversions": {"ods": ["libreoffice", "msoffice_excel"]}}}
        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(
                str(sample_xlsx_path), staging, target_format="ods", config_snapshot=config_snapshot
            )
            result = SmartSheetConverter().convert(context)

        assert result.success is True
        assert calls[0]["backend_priority"] == ["libreoffice", "msoffice_excel"]
        assert set(calls[0]["com_candidates"]) == {"msoffice_excel"}
        assert calls[0]["com_timeout_s"] == 300.0
        assert calls[0]["libreoffice_timeout_s"] == 300.0

    def test_policy02_flattens_external_formula_from_package_cache_without_target_access(self, tmp_path: Path) -> None:
        from docwen_plugin_spreadsheet.format_conversion.xlsx_ods_policy import (
            inspect_xlsx_ods_policy,
            prepare_xlsx_for_ods,
        )

        source = tmp_path / "owner.xlsx"
        prepared = tmp_path / "prepared.xlsx"
        _write_policy02_external_link_workbook(source)
        _add_policy02_structured_formula(source)
        source_before = source.read_bytes()
        assert not (tmp_path / "missing-target.xlsx").exists()

        inspection = inspect_xlsx_ods_policy(source)
        assert inspection.external_formula_cells == ("Sheet!A1",)
        result = prepare_xlsx_for_ods(
            source,
            prepared,
            password=None,
            allow_protection_loss=False,
        )

        assert result.external_links_flattened is True
        assert result.protection_removed is False
        assert source.read_bytes() == source_before
        assert not (tmp_path / "missing-target.xlsx").exists()
        with zipfile.ZipFile(prepared) as package:
            names = set(package.namelist())
            sheet_xml = package.read("xl/worksheets/sheet1.xml").decode("utf-8")
            workbook_xml = package.read("xl/workbook.xml").decode("utf-8")
            rels_xml = package.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            content_types = package.read("[Content_Types].xml").decode("utf-8")
        assert "SUM('[missing-target.xlsx]Sheet0'!A1:B1)" not in sheet_xml
        assert "SUM(ExampleTable[2022])" in sheet_xml
        assert "<v>30</v>" in sheet_xml
        assert "externalReferences" not in workbook_xml
        assert "relationships/externalLink" not in rels_xml
        assert "spreadsheetml.externalLink+xml" not in content_types
        assert not any(name.startswith("xl/externalLinks/") for name in names)

    def test_policy02_removes_unconsumed_external_defined_names_with_explicit_fact(
        self,
        tmp_path: Path,
    ) -> None:
        from docwen_plugin_spreadsheet.format_conversion.xlsx_ods_policy import (
            inspect_xlsx_ods_policy,
            prepare_xlsx_for_ods,
        )

        source = tmp_path / "owner.xlsx"
        prepared = tmp_path / "prepared.xlsx"
        _write_policy02_external_link_workbook(source)
        _rewrite_policy02_external_formula_as_defined_name(source)
        source_before = source.read_bytes()

        inspection = inspect_xlsx_ods_policy(source)
        assert inspection.external_formula_cells == ()
        assert inspection.external_defined_names == ("LegacyExternalName",)
        assert inspection.unsupported_external_references == ()

        result = prepare_xlsx_for_ods(
            source,
            prepared,
            password=None,
            allow_protection_loss=False,
        )

        assert result.external_links_flattened is True
        assert result.flattened_cached_values == ()
        assert result.removed_external_defined_names == ("LegacyExternalName",)
        assert source.read_bytes() == source_before
        with zipfile.ZipFile(prepared) as package:
            names = set(package.namelist())
            workbook_xml = package.read("xl/workbook.xml").decode("utf-8")
        assert "LegacyExternalName" not in workbook_xml
        assert "externalReferences" not in workbook_xml
        assert not any(name.startswith("xl/externalLinks/") for name in names)

    def test_policy02_does_not_misclassify_structured_table_references_as_external(
        self,
        tmp_path: Path,
    ) -> None:
        import openpyxl

        from docwen_plugin_spreadsheet.format_conversion.xlsx_ods_policy import inspect_xlsx_ods_policy

        source = tmp_path / "structured-reference.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet["A1"] = "=SUM(ExampleTable[2022])+SUM(ExampleTable[[#This Row],[2023]:[2026]])"
        workbook.save(source)
        workbook.close()

        inspection = inspect_xlsx_ods_policy(source)

        assert inspection.external_formula_cells == ()
        assert inspection.external_link_parts_present is False

    def test_policy02_records_complex_feature_fidelity_risks_without_rejecting(
        self,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from docwen_core.office_bridge import BridgeResult
        from docwen_plugin_spreadsheet.format_conversion import converter as sheet_module
        from docwen_plugin_spreadsheet.format_conversion.converter import SmartSheetConverter
        from docwen_plugin_spreadsheet.format_conversion.xlsx_ods_policy import inspect_xlsx_ods_policy

        source = tmp_path / "complex.xlsx"
        _write_policy02_complex_feature_workbook(source)
        source_before = source.read_bytes()

        inspection = inspect_xlsx_ods_policy(source)
        risk_counts = dict(inspection.fidelity_risk_counts)
        assert risk_counts == {
            "data_validations": 1,
            "conditional_formatting_ranges": 1,
            "charts": 1,
            "drawings": 1,
            "tables": 1,
            "defined_names": 1,
        }

        def fake_convert(input_path: str, output_path: str, **kwargs: Any) -> BridgeResult:
            _write_policy02_ods(Path(output_path), value="1")
            return BridgeResult(True, output_path=output_path, backend="fake-office")

        monkeypatch.setattr(sheet_module, "convert_with_backend_priority", fake_convert)
        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(source), staging, target_format="ods")
            result = SmartSheetConverter().convert(context)

        assert result.success is True
        assert source.read_bytes() == source_before
        assert [diagnostic.code for diagnostic in result.diagnostics] == [
            "SHEETFMT-OK",
            "ODS_FEATURE_FIDELITY_RISK",
        ]
        warning = result.diagnostics[-1].message
        assert "best-effort result" in warning
        assert "materialize hidden cache worksheets" in warning
        assert "original source was not changed" in warning
        assert result.artifacts[0].metadata["ods_fidelity_risk_counts"] == risk_counts
        assert result.metrics.extra["ods_fidelity_risk_counts"] == risk_counts

    @pytest.mark.parametrize(
        ("scope", "password", "expected_element"),
        [
            ("workbook", "test", "workbook"),
            ("sheet", "pwd", "Sheet"),
        ],
    )
    def test_policy02_validates_modern_password_and_removes_only_passworded_protection(
        self,
        tmp_path: Path,
        scope: str,
        password: str,
        expected_element: str,
    ) -> None:
        from docwen_plugin_spreadsheet.format_conversion.xlsx_ods_policy import (
            inspect_xlsx_ods_policy,
            prepare_xlsx_for_ods,
        )

        source = tmp_path / f"{scope}.xlsx"
        prepared = tmp_path / f"{scope}-prepared.xlsx"
        _write_policy02_protected_workbook(source, scope=scope, password=password)
        source_before = source.read_bytes()
        inspection = inspect_xlsx_ods_policy(source)
        assert inspection.password_protected_elements == (expected_element,)

        result = prepare_xlsx_for_ods(
            source,
            prepared,
            password=password,
            allow_protection_loss=True,
        )

        assert result.external_links_flattened is False
        assert result.protection_removed is True
        assert source.read_bytes() == source_before
        after = inspect_xlsx_ods_policy(prepared)
        assert after.password_protected_elements == ()
        assert after.unpassworded_protected_elements == ()

    @pytest.mark.parametrize(
        ("options", "expected_code"),
        [
            ({}, "PROTECTION_PASSWORD_REQUIRED"),
            (
                {"spreadsheet_password": "wrong", "allow_spreadsheet_protection_loss": True},
                "PROTECTION_PASSWORD_INVALID",
            ),
            ({"spreadsheet_password": "test"}, "PROTECTION_LOSS_CONSENT_REQUIRED"),
        ],
    )
    def test_policy02_rejects_password_admission_failures_before_backend(
        self,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
        options: dict[str, Any],
        expected_code: str,
    ) -> None:
        from docwen_plugin_spreadsheet.format_conversion import converter as sheet_module
        from docwen_plugin_spreadsheet.format_conversion.converter import SmartSheetConverter

        source = tmp_path / "protected.xlsx"
        _write_policy02_protected_workbook(source, scope="workbook", password="test")
        bridge_calls: list[str] = []

        def forbidden_bridge(input_path: str, output_path: str, **kwargs: Any) -> Any:
            bridge_calls.append(input_path)
            raise AssertionError("policy admission failure must not call a backend")

        monkeypatch.setattr(sheet_module, "convert_with_backend_priority", forbidden_bridge)
        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(source), staging, target_format="ods", options=options)
            result = SmartSheetConverter().convert(context)

        assert result.success is False
        assert result.error is not None
        assert result.error.error_type == "invalid_input"
        assert result.error.diagnostic_code == expected_code
        assert [diagnostic.code for diagnostic in result.diagnostics] == [expected_code]
        assert bridge_calls == []
        surfaced_text = "\n".join([result.error.message, *(diagnostic.message for diagnostic in result.diagnostics)])
        if password := options.get("spreadsheet_password"):
            assert password not in surfaced_text

    def test_policy02_converter_passes_only_private_transformed_copy_and_warns(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from docwen_core.office_bridge import BridgeResult
        from docwen_plugin_spreadsheet.format_conversion import converter as sheet_module
        from docwen_plugin_spreadsheet.format_conversion.converter import SmartSheetConverter

        source = tmp_path / "owner.xlsx"
        _write_policy02_external_link_workbook(source)
        source_before = source.read_bytes()
        bridge_inputs: list[Path] = []

        def fake_convert(input_path: str, output_path: str, **kwargs: Any) -> BridgeResult:
            bridge_input = Path(input_path)
            bridge_inputs.append(bridge_input)
            assert bridge_input != source
            with zipfile.ZipFile(bridge_input) as package:
                sheet_xml = package.read("xl/worksheets/sheet1.xml").decode("utf-8")
                assert "<f>" not in sheet_xml
                assert "<v>30</v>" in sheet_xml
                assert not any(name.startswith("xl/externalLinks/") for name in package.namelist())
            _write_policy02_ods(Path(output_path), value="30")
            return BridgeResult(True, output_path=output_path, backend="fake-office")

        monkeypatch.setattr(sheet_module, "convert_with_backend_priority", fake_convert)
        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(source), staging, target_format="ods")
            result = SmartSheetConverter().convert(context)

        assert result.success is True
        assert source.read_bytes() == source_before
        assert len(bridge_inputs) == 1
        assert [diagnostic.code for diagnostic in result.diagnostics] == [
            "SHEETFMT-OK",
            "EXTERNAL_LINK_FLATTENED",
        ]

    def test_policy02_converter_warns_when_external_defined_names_are_removed(
        self,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from docwen_core.office_bridge import BridgeResult
        from docwen_plugin_spreadsheet.format_conversion import converter as sheet_module
        from docwen_plugin_spreadsheet.format_conversion.converter import SmartSheetConverter

        source = tmp_path / "owner.xlsx"
        _write_policy02_external_link_workbook(source)
        _rewrite_policy02_external_formula_as_defined_name(source)

        def fake_convert(input_path: str, output_path: str, **kwargs: Any) -> BridgeResult:
            with zipfile.ZipFile(input_path) as package:
                workbook_xml = package.read("xl/workbook.xml").decode("utf-8")
                assert "LegacyExternalName" not in workbook_xml
            _write_policy02_ods(Path(output_path), value="30")
            return BridgeResult(True, output_path=output_path, backend="fake-office")

        monkeypatch.setattr(sheet_module, "convert_with_backend_priority", fake_convert)
        with tempfile.TemporaryDirectory() as staging:
            context = _build_fake_context(str(source), staging, target_format="ods")
            result = SmartSheetConverter().convert(context)

        assert result.success is True
        warning = next(item for item in result.diagnostics if item.code == "EXTERNAL_LINK_FLATTENED")
        assert "1 external defined name without a safe cached value was removed" in warning.message
        assert result.artifacts[0].metadata["removed_external_defined_names"] == 1
        assert result.metrics.extra["removed_external_defined_names"] == 1
