"""Tests for CSV/TSV ↔ XLSX interconversion routes."""

from __future__ import annotations

import json
import os
import re
import tempfile
import zipfile
from base64 import b64decode
from hashlib import sha512
from pathlib import Path
from struct import pack
from typing import Any

import pytest

pytestmark = [pytest.mark.golden, pytest.mark.contract]

_PROJECT_ROOT = Path(__file__).resolve().parents[4]

_MERGE_TABLES_OLD_SYSTEM_FIXTURE = (
    _PROJECT_ROOT / "tests" / "fixtures" / "golden" / "old_system_merge_tables_semantics.json"
)

_MERGE_TABLES_BROADER_OLD_SYSTEM_FIXTURE = (
    _PROJECT_ROOT / "tests" / "fixtures" / "golden" / "old_system_merge_tables_broader_workbook_semantics.json"
)


def _load_merge_tables_old_system_fixture() -> dict[str, Any]:
    return json.loads(_MERGE_TABLES_OLD_SYSTEM_FIXTURE.read_text(encoding="utf-8"))


def _load_merge_tables_broader_old_system_fixture() -> dict[str, Any]:
    return json.loads(_MERGE_TABLES_BROADER_OLD_SYSTEM_FIXTURE.read_text(encoding="utf-8"))


def _write_workbook(path: Path, rows: list[list[Any]]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet"
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _write_policy02_external_link_workbook(path: Path) -> None:
    """Create a self-contained XLSX with an external formula and cached value."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "=SUM('[missing-target.xlsx]Sheet0'!A1:B1)"
    wb.save(path)
    wb.close()

    with zipfile.ZipFile(path) as source:
        parts = {info.filename: (info, source.read(info.filename)) for info in source.infolist()}

    sheet_name = "xl/worksheets/sheet1.xml"
    sheet_xml = parts[sheet_name][1].decode("utf-8")
    sheet_xml, count = re.subn(
        r"<f>SUM\('\[missing-target\.xlsx\]Sheet0'!A1:B1\)</f><v></v>",
        "<f>SUM('[missing-target.xlsx]Sheet0'!A1:B1)</f><v>30</v>",
        sheet_xml,
    )
    assert count == 1
    parts[sheet_name] = (parts[sheet_name][0], sheet_xml.encode("utf-8"))

    workbook_name = "xl/workbook.xml"
    workbook_xml = parts[workbook_name][1].decode("utf-8")
    workbook_xml = workbook_xml.replace(
        "</workbook>",
        '<externalReferences xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<externalReference r:id="rIdPolicy02"/></externalReferences></workbook>',
    )
    parts[workbook_name] = (parts[workbook_name][0], workbook_xml.encode("utf-8"))

    rels_name = "xl/_rels/workbook.xml.rels"
    rels_xml = parts[rels_name][1].decode("utf-8")
    rels_xml = rels_xml.replace(
        "</Relationships>",
        '<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/'
        'externalLink" Target="externalLinks/externalLink1.xml" Id="rIdPolicy02"/></Relationships>',
    )
    parts[rels_name] = (parts[rels_name][0], rels_xml.encode("utf-8"))

    content_types_name = "[Content_Types].xml"
    content_types_xml = parts[content_types_name][1].decode("utf-8")
    content_types_xml = content_types_xml.replace(
        "</Types>",
        '<Override PartName="/xl/externalLinks/externalLink1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"/>'
        "</Types>",
    )
    parts[content_types_name] = (parts[content_types_name][0], content_types_xml.encode("utf-8"))

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for info, payload in parts.values():
            target.writestr(info, payload)
        target.writestr(
            "xl/externalLinks/externalLink1.xml",
            (
                '<?xml version="1.0" encoding="UTF-8"?>'
                '<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                '<externalBook r:id="rId1"/></externalLink>'
            ),
        )
        target.writestr(
            "xl/externalLinks/_rels/externalLink1.xml.rels",
            (
                '<?xml version="1.0" encoding="UTF-8"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                'relationships/externalLinkPath" Target="missing-target.xlsx" TargetMode="External"/>'
                "</Relationships>"
            ),
        )


def _rewrite_policy02_external_formula_as_defined_name(path: Path) -> None:
    """Replace the cached external cell formula with an unconsumed external name."""

    with zipfile.ZipFile(path) as source:
        parts = {info.filename: (info, source.read(info.filename)) for info in source.infolist()}

    sheet_name = "xl/worksheets/sheet1.xml"
    sheet_xml = parts[sheet_name][1].decode("utf-8")
    sheet_xml, count = re.subn(
        r"<f>SUM\('\[missing-target\.xlsx\]Sheet0'!A1:B1\)</f><v>30</v>",
        "<v>30</v>",
        sheet_xml,
    )
    assert count == 1
    parts[sheet_name] = (parts[sheet_name][0], sheet_xml.encode("utf-8"))

    workbook_name = "xl/workbook.xml"
    workbook_xml = parts[workbook_name][1].decode("utf-8")
    workbook_xml = workbook_xml.replace(
        "<externalReferences",
        (
            '<definedNames><definedName name="LegacyExternalName">'
            "'[1]Missing'!#REF!"
            "</definedName></definedNames><externalReferences"
        ),
    )
    parts[workbook_name] = (parts[workbook_name][0], workbook_xml.encode("utf-8"))

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for info, payload in parts.values():
            target.writestr(info, payload)


def _add_policy02_structured_formula(path: Path) -> None:
    with zipfile.ZipFile(path) as source:
        parts = {info.filename: (info, source.read(info.filename)) for info in source.infolist()}

    sheet_name = "xl/worksheets/sheet1.xml"
    sheet_xml = parts[sheet_name][1].decode("utf-8")
    sheet_xml, count = re.subn(
        r"</row>",
        ('<c r="B1"><f>SUM(ExampleTable[2022])+SUM(ExampleTable[[#This Row],[2023]:[2026]])</f><v>1</v></c></row>'),
        sheet_xml,
        count=1,
    )
    assert count == 1
    parts[sheet_name] = (parts[sheet_name][0], sheet_xml.encode("utf-8"))

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for info, payload in parts.values():
            target.writestr(info, payload)


def _write_policy02_complex_feature_workbook(path: Path) -> None:
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["Category", "Value"])
    sheet.append(["A", 1])
    sheet.append(["B", 2])

    validation = DataValidation(type="list", formula1='"A,B"')
    sheet.add_data_validation(validation)
    validation.add(sheet["A2"])
    sheet.conditional_formatting.add("B2:B3", CellIsRule(operator="greaterThan", formula=["1"]))
    sheet.add_table(Table(displayName="RiskTable", ref="A1:B3"))

    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    chart.anchor = "D1"
    sheet.add_chart(chart)

    workbook.defined_names.add(DefinedName("RiskRange", attr_text="'Sheet'!$A$1:$B$3"))
    workbook.save(path)
    workbook.close()


def _policy02_hash(password: str, salt_b64: str, spin_count: int = 100_000) -> str:
    import base64

    digest = sha512(b64decode(salt_b64) + password.encode("utf-16le")).digest()
    for index in range(spin_count):
        digest = sha512(digest + pack("<I", index)).digest()
    return base64.b64encode(digest).decode("ascii")


def _write_policy02_protected_workbook(path: Path, *, scope: str, password: str | None) -> None:
    import openpyxl
    from openpyxl.workbook.protection import WorkbookProtection

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "POLICY-02"
    if scope == "workbook":
        assert password is not None
        salt = "Wq5e2oy8ZLa/369T8z/Jaw=="
        wb.security = WorkbookProtection(
            lockStructure=True,
            workbookAlgorithmName="SHA-512",
            workbookHashValue=_policy02_hash(password, salt),
            workbookSaltValue=salt,
            workbookSpinCount=100_000,
        )
    elif password is None:
        ws.protection.sheet = True
        ws.protection.objects = True
        ws.protection.scenarios = True
    else:
        salt = "R040EdN/Ec7il6MJ8JrRLQ=="
        ws.protection.sheet = True
        ws.protection.objects = True
        ws.protection.scenarios = True
        ws.protection.algorithmName = "SHA-512"
        ws.protection.hashValue = _policy02_hash(password, salt)
        ws.protection.saltValue = salt
        ws.protection.spinCount = 100_000
    wb.save(path)
    wb.close()


def _add_policy02_workbook_protection(path: Path, password: str) -> None:
    salt = "Wq5e2oy8ZLa/369T8z/Jaw=="
    with zipfile.ZipFile(path) as source:
        parts = {info.filename: (info, source.read(info.filename)) for info in source.infolist()}
    workbook_name = "xl/workbook.xml"
    workbook_xml = parts[workbook_name][1].decode("utf-8")
    protection = (
        '<workbookProtection workbookAlgorithmName="SHA-512" '
        f'workbookHashValue="{_policy02_hash(password, salt)}" '
        f'workbookSaltValue="{salt}" workbookSpinCount="100000" lockStructure="1"/>'
    )
    workbook_xml = workbook_xml.replace("<bookViews>", f"{protection}<bookViews>")
    parts[workbook_name] = (parts[workbook_name][0], workbook_xml.encode("utf-8"))
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for info, payload in parts.values():
            target.writestr(info, payload)


def _write_policy02_ods(path: Path, *, value: str | None = None, protected_sheet: str | None = None) -> None:
    value_xml = ""
    if value is not None:
        value_xml = (
            f'<table:table-cell office:value-type="float" office:value="{value}">'
            f"<text:p>{value}</text:p></table:table-cell>"
        )
    protected_xml = ' table:protected="true"' if protected_sheet else ""
    sheet_name = protected_sheet or "Sheet"
    content = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<office:document-content xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0" '
        'xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0" '
        'xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0">'
        f'<office:body><office:spreadsheet><table:table table:name="{sheet_name}"{protected_xml}>'
        f"<table:table-row>{value_xml}</table:table-row>"
        "</table:table></office:spreadsheet></office:body></office:document-content>"
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as package:
        package.writestr("mimetype", "application/vnd.oasis.opendocument.spreadsheet")
        package.writestr("content.xml", content)


def _write_multisheet_marker_workbook(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    first = wb.active
    assert first is not None
    first.title = "Alpha Data"
    first.append(["ALPHA_MARKER", 11])
    second = wb.create_sheet("Beta-Summary")
    second.append(["BETA_MARKER", 22])
    hidden = wb.create_sheet("Hidden Sheet")
    hidden.sheet_state = "hidden"
    hidden.append(["HIDDEN_MARKER", 33])
    wb.save(path)
    wb.close()


def _write_broader_merge_workbook(path: Path, spec: dict[str, Any]) -> None:
    import openpyxl
    from openpyxl.cell import Cell
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = spec["active_sheet"]
    for cell_ref, value in spec["cells"].items():
        ws[cell_ref] = value
    for merged_range in spec.get("merged_ranges", []):
        ws.merge_cells(merged_range)
    for cell_ref, style in spec.get("styles", {}).items():
        cell = ws[cell_ref]
        assert isinstance(cell, Cell)
        if fill := style.get("fill"):
            cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(
            bold=bool(style.get("bold", False)),
            italic=bool(style.get("italic", False)),
        )
    ws.protection.sheet = bool(spec.get("sheet_protection", False))
    for sheet_name, rows in spec.get("extra_sheets", {}).items():
        extra = wb.create_sheet(sheet_name)
        for row in rows:
            extra.append(row)
    wb.save(path)
    wb.close()


def _write_cached_formula_merge_workbook(path: Path, spec: dict[str, Any]) -> None:
    import re
    import zipfile

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = spec["worksheet"]
    for row in spec["rows"]:
        ws.append(row)
    wb.save(path)
    wb.close()

    extracted = path.parent / f"{path.stem}_xlsx_parts"
    with zipfile.ZipFile(path, "r") as zip_in:
        zip_in.extractall(extracted)
    sheet_xml_path = extracted / "xl" / "worksheets" / "sheet1.xml"
    sheet_xml = sheet_xml_path.read_text(encoding="utf-8")
    formula_cell = spec["formula_cell"]
    formula_body = spec["formula_text"].lstrip("=")
    formula_pattern = re.escape(formula_body)

    def _replacement(match: re.Match[str]) -> str:
        return f'<c r="{formula_cell}"{match.group(1)}><f>{formula_body}</f><v>{spec["cached_value"]}</v></c>'

    sheet_xml, count = re.subn(
        rf'<c r="{formula_cell}"([^>]*)><f>{formula_pattern}</f><v></v></c>',
        _replacement,
        sheet_xml,
    )
    assert count == 1
    sheet_xml_path.write_text(sheet_xml, encoding="utf-8")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zip_out:
        for file in extracted.rglob("*"):
            if file.is_file():
                zip_out.write(file, file.relative_to(extracted).as_posix())


def _read_used_values(path: str | Path) -> list[list[Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    assert ws is not None
    values: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        row_values = list(row)
        if any(value is not None for value in row_values):
            values.append(row_values)
    wb.close()
    return values


def _read_broader_merge_projection(path: str | Path) -> dict[str, Any]:
    import openpyxl

    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    try:
        ws = wb_values.active
        formulas_ws = wb_formulas.active
        assert ws is not None
        assert formulas_ws is not None
        active_values = [[ws.cell(row_idx, col_idx).value for col_idx in range(1, 6)] for row_idx in range(1, 6)]
        active_formulas = [
            [formulas_ws.cell(row_idx, col_idx).value for col_idx in range(1, 6)] for row_idx in range(1, 6)
        ]
        sheets: dict[str, list[list[Any]]] = {}
        for sheet_name in wb_values.sheetnames:
            sheet = wb_values[sheet_name]
            sheets[sheet_name] = [
                [sheet.cell(row_idx, col_idx).value for col_idx in range(1, min(sheet.max_column, 3) + 1)]
                for row_idx in range(1, min(sheet.max_row, 3) + 1)
            ]
        return {
            "sheetnames": wb_values.sheetnames,
            "active_values": active_values,
            "active_formulas": active_formulas,
            "merged_ranges": [str(merged_range) for merged_range in ws.merged_cells.ranges],
            "sheet_protection": bool(ws.protection.sheet),
            "a1_fill": ws["A1"].fill.fgColor.rgb,
            "a1_bold": bool(ws["A1"].font.bold),
            "b1_fill": ws["B1"].fill.fgColor.rgb,
            "b1_bold": bool(ws["B1"].font.bold),
            "sheets": sheets,
        }
    finally:
        wb_values.close()
        wb_formulas.close()


def _read_cached_formula_merge_projection(path: str | Path) -> dict[str, Any]:
    import openpyxl

    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    try:
        ws = wb_values.active
        formulas_ws = wb_formulas.active
        assert ws is not None
        assert formulas_ws is not None
        active_values = [[ws.cell(row_idx, col_idx).value for col_idx in range(1, 3)] for row_idx in range(1, 5)]
        active_formulas = [
            [formulas_ws.cell(row_idx, col_idx).value for col_idx in range(1, 3)] for row_idx in range(1, 5)
        ]
        return {
            "active_values": active_values,
            "active_formulas": active_formulas,
            "contains_formula_text": any(value == "=SUM(B2:B3)" for row in active_formulas for value in row),
            "contains_cached_value_sum": active_values[3][1] == 20,
        }
    finally:
        wb_values.close()
        wb_formulas.close()


def _build_fake_context(
    input_path: str,
    staging_dir: str,
    target_format: str = "xlsx",
    options: dict[str, Any] | None = None,
    config_snapshot: dict[str, Any] | None = None,
    *,
    source_format: str | None = None,
) -> Any:
    """Build a fake PluginExecutionContext."""
    from tests.support.config import FakeConfigView
    from tests.support.execution import FakeExecutionContext
    from tests.support.logging import FakePluginLogger
    from tests.support.progress import FakeProgressSink
    from tests.support.workspace import FakeWorkspaceHandle

    from docwen_core.cancellation import CancellationToken
    from docwen_core.models.file_ref import FileRef
    from docwen_core.models.request import ConversionRequest, OutputPolicy

    token = CancellationToken()
    file_ref = FileRef(
        path=input_path,
        format=source_format or Path(input_path).suffix.lstrip("."),
        category="spreadsheet",
    )
    request = ConversionRequest(
        request_id="test-csv-001",
        input_refs=[file_ref],
        target_format=target_format,
        options=options or {},
        output_policy=OutputPolicy(),
        config_snapshot=config_snapshot or {},
    )
    workspace = FakeWorkspaceHandle(input_path, staging_dir)
    progress = FakeProgressSink()
    config = FakeConfigView(config_snapshot)

    return FakeExecutionContext(
        request=request,
        workspace=workspace,
        config=config,
        progress=progress,
        cancellation=token,
        logger=FakePluginLogger(),
    )


__all__ = (
    "Any",
    "Path",
    "_add_policy02_structured_formula",
    "_add_policy02_workbook_protection",
    "_build_fake_context",
    "_load_merge_tables_broader_old_system_fixture",
    "_load_merge_tables_old_system_fixture",
    "_read_broader_merge_projection",
    "_read_cached_formula_merge_projection",
    "_read_used_values",
    "_rewrite_policy02_external_formula_as_defined_name",
    "_write_broader_merge_workbook",
    "_write_cached_formula_merge_workbook",
    "_write_multisheet_marker_workbook",
    "_write_policy02_complex_feature_workbook",
    "_write_policy02_external_link_workbook",
    "_write_policy02_ods",
    "_write_policy02_protected_workbook",
    "_write_workbook",
    "os",
    "pytest",
    "pytestmark",
    "tempfile",
    "zipfile",
)
