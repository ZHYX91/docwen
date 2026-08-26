"""Shared fixtures for spreadsheet plugin tests."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[4]

LOCAL_SRC_PATHS = [
    PROJECT_ROOT,
    PROJECT_ROOT / "packages" / "core" / "src",
    PROJECT_ROOT / "packages" / "runtime" / "src",
    PROJECT_ROOT / "packages" / "plugins" / "spreadsheet" / "src",
]

for path in reversed(LOCAL_SRC_PATHS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))


@pytest.fixture
def sample_xlsx_path(tmp_path: Path) -> Path:
    """Create a sample XLSX file with known content for golden testing.

    Contains:
    - Sheet 1 ("Sales"): header row + 3 data rows
    - Sheet 2 ("Summary"): header row + 2 data rows
    - A merged cell region in Sheet 1

    Returns the path to the created XLSX file.
    """
    import openpyxl

    wb = openpyxl.Workbook()

    # Sheet 1: Sales
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Sales"

    # Header
    ws1.cell(row=1, column=1, value="Product")
    ws1.cell(row=1, column=2, value="Quantity")
    ws1.cell(row=1, column=3, value="Price")
    ws1.cell(row=1, column=4, value="Total")

    # Data rows
    ws1.cell(row=2, column=1, value="Alpha")
    ws1.cell(row=2, column=2, value=10)
    ws1.cell(row=2, column=3, value=9.99)
    ws1.cell(row=2, column=4, value=99.90)

    ws1.cell(row=3, column=1, value="Beta")
    ws1.cell(row=3, column=2, value=5)
    ws1.cell(row=3, column=3, value=19.50)
    ws1.cell(row=3, column=4, value=97.50)

    ws1.cell(row=4, column=1, value="Gamma")
    ws1.cell(row=4, column=2, value=20)
    ws1.cell(row=4, column=3, value=4.50)
    ws1.cell(row=4, column=4, value=90.00)

    # Merged cell (title row spanning columns 1-4, below data, row 6)
    ws1.merge_cells("A6:D6")
    ws1.cell(row=6, column=1, value="Q1 Sales Summary")

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Metric")
    ws2.cell(row=1, column=2, value="Value")

    ws2.cell(row=2, column=1, value="Total Revenue")
    ws2.cell(row=2, column=2, value=287.40)

    ws2.cell(row=3, column=1, value="Total Units")
    ws2.cell(row=3, column=2, value=35)

    # Save
    output_path = tmp_path / "golden_test.xlsx"
    wb.save(str(output_path))
    wb.close()

    return output_path


@pytest.fixture
def sample_csv_path(tmp_path: Path) -> Path:
    """Create a sample CSV file with known content.

    Contains:
    - Header row: Name, Age, City
    - 3 data rows

    Returns the path to the created CSV file.
    """
    import csv

    output_path = tmp_path / "test_data.csv"
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Age", "City"])
        writer.writerow(["Alice", "30", "New York"])
        writer.writerow(["Bob", "25", "London"])
        writer.writerow(["Charlie", "35", "Tokyo"])

    return output_path


@pytest.fixture
def sample_tsv_path(tmp_path: Path) -> Path:
    """Create a sample TSV file with known content.

    Returns the path to the created TSV file.
    """
    import csv

    output_path = tmp_path / "test_data.tsv"
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(["ID", "Score", "Grade"])
        writer.writerow(["1", "95", "A"])
        writer.writerow(["2", "87", "B"])
        writer.writerow(["3", "73", "C"])

    return output_path


@pytest.fixture
def sample_base_xlsx_path(tmp_path: Path) -> Path:
    """Create a base XLSX for table merge testing."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Base"

    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=2, value="Name")
    ws.cell(row=1, column=3, value="Score")

    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="Alice")
    ws.cell(row=2, column=3, value=90)

    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=2, value="Bob")
    ws.cell(row=3, column=3, value=85)

    output_path = tmp_path / "base.xlsx"
    wb.save(str(output_path))
    wb.close()
    return output_path


@pytest.fixture
def sample_collect_xlsx_path(tmp_path: Path) -> Path:
    """Create a collection XLSX for table merge testing."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Collect"

    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=2, value="Name")
    ws.cell(row=1, column=3, value="Score")

    ws.cell(row=2, column=1, value=3)
    ws.cell(row=2, column=2, value="Charlie")
    ws.cell(row=2, column=3, value=78)

    output_path = tmp_path / "collect.xlsx"
    wb.save(str(output_path))
    wb.close()
    return output_path
