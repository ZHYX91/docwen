"""Focused tests split from test_gui_e2e_conversion.py."""

from __future__ import annotations

from ._gui_e2e_conversion_support import (
    _E2E_CONVERSION_TIMEOUT_MS,
    BatchFileEntry,
    Path,
    _wait_for,
    pytest,
)


@pytest.mark.gui
class TestAggregateGuiExecution:
    @staticmethod
    def _wait_for_entries_finished(window, paths: list[str]) -> None:
        from docwen_gui.main_window import _normalize_path

        normalized = [_normalize_path(path) for path in paths]

        def _finished() -> bool:
            entries = [window._batch_list_vm.get_file_entry(path) for path in normalized]
            if any(entry is None for entry in entries):
                return False
            return all(entry.status in ("completed", "failed") for entry in entries if entry is not None)

        assert _wait_for(_finished, timeout_ms=_E2E_CONVERSION_TIMEOUT_MS)

    @staticmethod
    def _completed_entries(window, paths: list[str]) -> list[BatchFileEntry]:
        from docwen_gui.main_window import _normalize_path

        entries = [window._batch_list_vm.get_file_entry(_normalize_path(path)) for path in paths]
        assert all(entry is not None and entry.status == "completed" for entry in entries)
        return [entry for entry in entries if entry is not None]

    def test_merge_pdfs_runs_through_gui_thread_and_places_output(
        self, main_window_with_controller, tmp_path: Path
    ) -> None:
        import fitz
        from PySide6.QtWidgets import QApplication

        first = tmp_path / "first.pdf"
        second = tmp_path / "second.pdf"
        for path, label in ((first, "first"), (second, "second")):
            doc = fitz.open()
            page = doc.new_page(width=240, height=160)
            page.insert_text((48, 80), label)
            doc.save(path)
            doc.close()

        window = main_window_with_controller
        vm = window.view_model
        vm.add_files([str(first), str(second)])

        app = QApplication.instance()
        if app is not None:
            app.processEvents()

        assert window._conversion_panel._merge_pdfs_button is not None
        window._conversion_panel._merge_pdfs_button.click()

        if app is not None:
            app.processEvents()

        self._wait_for_entries_finished(window, [str(first), str(second)])
        entries = self._completed_entries(window, [str(first), str(second)])

        output_paths = {entry.output_path for entry in entries if entry.output_path is not None}
        assert len(output_paths) == 1
        output_path = Path(next(iter(output_paths)))
        assert output_path.name == "merged.pdf"
        assert output_path.exists()

        merged = fitz.open(output_path)
        try:
            assert merged.page_count == 2
        finally:
            merged.close()

    def test_merge_pdfs_uses_custom_output_directory_and_date_subfolder(
        self, main_window_with_controller, tmp_path: Path
    ) -> None:
        import re

        import fitz
        from PySide6.QtWidgets import QApplication

        first = tmp_path / "first.pdf"
        second = tmp_path / "second.pdf"
        for path, label in ((first, "first"), (second, "second")):
            doc = fitz.open()
            page = doc.new_page(width=240, height=160)
            page.insert_text((48, 80), label)
            doc.save(path)
            doc.close()

        window = main_window_with_controller
        controller = window.view_model.controller
        assert controller is not None
        assert controller.config_port is not None

        output_dir = tmp_path / "aggregate_exports"
        assert controller.config_port.set("output.directory.mode", "custom")
        assert controller.config_port.set("output.directory.custom_path", str(output_dir))
        assert controller.config_port.set("output.directory.create_date_subfolder", True)
        assert controller.config_port.set("output.directory.date_folder_format", "%Y-%m-%d")

        window.view_model.add_files([str(first), str(second)])

        app = QApplication.instance()
        if app is not None:
            app.processEvents()

        assert window._conversion_panel._merge_pdfs_button is not None
        window._conversion_panel._merge_pdfs_button.click()

        if app is not None:
            app.processEvents()

        self._wait_for_entries_finished(window, [str(first), str(second)])
        entries = self._completed_entries(window, [str(first), str(second)])

        output_paths = {entry.output_path for entry in entries if entry.output_path is not None}
        assert len(output_paths) == 1
        output_path = Path(next(iter(output_paths)))
        assert output_path.name == "merged.pdf"
        assert output_path.exists()
        assert output_path.parent.parent == output_dir
        assert re.fullmatch(r"\d{4}-\d{2}-\d{2}", output_path.parent.name)

        merged = fitz.open(output_path)
        try:
            assert merged.page_count == 2
        finally:
            merged.close()

    def test_merge_tables_runs_through_gui_thread_and_places_output(
        self, main_window_with_controller, tmp_path: Path
    ) -> None:
        from openpyxl import Workbook, load_workbook
        from PySide6.QtWidgets import QApplication

        first = tmp_path / "base.xlsx"
        second = tmp_path / "collect.xlsx"
        for path, rows in (
            (first, [["Name", "Score"], ["Alice", 90], ["Bob", 80]]),
            (second, [["Name", "Score"], ["Charlie", 78]]),
        ):
            workbook = Workbook()
            sheet = workbook.active
            assert sheet is not None
            for row in rows:
                sheet.append(row)
            workbook.save(path)
            workbook.close()

        window = main_window_with_controller
        vm = window.view_model
        vm.add_files([str(first), str(second)])

        app = QApplication.instance()
        if app is not None:
            app.processEvents()

        window._conversion_panel_vm.merge_mode = 1
        if app is not None:
            app.processEvents()

        assert window._conversion_panel._merge_tables_button is not None
        window._conversion_panel._merge_tables_button.click()

        if app is not None:
            app.processEvents()

        self._wait_for_entries_finished(window, [str(first), str(second)])
        entries = self._completed_entries(window, [str(first), str(second)])

        output_paths = {entry.output_path for entry in entries if entry.output_path is not None}
        assert len(output_paths) == 1
        output_path = Path(next(iter(output_paths)))
        assert output_path.name == "base_merged.xlsx"
        assert output_path.exists()

        workbook = load_workbook(output_path)
        try:
            sheet = workbook.active
            assert sheet is not None
            values = [cell for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None]
            assert "Alice" in values
            assert "Bob" in values
            assert "Charlie" in values
        finally:
            workbook.close()

    def test_merge_tables_uses_custom_output_directory_and_date_subfolder(
        self, main_window_with_controller, tmp_path: Path
    ) -> None:
        import re

        from openpyxl import Workbook, load_workbook
        from PySide6.QtWidgets import QApplication

        first = tmp_path / "base.xlsx"
        second = tmp_path / "collect.xlsx"
        for path, rows in (
            (first, [["Name", "Score"], ["Alice", 90]]),
            (second, [["Name", "Score"], ["Charlie", 78]]),
        ):
            workbook = Workbook()
            sheet = workbook.active
            assert sheet is not None
            for row in rows:
                sheet.append(row)
            workbook.save(path)
            workbook.close()

        window = main_window_with_controller
        controller = window.view_model.controller
        assert controller is not None
        assert controller.config_port is not None

        output_dir = tmp_path / "table_exports"
        assert controller.config_port.set("output.directory.mode", "custom")
        assert controller.config_port.set("output.directory.custom_path", str(output_dir))
        assert controller.config_port.set("output.directory.create_date_subfolder", True)
        assert controller.config_port.set("output.directory.date_folder_format", "%Y%m%d")

        window.view_model.add_files([str(first), str(second)])

        app = QApplication.instance()
        if app is not None:
            app.processEvents()

        window._conversion_panel_vm.merge_mode = 1
        if app is not None:
            app.processEvents()

        assert window._conversion_panel._merge_tables_button is not None
        window._conversion_panel._merge_tables_button.click()

        if app is not None:
            app.processEvents()

        self._wait_for_entries_finished(window, [str(first), str(second)])
        entries = self._completed_entries(window, [str(first), str(second)])

        output_paths = {entry.output_path for entry in entries if entry.output_path is not None}
        assert len(output_paths) == 1
        output_path = Path(next(iter(output_paths)))
        assert output_path.name == "base_merged.xlsx"
        assert output_path.exists()
        assert output_path.parent.parent == output_dir
        assert re.fullmatch(r"\d{8}", output_path.parent.name)

        workbook = load_workbook(output_path)
        try:
            sheet = workbook.active
            assert sheet is not None
            values = [cell for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None]
            assert "Alice" in values
            assert "Charlie" in values
        finally:
            workbook.close()

    def test_merge_images_to_tiff_runs_through_gui_thread_and_places_output(
        self, main_window_with_controller, tmp_path: Path
    ) -> None:
        from PIL import Image, ImageSequence
        from PySide6.QtWidgets import QApplication

        first = tmp_path / "red.png"
        second = tmp_path / "blue.jpg"
        Image.new("RGB", (32, 24), (255, 0, 0)).save(first)
        Image.new("RGB", (32, 24), (0, 0, 255)).save(second)

        window = main_window_with_controller
        vm = window.view_model
        vm.add_files([str(first), str(second)])

        app = QApplication.instance()
        if app is not None:
            app.processEvents()

        window._conversion_panel_vm.tiff_mode = "rgb"
        if app is not None:
            app.processEvents()

        assert window._conversion_panel._merge_tiff_button is not None
        window._conversion_panel._merge_tiff_button.click()

        if app is not None:
            app.processEvents()

        self._wait_for_entries_finished(window, [str(first), str(second)])
        entries = self._completed_entries(window, [str(first), str(second)])

        output_paths = {entry.output_path for entry in entries if entry.output_path is not None}
        assert len(output_paths) == 1
        output_path = Path(next(iter(output_paths)))
        assert output_path.name == "red_merged.tif"
        assert output_path.exists()

        image = Image.open(output_path)
        try:
            assert image.format == "TIFF"
            assert sum(1 for _ in ImageSequence.Iterator(image)) == 2
        finally:
            image.close()

    def test_merge_images_to_tiff_uses_custom_output_directory_and_date_subfolder(
        self, main_window_with_controller, tmp_path: Path
    ) -> None:
        import re

        from PIL import Image, ImageSequence
        from PySide6.QtWidgets import QApplication

        first = tmp_path / "red.png"
        second = tmp_path / "blue.jpg"
        Image.new("RGB", (32, 24), (255, 0, 0)).save(first)
        Image.new("RGB", (32, 24), (0, 0, 255)).save(second)

        window = main_window_with_controller
        controller = window.view_model.controller
        assert controller is not None
        assert controller.config_port is not None

        output_dir = tmp_path / "image_exports"
        assert controller.config_port.set("output.directory.mode", "custom")
        assert controller.config_port.set("output.directory.custom_path", str(output_dir))
        assert controller.config_port.set("output.directory.create_date_subfolder", True)
        assert controller.config_port.set("output.directory.date_folder_format", "%Y年%m月%d日")

        window.view_model.add_files([str(first), str(second)])

        app = QApplication.instance()
        if app is not None:
            app.processEvents()

        window._conversion_panel_vm.tiff_mode = "rgb"
        if app is not None:
            app.processEvents()

        assert window._conversion_panel._merge_tiff_button is not None
        window._conversion_panel._merge_tiff_button.click()

        if app is not None:
            app.processEvents()

        self._wait_for_entries_finished(window, [str(first), str(second)])
        entries = self._completed_entries(window, [str(first), str(second)])

        output_paths = {entry.output_path for entry in entries if entry.output_path is not None}
        assert len(output_paths) == 1
        output_path = Path(next(iter(output_paths)))
        assert output_path.name == "red_merged.tif"
        assert output_path.exists()
        assert output_path.parent.parent == output_dir
        assert re.fullmatch(r"\d{4}年\d{2}月\d{2}日", output_path.parent.name)

        image = Image.open(output_path)
        try:
            assert image.format == "TIFF"
            assert sum(1 for _ in ImageSequence.Iterator(image)) == 2
        finally:
            image.close()

    def test_merge_pdfs_failure_marks_all_participants_failed(
        self, main_window_with_controller, tmp_path: Path
    ) -> None:
        import fitz
        from PySide6.QtWidgets import QApplication

        good = tmp_path / "good.pdf"
        broken = tmp_path / "broken.pdf"

        doc = fitz.open()
        page = doc.new_page(width=240, height=160)
        page.insert_text((48, 80), "good")
        doc.save(good)
        doc.close()
        broken.write_bytes(b"%PDF-1.4\nnot a real pdf body\n")

        window = main_window_with_controller
        vm = window.view_model
        vm.add_files([str(good), str(broken)])

        app = QApplication.instance()
        if app is not None:
            app.processEvents()

        assert window._conversion_panel._merge_pdfs_button is not None
        window._conversion_panel._merge_pdfs_button.click()

        if app is not None:
            app.processEvents()

        self._wait_for_entries_finished(window, [str(good), str(broken)])
        entries = [
            window._batch_list_vm.get_file_entry(path)
            for path in (str(good).replace("\\", "/"), str(broken).replace("\\", "/"))
        ]
        assert all(entry is not None and entry.status == "failed" for entry in entries)
        assert all(entry is not None and not entry.output_path for entry in entries)
        assert all(entry is not None and entry.error_message for entry in entries)
        assert any("broken.pdf" in (entry.error_message or "") for entry in entries if entry is not None)
        assert any("Failed to merge" in (entry.error_message or "") for entry in entries if entry is not None)

        assert window._info_area_vm.has_task_summary
        assert window._info_area_vm.history_rows
        latest = window._info_area_vm.history_rows[-1]
        assert latest.message_type == "danger"
        assert "Failed to merge" in latest.message

    def test_merge_tables_failure_marks_all_participants_failed(
        self,
        main_window_with_controller,
        tmp_path: Path,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from openpyxl import Workbook
        from PySide6.QtWidgets import QApplication

        from docwen_core.models.result import ConversionErrorInfo, ConversionResult
        from docwen_core.protocols.execution_context import ConverterContext
        from docwen_plugin_spreadsheet.table_merger.converter import TableMergerConverter

        good = tmp_path / "good.xlsx"
        broken = tmp_path / "broken.xlsx"

        for path, name in ((good, "Alice"), (broken, "Bob")):
            workbook = Workbook()
            sheet = workbook.active
            assert sheet is not None
            sheet.append(["Name", "Score"])
            sheet.append([name, 90])
            workbook.save(path)
            workbook.close()

        def fail_merge(_converter: TableMergerConverter, context: ConverterContext) -> ConversionResult:
            request = context.request
            message = "Failed to merge 'broken.xlsx': simulated converter failure"
            return ConversionResult(
                task_id=request.request_id,
                success=False,
                error=ConversionErrorInfo(
                    error_type="conversion_failed",
                    message=message,
                    diagnostic_code="MERGE-PARSE-ERROR",
                ),
            )

        monkeypatch.setattr(TableMergerConverter, "convert", fail_merge)

        window = main_window_with_controller
        window.view_model.add_files([str(good), str(broken)])

        app = QApplication.instance()
        if app is not None:
            app.processEvents()

        assert window._conversion_panel._merge_tables_button is not None
        window._conversion_panel._merge_tables_button.click()

        if app is not None:
            app.processEvents()

        self._wait_for_entries_finished(window, [str(good), str(broken)])
        entries = [
            window._batch_list_vm.get_file_entry(path)
            for path in (str(good).replace("\\", "/"), str(broken).replace("\\", "/"))
        ]
        assert all(entry is not None and entry.status == "failed" for entry in entries)
        assert all(entry is not None and not entry.output_path for entry in entries)
        assert all(entry is not None and entry.error_message for entry in entries)
        assert any("broken.xlsx" in (entry.error_message or "") for entry in entries if entry is not None)
        assert any("Failed to merge" in (entry.error_message or "") for entry in entries if entry is not None)

        assert window._info_area_vm.has_task_summary
        assert window._info_area_vm.history_rows
        latest = window._info_area_vm.history_rows[-1]
        assert latest.message_type == "danger"
        assert "broken.xlsx" in latest.message
