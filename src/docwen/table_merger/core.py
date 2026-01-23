"""
表格汇总核心模块

功能说明：
- 支持将多个表格文件合并到一个基准表格中
- 支持三种汇总模式：按行汇总、按列汇总、按单元格汇总
- 使用滑块算法自动对齐表格（处理行列偏移问题）
- 支持多种表格格式（xlsx、xls、csv等）

核心算法：
1. 滑块对齐：通过计算重合度找到最佳偏移量
2. 按行汇总：基于涵盖关系合并行
3. 按列汇总：基于涵盖关系合并列
4. 按单元格汇总：数值相加、文本拼接
"""

import os
import logging
import tempfile
import shutil
from typing import List, Tuple, Optional, Callable
from pathlib import Path
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from docwen.utils.file_type_utils import detect_actual_file_format
from docwen.utils.path_utils import generate_output_path
from docwen.converter.formats.spreadsheet import csv_to_xlsx
from docwen.i18n import t

logger = logging.getLogger(__name__)


class TableMerger:
    """
    表格汇总主类
    
    负责完整的表格汇总流程：
    1. 预处理：格式转换、临时文件管理
    2. 对齐：滑块算法找到最佳偏移量
    3. 汇总：根据模式执行相应的汇总算法
    4. 清理：删除临时文件
    """
    
    # 汇总模式常量
    MODE_BY_ROW = 1      # 按行汇总
    MODE_BY_COLUMN = 2   # 按列汇总
    MODE_BY_CELL = 3     # 按单元格汇总
    
    # 偏移范围配置
    OFFSET_RANGE = 10    # 行列偏移范围：[-10, +10]
    
    def __init__(self):
        """初始化表格汇总器"""
        self.temp_dir = None
        self.temp_files = []
        logger.debug("TableMerger初始化完成")
    
    def merge_tables(
        self,
        base_file: str,
        collect_files: List[str],
        mode: int,
        progress_callback: Optional[Callable[[str], None]] = None
    ) -> Tuple[bool, str, Optional[str]]:
        """
        汇总多个表格文件
        
        参数:
            base_file: 基准表格文件路径
            collect_files: 待汇总表格文件路径列表
            mode: 汇总模式（1=按行, 2=按列, 3=按单元格）
            progress_callback: 进度回调函数
            
        返回:
            Tuple[bool, str, Optional[str]]: (成功标志, 消息, 输出文件路径)
        """
        logger.info(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        logger.info(f"开始表格汇总")
        logger.info(f"  基准表格: {base_file}")
        logger.info(f"  待汇总文件数: {len(collect_files)}")
        logger.info(f"  汇总模式: {self._get_mode_name(mode)}")
        logger.info(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        
        try:
            # 步骤1：创建临时目录
            self.temp_dir = tempfile.mkdtemp(prefix="table_merge_")
            logger.info(f"创建临时目录: {self.temp_dir}")
            
            # 步骤2：预处理基准表格
            if progress_callback:
                progress_callback(t('conversion.progress.preprocessing_base_table'))
            
            base_xlsx = self._preprocess_table(base_file, is_base=True)
            logger.info(f"基准表格预处理完成: {base_xlsx}")
            
            # 加载基准表格工作簿（data_only=True以读取公式的计算结果而非公式本身）
            base_wb = openpyxl.load_workbook(base_xlsx, data_only=True)
            base_ws = base_wb.active
            logger.info(f"基准表格尺寸: {base_ws.max_row}行 × {base_ws.max_column}列")
            
            # 步骤2.1：拆散基准表格的所有合并单元格
            self._unmerge_all_cells(base_ws)
            logger.info("基准表格合并单元格拆散完成")
            
            # 步骤3：依次处理每个待汇总表格
            for i, collect_file in enumerate(collect_files, 1):
                if progress_callback:
                    progress_callback(t('conversion.progress.merging_table_progress', current=i, total=len(collect_files)))
                
                logger.info(f"┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
                logger.info(f"┃ 处理待汇总表格 {i}/{len(collect_files)}                 ┃")
                logger.info(f"┣━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┫")
                logger.info(f"  文件: {Path(collect_file).name}")
                
                # 预处理当前表格
                collect_xlsx = self._preprocess_table(collect_file, is_base=False)
                
                # 加载当前表格工作簿（data_only=True以读取公式的计算结果而非公式本身）
                collect_wb = openpyxl.load_workbook(collect_xlsx, data_only=True)
                collect_ws = collect_wb.active
                logger.info(f"  表格尺寸: {collect_ws.max_row}行 × {collect_ws.max_column}列")
                
                # 步骤3.1：拆散收集表格的所有合并单元格
                self._unmerge_all_cells(collect_ws)
                logger.info("收集表格合并单元格拆散完成")
                
                # 步骤3.2：寻找最佳偏移量
                logger.info(f"  [步骤1] 滑块对齐算法")
                row_offset, col_offset = self._find_best_offset(base_ws, collect_ws)
                logger.info(f"  ✓ 最佳偏移: 行{row_offset:+d}, 列{col_offset:+d}")
                
                # 步骤3.2：执行汇总
                logger.info(f"  [步骤2] 执行{self._get_mode_name(mode)}")
                if mode == self.MODE_BY_ROW:
                    self._merge_by_row(base_ws, collect_ws, row_offset, col_offset)
                elif mode == self.MODE_BY_COLUMN:
                    self._merge_by_column(base_ws, collect_ws, row_offset, col_offset)
                elif mode == self.MODE_BY_CELL:
                    self._merge_by_cell(base_ws, collect_ws, row_offset, col_offset)
                else:
                    raise ValueError(f"不支持的汇总模式: {mode}")
                
                logger.info(f"  ✓ 汇总完成")
                logger.info(f"┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
                
                # 关闭当前表格工作簿
                collect_wb.close()
            
            # 步骤4：保存汇总结果
            if progress_callback:
                progress_callback(t('conversion.progress.saving_merge_result'))
            
            output_path = self._save_result(base_wb, base_file)
            logger.info(f"汇总结果已保存: {output_path}")
            
            # 关闭基准表格工作簿
            base_wb.close()
            
            logger.info(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            logger.info(f"✓ 表格汇总完成")
            logger.info(f"  输出文件: {output_path}")
            logger.info(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            
            return True, f"成功汇总 {len(collect_files)} 个表格", output_path
            
        except Exception as e:
            logger.error(f"表格汇总失败: {str(e)}", exc_info=True)
            return False, f"汇总失败: {str(e)}", None
            
        finally:
            # 清理临时文件
            self._cleanup_temp_files()
    
    def _preprocess_table(self, file_path: str, is_base: bool = False) -> str:
        """
        预处理表格文件
        
        将表格转换为xlsx格式并存储在临时目录中。
        使用统一的输入文件保护机制，防止中途修改原文件。
        
        参数:
            file_path: 表格文件路径
            is_base: 是否为基准表格
            
        返回:
            str: 预处理后的xlsx文件路径
        """
        logger.debug(f"预处理表格: {file_path}")
        
        # 使用统一的输入文件保护机制
        from docwen.utils.workspace_manager import prepare_input_file
        
        # 检测实际格式
        actual_format = detect_actual_file_format(file_path)
        logger.debug(f"实际格式: {actual_format}")
        
        # 创建输入文件副本
        temp_input = prepare_input_file(file_path, self.temp_dir, actual_format)
        
        # 生成目标文件名
        file_name = Path(file_path).stem
        if is_base:
            temp_path = os.path.join(self.temp_dir, f"{file_name}_base.xlsx")
        else:
            temp_path = os.path.join(self.temp_dir, f"{file_name}_collect.xlsx")
        
        # 根据格式进行相应处理
        if actual_format == 'xlsx':
            # 直接复制xlsx文件
            shutil.copy2(temp_input, temp_path)
            logger.debug(f"直接复制xlsx文件: {temp_path}")
        
        elif actual_format == 'csv':
            # CSV转xlsx
            logger.debug(f"转换CSV到xlsx")
            # 先将CSV转为xlsx
            xlsx_path = csv_to_xlsx(temp_input)
            # 复制到临时目录
            shutil.copy2(xlsx_path, temp_path)
            # 删除中间文件
            try:
                os.remove(xlsx_path)
            except:
                pass
        
        elif actual_format in ['xls', 'et']:
            # 旧版表格格式：使用Office/LibreOffice转换
            logger.debug(f"转换{actual_format.upper()}到xlsx")
            from docwen.converter.formats.spreadsheet import office_to_xlsx
            office_to_xlsx(temp_input, temp_path)  # 修复：添加output_path参数
            # 不需要复制和删除中间文件，因为直接写入temp_path
        
        else:
            raise ValueError(f"不支持的表格格式: {actual_format}")
        
        # 记录临时文件
        self.temp_files.append(temp_path)
        
        logger.info(f"表格预处理完成: {temp_path}")
        return temp_path
    
    def _find_best_offset(self, base_ws: Worksheet, collect_ws: Worksheet) -> Tuple[int, int]:
        """
        滑块对齐算法：寻找最佳偏移量
        
        遍历所有可能的行列偏移组合，计算重合度，
        选择重合度最高的偏移量。
        
        参数:
            base_ws: 基准表格工作表
            collect_ws: 待汇总表格工作表
            
        返回:
            Tuple[int, int]: (行偏移, 列偏移)
            
        异常:
            ValueError: 如果重合度为0（表格无法对齐）
        """
        logger.debug("开始滑块对齐算法")
        
        max_overlap = 0
        best_offset = (0, 0)
        
        # 遍历所有可能的偏移量
        for row_offset in range(-self.OFFSET_RANGE, self.OFFSET_RANGE + 1):
            for col_offset in range(-self.OFFSET_RANGE, self.OFFSET_RANGE + 1):
                overlap = self._calculate_overlap(
                    base_ws, collect_ws, 
                    row_offset, col_offset
                )
                
                if overlap > max_overlap:
                    max_overlap = overlap
                    best_offset = (row_offset, col_offset)
                    logger.debug(f"找到更好的偏移: ({row_offset:+d}, {col_offset:+d}), 重合度={overlap}")
        
        # 如果重合度为0，报错
        if max_overlap == 0:
            raise ValueError(
                "表格与基准表格无法对齐（重合度为0）。"
                "可能原因：表格内容完全不同，或者偏移量超过搜索范围。"
            )
        
        logger.info(f"滑块对齐完成: 偏移=({best_offset[0]:+d}, {best_offset[1]:+d}), 重合度={max_overlap}")
        return best_offset
    
    def _calculate_overlap(
        self,
        base_ws: Worksheet,
        collect_ws: Worksheet,
        row_offset: int,
        col_offset: int
    ) -> int:
        """
        计算给定偏移量下的重合度
        
        重合度定义：收集表格中有多少个非空单元格，
        在基准表格的对应位置也有相同的非空值。
        
        参数:
            base_ws: 基准表格工作表
            collect_ws: 待汇总表格工作表
            row_offset: 行偏移量
            col_offset: 列偏移量
            
        返回:
            int: 重合度（相同非空单元格的数量）
        """
        overlap = 0
        
        # 遍历收集表格的所有单元格
        for row in collect_ws.iter_rows():
            for cell in row:
                # 跳过空单元格
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                
                # 计算在基准表格中的对应位置
                base_row = cell.row + row_offset
                base_col = cell.column + col_offset
                
                # 检查位置是否有效
                if base_row < 1 or base_col < 1:
                    continue
                if base_row > base_ws.max_row or base_col > base_ws.max_column:
                    continue
                
                # 获取基准表格对应位置的单元格
                base_cell = base_ws.cell(base_row, base_col)
                
                # 如果两个单元格都非空且值相同，重合度+1
                if base_cell.value is not None and str(base_cell.value).strip() != "":
                    if str(cell.value).strip() == str(base_cell.value).strip():
                        overlap += 1
        
        return overlap
    
    def _merge_by_row(
        self,
        base_ws: Worksheet,
        collect_ws: Worksheet,
        row_offset: int,
        col_offset: int
    ):
        """
        按行汇总两个表格（智能插入）
        
        处理逻辑：
        1. 对收集表格的每一行，只使用列偏移量来对齐列
        2. 跳过收集表和基准表的空行
        3. 遍历基准表的所有行，计算与收集行的相似度
        4. 找到相似度最高的基准行，在其下方插入收集行
        
        参数:
            base_ws: 基准表格工作表
            collect_ws: 待汇总表格工作表
            row_offset: 行偏移量（不用于按行汇总）
            col_offset: 列偏移量（用于对齐列）
        """
        logger.debug("按行汇总（智能插入，仅使用列偏移，跳过空行）")
        
        # 遍历收集表格的每一行
        for collect_row_idx in range(1, collect_ws.max_row + 1):
            # 检查收集行是否为空行
            collect_row_empty = True
            for col_idx in range(1, collect_ws.max_column + 1):
                cell_value = collect_ws.cell(collect_row_idx, col_idx).value
                if cell_value is not None and str(cell_value).strip() != "":
                    collect_row_empty = False
                    break
            
            # 如果收集行是空行，跳过
            if collect_row_empty:
                logger.debug(f"跳过收集表空行: {collect_row_idx}")
                continue
            
            # 检查涵盖关系（用于判断是否需要替换或保持）
            has_coverage = False
            
            # 遍历基准表的所有行，检查是否有涵盖关系
            for base_row_idx in range(1, base_ws.max_row + 1):
                coverage = self._check_row_coverage(
                    base_ws, base_row_idx,
                    collect_ws, collect_row_idx,
                    col_offset
                )
                
                # 跳过基准表的空行
                if coverage == "skip_empty":
                    continue
                
                if coverage == "collect_covers_base":
                    # 收集行涵盖基准行：用收集行替换基准行
                    logger.debug(f"  行{base_row_idx}: 收集行涵盖基准行，替换")
                    self._replace_row(base_ws, base_row_idx, collect_ws, collect_row_idx, col_offset)
                    has_coverage = True
                    break
                
                elif coverage == "base_covers_collect":
                    # 基准行涵盖收集行：保持基准行不变
                    logger.debug(f"  行{base_row_idx}: 基准行涵盖收集行，保持不变")
                    has_coverage = True
                    break
            
            # 如果没有涵盖关系，寻找最相似行并插入
            if not has_coverage:
                logger.debug(f"  收集行{collect_row_idx}: 无涵盖关系，寻找最相似行并插入")
                self._insert_at_most_similar_row(
                    base_ws, collect_ws, collect_row_idx, col_offset
                )
    
    def _check_row_coverage(
        self,
        base_ws: Worksheet,
        base_row: int,
        collect_ws: Worksheet,
        collect_row: int,
        col_offset: int
    ) -> str:
        """
        检查两行之间的涵盖关系
        
        涵盖关系定义：
        - 如果行A的所有非空单元格在行B中对应位置的值都相同，
          且行B有更多非空单元格，则称"行B涵盖行A"
        
        参数:
            base_ws: 基准表格工作表
            base_row: 基准行号
            collect_ws: 待汇总表格工作表
            collect_row: 收集行号
            col_offset: 列偏移量
            
        返回:
            str: 涵盖关系类型
                - "collect_covers_base": 收集行涵盖基准行
                - "base_covers_collect": 基准行涵盖收集行
                - "no_coverage": 互不涵盖
                - "conflict": 有冲突
                - "out_of_range": 超出基准表格范围
                - "skip_empty": 基准行为空行，应跳过
        """
        # 检查是否超出基准表格范围
        if base_row < 1 or base_row > base_ws.max_row:
            return "out_of_range"
        
        # 检查基准行是否为空行
        base_row_empty = True
        for col_idx in range(1, base_ws.max_column + 1):
            cell_value = base_ws.cell(base_row, col_idx).value
            if cell_value is not None and str(cell_value).strip() != "":
                base_row_empty = False
                break
        
        # 如果基准行是空行，返回skip_empty
        if base_row_empty:
            logger.debug(f"基准行{base_row}为空行，跳过")
            return "skip_empty"
        
        base_has_unique = False  # 基准行是否有独有的非空值
        collect_has_unique = False  # 收集行是否有独有的非空值
        has_conflict = False  # 是否有冲突
        
        # 收集行和基准行的内容（用于日志）
        collect_row_content = []
        base_row_content = []
        
        # 计算需要检查的最大列数
        max_col = max(base_ws.max_column, collect_ws.max_column + col_offset)
        
        for col_idx in range(1, max_col + 1):
            # 获取基准行的单元格值
            if col_idx <= base_ws.max_column:
                base_value = base_ws.cell(base_row, col_idx).value
                base_empty = (base_value is None or str(base_value).strip() == "")
                if not base_empty:
                    base_row_content.append(f"列{col_idx}:{base_value}")
            else:
                base_empty = True
            
            # 计算收集行的对应列号
            collect_col = col_idx - col_offset
            if collect_col < 1 or collect_col > collect_ws.max_column:
                if not base_empty:
                    base_has_unique = True
                continue
            
            # 获取收集行的单元格值
            collect_value = collect_ws.cell(collect_row, collect_col).value
            collect_empty = (collect_value is None or str(collect_value).strip() == "")
            if not collect_empty:
                collect_row_content.append(f"列{collect_col}:{collect_value}")
            
            # 判断关系
            if base_empty and collect_empty:
                continue  # 都为空，跳过
            
            if base_empty and not collect_empty:
                collect_has_unique = True  # 收集行有独有值
            
            elif not base_empty and collect_empty:
                base_has_unique = True  # 基准行有独有值
            
            else:  # 都非空
                if str(base_value).strip() != str(collect_value).strip():
                    has_conflict = True  # 冲突
                    break
        
        # 记录详细的内容信息
        logger.debug(f"检查行涵盖关系:")
        logger.debug(f"  收集行{collect_row}内容: {', '.join(collect_row_content)}")
        logger.debug(f"  基准行{base_row}内容: {', '.join(base_row_content)}")
        logger.debug(f"  关系判断: base_has_unique={base_has_unique}, collect_has_unique={collect_has_unique}, has_conflict={has_conflict}")
        
        # 根据标志位判断涵盖关系
        if has_conflict:
            logger.debug(f"  结果: conflict (有冲突)")
            return "conflict"
        
        if collect_has_unique and not base_has_unique:
            logger.debug(f"  结果: collect_covers_base (收集行涵盖基准行)")
            return "collect_covers_base"
        
        if base_has_unique and not collect_has_unique:
            logger.debug(f"  结果: base_covers_collect (基准行涵盖收集行)")
            return "base_covers_collect"
        
        # 当两行完全相同时（都没有独有值，也没有冲突），当作互相涵盖
        # 保持基准行不变，避免重复插入
        logger.debug(f"  结果: base_covers_collect (两行完全相同，保持基准行不变)")
        return "base_covers_collect"
    
    def _replace_row(
        self,
        base_ws: Worksheet,
        base_row: int,
        collect_ws: Worksheet,
        collect_row: int,
        col_offset: int
    ):
        """
        用收集行替换基准行
        
        参数:
            base_ws: 基准表格工作表
            base_row: 基准行号
            collect_ws: 待汇总表格工作表
            collect_row: 收集行号
            col_offset: 列偏移量
        """
        # 遍历收集行的所有单元格
        for col_idx in range(1, collect_ws.max_column + 1):
            collect_cell = collect_ws.cell(collect_row, col_idx)
            base_col = col_idx + col_offset
            
            if base_col < 1:
                continue
            
            # 复制单元格值到基准表格
            base_ws.cell(base_row, base_col).value = collect_cell.value
    
    def _insert_row(
        self,
        base_ws: Worksheet,
        base_row: int,
        collect_ws: Worksheet,
        collect_row: int,
        col_offset: int
    ):
        """
        在基准表格中插入收集行
        
        参数:
            base_ws: 基准表格工作表
            base_row: 插入位置（基准行号）
            collect_ws: 待汇总表格工作表
            collect_row: 收集行号
            col_offset: 列偏移量
        """
        # 插入新行
        base_ws.insert_rows(base_row)
        
        # 复制收集行的内容
        for col_idx in range(1, collect_ws.max_column + 1):
            collect_cell = collect_ws.cell(collect_row, col_idx)
            base_col = col_idx + col_offset
            
            if base_col < 1:
                continue
            
            base_ws.cell(base_row, base_col).value = collect_cell.value
    
    def _append_row(
        self,
        base_ws: Worksheet,
        collect_ws: Worksheet,
        collect_row: int,
        col_offset: int
    ):
        """
        在基准表格末尾追加收集行
        
        参数:
            base_ws: 基准表格工作表
            collect_ws: 待汇总表格工作表
            collect_row: 收集行号
            col_offset: 列偏移量
        """
        # 追加行号 = 当前最大行号 + 1
        append_row = base_ws.max_row + 1
        
        # 复制收集行的内容
        for col_idx in range(1, collect_ws.max_column + 1):
            collect_cell = collect_ws.cell(collect_row, col_idx)
            base_col = col_idx + col_offset
            
            if base_col < 1:
                continue
            
            base_ws.cell(append_row, base_col).value = collect_cell.value
    
    def _merge_by_column(
        self,
        base_ws: Worksheet,
        collect_ws: Worksheet,
        row_offset: int,
        col_offset: int
    ):
        """
        按列汇总两个表格（智能插入）
        
        处理逻辑：
        1. 对收集表格的每一列，只使用行偏移量来对齐行
        2. 跳过收集表和基准表的空列
        3. 遍历基准表的所有列，计算与收集列的相似度
        4. 找到相似度最高的基准列，在其右侧插入收集列
        
        参数:
            base_ws: 基准表格工作表
            collect_ws: 待汇总表格工作表
            row_offset: 行偏移量（用于对齐行）
            col_offset: 列偏移量（不用于按列汇总）
        """
        logger.debug("按列汇总（智能插入，仅使用行偏移，跳过空列）")
        
        # 遍历收集表格的每一列
        for collect_col_idx in range(1, collect_ws.max_column + 1):
            # 检查收集列是否为空列
            collect_col_empty = True
            for row_idx in range(1, collect_ws.max_row + 1):
                cell_value = collect_ws.cell(row_idx, collect_col_idx).value
                if cell_value is not None and str(cell_value).strip() != "":
                    collect_col_empty = False
                    break
            
            # 如果收集列是空列，跳过
            if collect_col_empty:
                logger.debug(f"跳过收集表空列: {collect_col_idx}")
                continue
            
            # 检查涵盖关系（用于判断是否需要替换或保持）
            has_coverage = False
            
            # 遍历基准表的所有列，检查是否有涵盖关系
            for base_col_idx in range(1, base_ws.max_column + 1):
                coverage = self._check_column_coverage(
                    base_ws, base_col_idx,
                    collect_ws, collect_col_idx,
                    row_offset
                )
                
                # 跳过基准表的空列
                if coverage == "skip_empty":
                    continue
                
                if coverage == "collect_covers_base":
                    # 收集列涵盖基准列：用收集列替换基准列
                    logger.debug(f"  列{base_col_idx}: 收集列涵盖基准列，替换")
                    self._replace_column(base_ws, base_col_idx, collect_ws, collect_col_idx, row_offset)
                    has_coverage = True
                    break
                
                elif coverage == "base_covers_collect":
                    # 基准列涵盖收集列：保持基准列不变
                    logger.debug(f"  列{base_col_idx}: 基准列涵盖收集列，保持不变")
                    has_coverage = True
                    break
            
            # 如果没有涵盖关系，寻找最相似列并插入
            if not has_coverage:
                logger.debug(f"  收集列{collect_col_idx}: 无涵盖关系，寻找最相似列并插入")
                self._insert_at_most_similar_column(
                    base_ws, collect_ws, collect_col_idx, row_offset
                )
    
    def _check_column_coverage(
        self,
        base_ws: Worksheet,
        base_col: int,
        collect_ws: Worksheet,
        collect_col: int,
        row_offset: int
    ) -> str:
        """
        检查两列之间的涵盖关系
        
        逻辑与_check_row_coverage相同。
        
        参数:
            base_ws: 基准表格工作表
            base_col: 基准列号
            collect_ws: 待汇总表格工作表
            collect_col: 收集列号
            row_offset: 行偏移量
            
        返回:
            str: 涵盖关系类型
                - "collect_covers_base": 收集列涵盖基准列
                - "base_covers_collect": 基准列涵盖收集列
                - "conflict": 有冲突
                - "out_of_range": 超出基准表格范围
                - "skip_empty": 基准列为空列，应跳过
        """
        # 检查是否超出基准表格范围
        if base_col < 1 or base_col > base_ws.max_column:
            return "out_of_range"
        
        # 检查基准列是否为空列
        base_col_empty = True
        for row_idx in range(1, base_ws.max_row + 1):
            cell_value = base_ws.cell(row_idx, base_col).value
            if cell_value is not None and str(cell_value).strip() != "":
                base_col_empty = False
                break
        
        # 如果基准列是空列，返回skip_empty
        if base_col_empty:
            logger.debug(f"基准列{base_col}为空列，跳过")
            return "skip_empty"
        
        base_has_unique = False
        collect_has_unique = False
        has_conflict = False
        
        # 收集列和基准列的内容（用于日志）
        collect_col_content = []
        base_col_content = []
        
        # 计算需要检查的最大行数
        max_row = max(base_ws.max_row, collect_ws.max_row + row_offset)
        
        for row_idx in range(1, max_row + 1):
            # 获取基准列的单元格值
            if row_idx <= base_ws.max_row:
                base_value = base_ws.cell(row_idx, base_col).value
                base_empty = (base_value is None or str(base_value).strip() == "")
                if not base_empty:
                    base_col_content.append(f"行{row_idx}:{base_value}")
            else:
                base_empty = True
            
            # 计算收集列的对应行号
            collect_row = row_idx - row_offset
            if collect_row < 1 or collect_row > collect_ws.max_row:
                if not base_empty:
                    base_has_unique = True
                continue
            
            # 获取收集列的单元格值
            collect_value = collect_ws.cell(collect_row, collect_col).value
            collect_empty = (collect_value is None or str(collect_value).strip() == "")
            if not collect_empty:
                collect_col_content.append(f"行{collect_row}:{collect_value}")
            
            # 判断关系
            if base_empty and collect_empty:
                continue
            
            if base_empty and not collect_empty:
                collect_has_unique = True
            
            elif not base_empty and collect_empty:
                base_has_unique = True
            
            else:  # 都非空
                if str(base_value).strip() != str(collect_value).strip():
                    has_conflict = True
                    break
        
        # 记录详细的内容信息
        logger.debug(f"检查列涵盖关系:")
        logger.debug(f"  收集列{collect_col}内容: {', '.join(collect_col_content)}")
        logger.debug(f"  基准列{base_col}内容: {', '.join(base_col_content)}")
        logger.debug(f"  关系判断: base_has_unique={base_has_unique}, collect_has_unique={collect_has_unique}, has_conflict={has_conflict}")
        
        # 根据标志位判断涵盖关系
        if has_conflict:
            logger.debug(f"  结果: conflict (有冲突)")
            return "conflict"
        
        if collect_has_unique and not base_has_unique:
            logger.debug(f"  结果: collect_covers_base (收集列涵盖基准列)")
            return "collect_covers_base"
        
        if base_has_unique and not collect_has_unique:
            logger.debug(f"  结果: base_covers_collect (基准列涵盖收集列)")
            return "base_covers_collect"
        
        # 当两列完全相同时（都没有独有值，也没有冲突），当作互相涵盖
        # 保持基准列不变，避免重复插入
        logger.debug(f"  结果: base_covers_collect (两列完全相同，保持基准列不变)")
        return "base_covers_collect"
    
    def _replace_column(
        self,
        base_ws: Worksheet,
        base_col: int,
        collect_ws: Worksheet,
        collect_col: int,
        row_offset: int
    ):
        """用收集列替换基准列"""
        for row_idx in range(1, collect_ws.max_row + 1):
            collect_cell = collect_ws.cell(row_idx, collect_col)
            base_row = row_idx + row_offset
            
            if base_row < 1:
                continue
            
            base_ws.cell(base_row, base_col).value = collect_cell.value
    
    def _insert_column(
        self,
        base_ws: Worksheet,
        base_col: int,
        collect_ws: Worksheet,
        collect_col: int,
        row_offset: int
    ):
        """在基准表格中插入收集列"""
        base_ws.insert_cols(base_col)
        
        for row_idx in range(1, collect_ws.max_row + 1):
            collect_cell = collect_ws.cell(row_idx, collect_col)
            base_row = row_idx + row_offset
            
            if base_row < 1:
                continue
            
            base_ws.cell(base_row, base_col).value = collect_cell.value
    
    def _append_column(
        self,
        base_ws: Worksheet,
        collect_ws: Worksheet,
        collect_col: int,
        row_offset: int
    ):
        """在基准表格末尾追加收集列"""
        append_col = base_ws.max_column + 1
        
        for row_idx in range(1, collect_ws.max_row + 1):
            collect_cell = collect_ws.cell(row_idx, collect_col)
            base_row = row_idx + row_offset
            
            if base_row < 1:
                continue
            
            base_ws.cell(base_row, append_col).value = collect_cell.value
    
    def _merge_by_cell(
        self,
        base_ws: Worksheet,
        collect_ws: Worksheet,
        row_offset: int,
        col_offset: int
    ):
        """
        按单元格汇总两个表格
        
        遍历收集表格的每个单元格，与基准表格对应位置的单元格合并。
        注意：在调用此方法前，所有合并单元格已被拆散并填充值。
        
        参数:
            base_ws: 基准表格工作表
            collect_ws: 待汇总表格工作表
            row_offset: 行偏移量
            col_offset: 列偏移量
        """
        logger.debug("按单元格汇总")
        
        # 遍历收集表格的所有单元格
        for row in collect_ws.iter_rows():
            for cell in row:
                # 跳过空单元格
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                
                # 计算在基准表格中的对应位置
                base_row = cell.row + row_offset
                base_col = cell.column + col_offset
                
                # 如果位置无效，跳过
                if base_row < 1 or base_col < 1:
                    continue
                
                # 获取或创建基准表格的对应单元格
                base_cell = base_ws.cell(base_row, base_col)
                
                # 合并单元格值
                merged_value = self._merge_cell_values(base_cell.value, cell.value)
                
                # 更新基准单元格
                base_cell.value = merged_value
                
                logger.debug(f"合并单元格: 收集({cell.row}, {cell.column})={cell.value} -> "
                           f"基准({base_row}, {base_col})={base_cell.value} -> {merged_value}")
    
    def _merge_cell_values(self, base_value, collect_value):
        """
        合并两个单元格的值
        
        规则：
        1. 都为空 → 空
        2. 一个为空 → 非空值
        3. 都为数字 → 相加
        4. 都为文本且相同 → 保持不变
        5. 都为文本但不同 → 逗号拼接
        
        参数:
            base_value: 基准单元格值
            collect_value: 收集单元格值
            
        返回:
            合并后的值
        """
        # 判断是否为空
        base_empty = (base_value is None or str(base_value).strip() == "")
        collect_empty = (collect_value is None or str(collect_value).strip() == "")
        
        # 规则1: 都为空
        if base_empty and collect_empty:
            return None
        
        # 规则2: 一个为空
        if base_empty:
            return collect_value
        if collect_empty:
            return base_value
        
        # 规则3: 尝试转换为数字
        base_num = self._try_convert_to_number(base_value)
        collect_num = self._try_convert_to_number(collect_value)
        
        if base_num is not None and collect_num is not None:
            result = base_num + collect_num
            logger.debug(f"数值相加: {base_num} + {collect_num} = {result}")
            return result
        
        # 规则4: 相同文本
        if str(base_value).strip() == str(collect_value).strip():
            return base_value
        
        # 规则5: 不同文本，逗号拼接
        result = f"{base_value},{collect_value}"
        logger.debug(f"文本拼接: {base_value} + {collect_value} = {result}")
        return result
    
    def _try_convert_to_number(self, value):
        """
        尝试将值转换为数字
        
        参数:
            value: 要转换的值
            
        返回:
            float/int: 如果转换成功
            None: 如果转换失败
        """
        # 如果已经是数字类型，直接返回
        if isinstance(value, (int, float)):
            return value
        
        try:
            # 尝试转换为浮点数
            num = float(str(value).strip())
            # 如果是整数，返回int
            if num.is_integer():
                return int(num)
            return num
        except (ValueError, AttributeError):
            return None
    
    def _save_result(self, workbook: Workbook, base_file: str) -> str:
        """
        保存汇总结果
        
        参数:
            workbook: 汇总后的工作簿
            base_file: 基准表格文件路径
            
        返回:
            str: 输出文件路径
        """
        # 获取输出目录
        from docwen.utils.workspace_manager import get_output_directory
        output_dir = get_output_directory(base_file)
        
        # 生成时间戳
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 构建输出文件名（与合并功能保持一致）
        output_filename = f"{t('conversion.filenames.merged_table')}_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        # 保存工作簿
        workbook.save(output_path)
        logger.info(f"汇总结果已保存: {output_path}")
        
        return output_path
    
    def _cleanup_temp_files(self):
        """清理临时文件和目录"""
        if not self.temp_dir:
            return
        
        try:
            logger.debug(f"清理临时目录: {self.temp_dir}")
            shutil.rmtree(self.temp_dir, ignore_errors=True)
            logger.info("临时文件清理完成")
        except Exception as e:
            logger.warning(f"清理临时文件失败: {e}")
    
    def _unmerge_all_cells(self, worksheet: Worksheet):
        """
        拆散工作表中的所有合并单元格
        
        处理逻辑：
        1. 获取所有合并单元格区域
        2. 拆散合并单元格
        3. 将合并区域的值填充到所有单元格
        
        参数:
            worksheet: 要处理的工作表
        """
        try:
            # 获取所有合并单元格区域
            merged_ranges = list(worksheet.merged_cells.ranges)
            
            if not merged_ranges:
                logger.debug("工作表没有合并单元格")
                return
            
            logger.info(f"开始拆散 {len(merged_ranges)} 个合并单元格区域")
            
            for merged_range in merged_ranges:
                # 获取合并区域的值（取左上角单元格的值）
                top_left_cell = worksheet.cell(
                    row=merged_range.min_row, 
                    column=merged_range.min_col
                )
                merged_value = top_left_cell.value
                
                # 拆散合并单元格
                worksheet.unmerge_cells(str(merged_range))
                logger.debug(f"拆散合并区域: {merged_range.min_row}-{merged_range.max_row}, "
                           f"{merged_range.min_col}-{merged_range.max_col}")
                
                # 将合并区域的值填充到所有单元格
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.value = merged_value
                
                logger.debug(f"填充合并区域值: {merged_value}")
            
            logger.info("合并单元格拆散完成")
            
        except Exception as e:
            logger.warning(f"拆散合并单元格时发生异常: {str(e)}")

    def _get_mode_name(self, mode: int) -> str:
        """获取汇总模式名称"""
        mode_names = {
            self.MODE_BY_ROW: "按行汇总",
            self.MODE_BY_COLUMN: "按列汇总", 
            self.MODE_BY_CELL: "按单元格汇总"
        }
        return mode_names.get(mode, f"未知模式({mode})")

    def _calculate_row_similarity(
        self,
        base_ws: Worksheet,
        base_row: int,
        collect_ws: Worksheet,
        collect_row: int,
        col_offset: int
    ) -> float:
        """
        计算两行之间的相似度
        
        相似度定义：基于单元格内容匹配计算相似度分数 (0-1)
        
        参数:
            base_ws: 基准表格工作表
            base_row: 基准行号
            collect_ws: 收集表格工作表
            collect_row: 收集行号
            col_offset: 列偏移量
            
        返回:
            float: 相似度分数 (0-1)
        """
        total_cells = 0
        matching_cells = 0
        
        # 计算需要比较的列范围
        max_col = max(base_ws.max_column, collect_ws.max_column + col_offset)
        
        for col_idx in range(1, max_col + 1):
            # 获取基准单元格值
            if col_idx <= base_ws.max_column:
                base_value = base_ws.cell(base_row, col_idx).value
                base_empty = (base_value is None or str(base_value).strip() == "")
            else:
                base_empty = True
                
            # 计算收集表格的对应列
            collect_col = col_idx - col_offset
            if collect_col < 1 or collect_col > collect_ws.max_column:
                if not base_empty:
                    total_cells += 1
                continue
                
            # 获取收集单元格值
            collect_value = collect_ws.cell(collect_row, collect_col).value
            collect_empty = (collect_value is None or str(collect_value).strip() == "")
            
            total_cells += 1
            
            # 判断是否匹配
            if base_empty and collect_empty:
                matching_cells += 1
            elif not base_empty and not collect_empty:
                if str(base_value).strip() == str(collect_value).strip():
                    matching_cells += 1
        
        return matching_cells / total_cells if total_cells > 0 else 0.0

    def _calculate_column_similarity(
        self,
        base_ws: Worksheet,
        base_col: int,
        collect_ws: Worksheet,
        collect_col: int,
        row_offset: int
    ) -> float:
        """
        计算两列之间的相似度
        
        逻辑与_calculate_row_similarity相同，但操作对象为列。
        
        参数:
            base_ws: 基准表格工作表
            base_col: 基准列号
            collect_ws: 收集表格工作表
            collect_col: 收集列号
            row_offset: 行偏移量
            
        返回:
            float: 相似度分数 (0-1)
        """
        total_cells = 0
        matching_cells = 0
        
        # 计算需要比较的行范围
        max_row = max(base_ws.max_row, collect_ws.max_row + row_offset)
        
        for row_idx in range(1, max_row + 1):
            # 获取基准单元格值
            if row_idx <= base_ws.max_row:
                base_value = base_ws.cell(row_idx, base_col).value
                base_empty = (base_value is None or str(base_value).strip() == "")
            else:
                base_empty = True
                
            # 计算收集表格的对应行
            collect_row = row_idx - row_offset
            if collect_row < 1 or collect_row > collect_ws.max_row:
                if not base_empty:
                    total_cells += 1
                continue
                
            # 获取收集单元格值
            collect_value = collect_ws.cell(collect_row, collect_col).value
            collect_empty = (collect_value is None or str(collect_value).strip() == "")
            
            total_cells += 1
            
            # 判断是否匹配
            if base_empty and collect_empty:
                matching_cells += 1
            elif not base_empty and not collect_empty:
                if str(base_value).strip() == str(collect_value).strip():
                    matching_cells += 1
        
        return matching_cells / total_cells if total_cells > 0 else 0.0

    def _insert_at_most_similar_row(
        self,
        base_ws: Worksheet,
        collect_ws: Worksheet,
        collect_row: int,
        col_offset: int
    ):
        """
        在基准表中找到与收集行最相似的行，并在其下方插入
        
        参数:
            base_ws: 基准表格工作表
            collect_ws: 收集表格工作表
            collect_row: 收集行号
            col_offset: 列偏移量
        """
        max_similarity = 0.0
        best_base_row = 1  # 默认在第一行下方插入
        
        # 收集收集行的内容（用于日志）
        collect_row_content = []
        for col_idx in range(1, collect_ws.max_column + 1):
            collect_cell = collect_ws.cell(collect_row, col_idx)
            if collect_cell.value is not None and str(collect_cell.value).strip() != "":
                collect_row_content.append(f"列{col_idx}:{collect_cell.value}")
        
        # 遍历基准表的所有行，找到最相似的行
        for base_row_idx in range(1, base_ws.max_row + 1):
            similarity = self._calculate_row_similarity(
                base_ws, base_row_idx,
                collect_ws, collect_row,
                col_offset
            )
            
            if similarity > max_similarity:
                max_similarity = similarity
                best_base_row = base_row_idx
        
        # 收集最相似基准行的内容（用于日志）
        best_base_row_content = []
        for col_idx in range(1, base_ws.max_column + 1):
            base_cell = base_ws.cell(best_base_row, col_idx)
            if base_cell.value is not None and str(base_cell.value).strip() != "":
                best_base_row_content.append(f"列{col_idx}:{base_cell.value}")
        
        logger.debug(f"智能插入行:")
        logger.debug(f"  收集行{collect_row}内容: {', '.join(collect_row_content)}")
        logger.debug(f"  最相似基准行{best_base_row}内容: {', '.join(best_base_row_content)}")
        logger.debug(f"  相似度: {max_similarity:.2f}, 插入位置: {best_base_row + 1}")
        
        # 在最相似行的下方插入
        insert_position = best_base_row + 1
        base_ws.insert_rows(insert_position)
        
        # 复制收集行的内容
        for col_idx in range(1, collect_ws.max_column + 1):
            collect_cell = collect_ws.cell(collect_row, col_idx)
            base_col = col_idx + col_offset
            
            if base_col < 1:
                continue
            
            base_ws.cell(insert_position, base_col).value = collect_cell.value

    def _insert_at_most_similar_column(
        self,
        base_ws: Worksheet,
        collect_ws: Worksheet,
        collect_col: int,
        row_offset: int
    ):
        """
        在基准表中找到与收集列最相似的列，并在其右侧插入
        
        参数:
            base_ws: 基准表格工作表
            collect_ws: 收集表格工作表
            collect_col: 收集列号
            row_offset: 行偏移量
        """
        max_similarity = 0.0
        best_base_col = 1  # 默认在第一列右侧插入
        
        # 遍历基准表的所有列，找到最相似的列
        for base_col_idx in range(1, base_ws.max_column + 1):
            similarity = self._calculate_column_similarity(
                base_ws, base_col_idx,
                collect_ws, collect_col,
                row_offset
            )
            
            if similarity > max_similarity:
                max_similarity = similarity
                best_base_col = base_col_idx
        
        logger.debug(f"  最相似列: {best_base_col}, 相似度: {max_similarity:.2f}")
        
        # 在最相似列的右侧插入
        insert_position = best_base_col + 1
        base_ws.insert_cols(insert_position)
        
        # 复制收集列的内容
        for row_idx in range(1, collect_ws.max_row + 1):
            collect_cell = collect_ws.cell(row_idx, collect_col)
            base_row = row_idx + row_offset
            
            if base_row < 1:
                continue
            
            base_ws.cell(base_row, insert_position).value = collect_cell.value
