"""
表格格式互转核心模块
支持 CSV ↔ XLSX 互转

核心特性:
- 支持扩展名不匹配的文件（使用临时文件机制）
- 自动检测文件真实格式
- 多工作表XLSX转CSV（每个工作表生成一个CSV文件）
"""

import logging
import csv
import re
from pathlib import Path

import openpyxl

from docwen.utils.path_utils import generate_output_path

logger = logging.getLogger(__name__)


def csv_to_xlsx(
    csv_path: str,
    output_path: str | None = None,
) -> str:
    """
    将CSV文件转换为XLSX文件

    纯粹的转换函数，假设：
    - csv_path 是真实的CSV文件（策略层已确保格式正确）
    - 不负责格式检测和临时文件管理

    参数:
        csv_path: CSV文件路径
        output_path: 输出XLSX文件路径（可选；如不提供则自动生成）

    返回:
        str: 生成的XLSX文件路径

    异常:
        Exception: 转换失败时抛出异常
    """
    logger.info(f"开始转换 CSV 到 XLSX: {csv_path}")

    try:

        def _maybe_number(v: str):
            s = (v or "").strip()
            if s == "":
                return ""
            if re.fullmatch(r"-?[0-9]+", s):
                if len(s) > 1 and s.startswith("0"):
                    return v
                if len(s) > 2 and s.startswith("-0"):
                    return v
                try:
                    return int(s)
                except Exception:
                    return v
            if re.fullmatch(r"-?[0-9]+\\.[0-9]+", s):
                try:
                    return float(s)
                except Exception:
                    return v
            return v

        if output_path is None:
            output_path = generate_output_path(
                csv_path, section="", add_timestamp=True, description="fromCsv", file_type="xlsx"
            )
        logger.debug(f"输出路径: {output_path}")

        output_dir = Path(output_path).parent
        if output_dir != Path():
            output_dir.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for r_idx, row in enumerate(reader, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=_maybe_number(value))

        wb.save(output_path)

        logger.info(f"CSV 转换为 XLSX 成功: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"CSV 转换为 XLSX 失败: {e}", exc_info=True)
        raise


def xlsx_to_csv(
    xlsx_path: str,
    output_dir: str | None = None,
    original_basename: str | None = None,
    unified_timestamp_desc: str | None = None,
) -> list[str]:
    """
    将XLSX文件转换为CSV文件（每个工作表一个CSV）

    纯粹的转换函数，假设：
    - xlsx_path 是真实的XLSX文件（策略层已确保格式正确）
    - 不负责格式检测和临时文件管理

    参数:
        xlsx_path: XLSX文件路径
        output_dir: 输出目录（可选），如果不提供则输出到源文件所在目录
        original_basename: 原始文件的basename（不含路径、扩展名、时间戳、描述），
                          如果提供则使用此值构建CSV文件名，否则从xlsx_path提取
        unified_timestamp_desc: 统一的时间戳和描述部分（如 "20251111_145257_fromMd"），
                               如果提供则使用此值，否则自动生成

    返回:
        List[str]: 生成的CSV文件路径列表（按工作表顺序）

    异常:
        Exception: 转换失败时抛出异常

    说明:
        所有工作表的CSV文件使用统一的时间戳
        如果提供了original_basename和unified_timestamp_desc，将使用它们构建文件名，
        确保文件名与子文件夹名一致
    """
    logger.info(f"开始转换 XLSX 到 CSV: {xlsx_path}")

    # 如果没有提供output_dir，使用源文件所在目录
    if output_dir is None:
        output_dir = str(Path(xlsx_path).parent)

    try:
        # 直接使用openpyxl打开XLSX文件
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        logger.debug(f"XLSX 文件打开成功，工作表数量: {len(wb.sheetnames)}")

        # 步骤1：确定原始文件名和时间戳描述部分
        if original_basename is not None and unified_timestamp_desc is not None:
            # 使用调用者提供的参数（确保文件名与子文件夹一致）
            logger.debug(f"使用提供的原始basename: {original_basename}")
            logger.debug(f"使用提供的时间戳描述: {unified_timestamp_desc}")
            timestamp_and_desc = unified_timestamp_desc
        else:
            # 降级方案：使用旧逻辑（兼容性）
            logger.debug("未提供原始basename，使用自动生成逻辑")
            base_path = generate_output_path(
                xlsx_path, section="", add_timestamp=True, description="fromXlsx", file_type="csv"
            )
            basename = Path(base_path).stem
            logger.debug(f"自动生成basename: {basename}")

            # 提取原始文件名（不含扩展名）
            original_basename = Path(xlsx_path).stem

            # 从basename中提取时间戳和description部分
            parts = basename.split("_")
            timestamp_and_desc = ""
            if len(parts) >= 3:
                # 找到时间戳的位置（格式: YYYYMMDD）
                timestamp_idx = None
                for i, part in enumerate(parts):
                    if len(part) == 8 and part.isdigit():
                        timestamp_idx = i
                        break

                if timestamp_idx is not None:
                    # 提取时间戳和description部分
                    timestamp_and_desc = "_".join(parts[timestamp_idx:])

            if not timestamp_and_desc:
                # 降级方案：使用完整basename
                timestamp_and_desc = "_".join(parts[1:]) if len(parts) > 1 else parts[0]

        csv_files = []

        # 步骤2：遍历所有工作表，使用统一时间戳
        for sheet_name in wb.sheetnames:
            logger.debug(f"处理工作表: {sheet_name}")
            ws = wb[sheet_name]

            # 处理工作表名（空格替换为下划线）
            clean_sheet_name = sheet_name.replace(" ", "_")
            logger.debug(f"清理后的工作表名: {clean_sheet_name}")

            # 构建输出文件名：原名_工作表名_时间戳_description.csv
            csv_filename = f"{original_basename}_{clean_sheet_name}_{timestamp_and_desc}.csv"
            output_path = str(Path(output_dir) / csv_filename)
            logger.debug(f"输出路径: {output_path}")

            with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])

            csv_files.append(output_path)
            logger.info(f"工作表 {sheet_name} 转换为 CSV 成功: {output_path}")

        logger.info(f"XLSX 转换为 CSV 完成，共生成 {len(csv_files)} 个文件")
        return csv_files

    except Exception as e:
        logger.error(f"XLSX 转换为 CSV 失败: {e}", exc_info=True)
        raise
