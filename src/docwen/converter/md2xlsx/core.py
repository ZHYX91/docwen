"""
Excel 转换器核心模块

统一的临时目录管理和转换流程
"""

import os
import re
import logging
import shutil
import tempfile
import threading
from typing import Callable, Optional
from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment
from copy import copy

from docwen.utils.text_utils import clean_text, is_pure_number
from docwen.utils.workspace_manager import prepare_input_file, should_save_intermediate_files
from docwen.utils.path_utils import generate_output_path
from docwen.utils import yaml_utils
from docwen.i18n import t, t_all_locales
from docwen.utils.markdown_utils import extract_markdown_tables
from .placeholder_processor import (
    extract_template_placeholders,
    create_replacements,
    find_placeholder_positions,
    process_cell_replacements,
    clean_remaining_placeholders,
    _is_fraction
)

logger = logging.getLogger(__name__)


def _get_fallback_title(yaml_data: dict, md_path: Optional[str] = None) -> str:
    """
    获取标题的回退值
    
    优先级：aliases → 文件名 → "标题"
    
    参数:
        yaml_data: YAML数据字典
        md_path: 原始MD文件路径（用于提取文件名）
    
    返回:
        str: 回退标题文本
    """
    # 优先级1：aliases
    if 'aliases' in yaml_data:
        aliases = yaml_data['aliases']
        if isinstance(aliases, str) and aliases.strip():
            logger.debug("使用YAML中的'aliases'键（字符串）作为回退标题")
            return aliases.strip()
        elif isinstance(aliases, list) and aliases:
            logger.debug("使用YAML中的'aliases'键（列表第一项）作为回退标题")
            return str(aliases[0]).strip()
    
    # 优先级2：文件名
    if md_path:
        filename = os.path.splitext(os.path.basename(md_path))[0]
        if filename:
            logger.debug(f"从文件名提取回退标题: {filename}")
            return filename
    
    # 最终默认值
    logger.debug("使用默认回退标题")
    return "标题"


def _ensure_title_fallbacks(yaml_data: dict, md_path: Optional[str] = None):
    """
    为所有语言版本的标题键设置回退值
    
    逻辑：对于每个标题键（如 'title', '标题', 'Titel'...）：
    1. 如果 YAML 中已有该键且非空 → 保持原值
    2. 如果 YAML 中没有该键或为空 → 设置回退值（aliases → 文件名）
    
    参数:
        yaml_data: YAML数据字典（会被原地修改）
        md_path: 原始MD文件路径（用于提取文件名）
    """
    # 获取所有语言版本的标题键名（从 placeholders.title）
    title_variants = set(t_all_locales('placeholders.title').values())
    logger.debug(f"标题键变体: {title_variants}")
    
    # 计算回退值
    fallback_value = _get_fallback_title(yaml_data, md_path)
    
    # 为每个变体设置回退值
    for title_key in title_variants:
        if title_key not in yaml_data or not yaml_data[title_key]:
            yaml_data[title_key] = fallback_value
            logger.debug(f"为 '{title_key}' 设置回退值: {fallback_value}")


def convert(
    md_path: str,
    output_path: str,
    *,
    template_name: str,
    progress_callback: Optional[Callable[[str], None]] = None,
    cancel_event: Optional[threading.Event] = None,
    original_source_path: Optional[str] = None
) -> Optional[str]:
    """
    将Markdown文件转换为Excel报表（XLSX格式）
    
    统一的临时目录管理流程：
    1. 创建临时目录
    2. 准备input.md副本
    3. 对副本进行转换
    4. 保存到指定输出路径
    5. 自动清理临时目录
    
    参数:
        md_path: Markdown文件路径
        output_path: 输出文件完整路径
        template_name: 模板名称（必需）
        progress_callback: 进度回调函数
        cancel_event: 取消事件
        original_source_path: 原始源文件路径（用于嵌入功能的路径解析，可选）
    
    返回:
        str: 输出文件完整路径，失败时返回None
    """
    try:
        logger.info("=" * 60)
        logger.info(f"开始MD转Excel: {os.path.basename(md_path)}")
        logger.info(f"使用模板: {template_name}")
        logger.info("=" * 60)
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # 1. 创建input.md副本
            temp_input = prepare_input_file(md_path, temp_dir, 'md')
            logger.debug(f"已创建输入副本: {os.path.basename(temp_input)}")
            
            # 2. 读取和解析MD文件（传递原始路径用于路径解析）
            if progress_callback:
                progress_callback(t('conversion.progress.parsing_markdown'))
            
            # 如果提供了original_source_path，使用它；否则使用md_path
            path_for_resolve = original_source_path if original_source_path else md_path
            yaml_data, md_body = _read_and_parse_md(temp_input, path_for_resolve)
            
            # 为所有语言版本的标题键设置回退值
            _ensure_title_fallbacks(yaml_data, path_for_resolve)
            logger.debug("标题回退值设置完成")
            
            if cancel_event and cancel_event.is_set():
                logger.info("操作已取消")
                return None
            
            # 3. 获取模板路径
            template_path = _get_template_path(template_name)
            
            # 4. 在临时目录生成临时输出文件
            temp_output = os.path.join(temp_dir, "temp_output.xlsx")
            
            # 5. 在临时目录转换
            success, warnings = _convert_internal(
                yaml_data, md_body, template_path, temp_output,
                progress_callback, cancel_event
            )
            
            if not success:
                logger.error("内部转换失败")
                return None
            
            if cancel_event and cancel_event.is_set():
                logger.info("操作已取消")
                return None
            
            # 6. 确保输出目录存在
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # 7. 移动到最终输出路径
            shutil.move(temp_output, output_path)
            
            # 9. 保留中间文件（如需要）
            if should_save_intermediate_files():
                logger.debug("保留中间文件到输出目录")
                for filename in os.listdir(temp_dir):
                    src = os.path.join(temp_dir, filename)
                    if os.path.isfile(src):
                        dst = os.path.join(output_dir, filename)
                        shutil.copy(src, dst)
                        logger.debug(f"保留中间文件: {filename}")
            
            # 10. 如果有警告消息，记录到日志
            if warnings:
                logger.warning(f"转换完成，但有 {len(warnings)} 处数据因只读单元格无法填充")
                if progress_callback:
                    progress_callback(t('conversion.progress.conversion_completed_with_readonly_warnings', count=len(warnings)))
            
            logger.info(f"MD转Excel成功: {output_path}")
            return output_path
        
    except Exception as e:
        logger.error(f"MD转Excel失败: {e}", exc_info=True)
        return None


def _read_and_parse_md(temp_md_path: str, original_md_path: Optional[str] = None) -> tuple:
    """
    读取并解析Markdown文件，返回YAML数据和YAML后的Markdown内容
    
    新增功能：
    - 在返回前展开Markdown内容中的嵌入链接（如果启用）
    
    参数:
        temp_md_path: 临时Markdown文件路径（用于读取内容）
        original_md_path: 原始Markdown文件路径（用于路径解析，可选）
    
    返回:
        tuple: (yaml_data字典, md_body字符串)
    """
    logger.debug(f"读取Markdown文件: {temp_md_path}")
    with open(temp_md_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 匹配YAML格式部分
    yaml_match = re.search(r'---\n(.*?)\n---', content, re.DOTALL)
    
    if yaml_match:
        yaml_content = yaml_match.group(1).replace('\t', '  ')
        md_body = content[yaml_match.end():].strip()
        logger.debug(f"提取YAML内容: {len(yaml_content)} 字符")
        logger.debug(f"提取YAML后的内容: {len(md_body)} 字符")
    else:
        logger.warning("未找到YAML头部，使用空YAML")
        yaml_content = ""
        md_body = content.strip()
    
    yaml_data = yaml_utils.parse_yaml(yaml_content)
    
    # 处理所有Markdown链接（嵌入和非嵌入）
    try:
        from docwen.utils.link_processing import process_markdown_links
        
        # 使用原始文件路径进行路径解析（如果提供），否则使用临时路径
        source_path_for_resolve = original_md_path if original_md_path else temp_md_path
        
        # 1. 处理YAML字段值中的链接
        logger.info("开始处理YAML字段值中的链接...")
        
        def process_yaml_links(data, source_path):
            """递归处理YAML数据结构中的所有链接"""
            if isinstance(data, dict):
                return {k: process_yaml_links(v, source_path) for k, v in data.items()}
            elif isinstance(data, list):
                return [process_yaml_links(item, source_path) for item in data]
            elif isinstance(data, str):
                # 对字符串调用链接处理（depth=0，允许嵌套展开）
                return process_markdown_links(data, source_path, set(), depth=0)
            else:
                return data
        
        yaml_data = process_yaml_links(yaml_data, source_path_for_resolve)
        logger.info("YAML字段链接处理完成")
        
        # 2. 处理Markdown内容中的链接
        logger.info("开始处理Markdown内容中的链接...")
        original_length = len(md_body)
        md_body = process_markdown_links(md_body, source_path_for_resolve)
        new_length = len(md_body)
        logger.info(f"Markdown内容链接处理完成 | 长度变化: {original_length} → {new_length}")
    except Exception as e:
        logger.error(f"处理Markdown链接失败: {e}", exc_info=True)
        logger.warning("将继续使用原始MD内容")
    
    return yaml_data, md_body


def _get_template_path(template_name: str) -> str:
    """
    获取Excel模板完整路径
    
    参数:
        template_name: 模板名称
        
    返回:
        str: 模板完整路径
    """
    try:
        # 使用模板加载器获取路径（自动处理扩展名）
        from docwen.template.loader import TemplateLoader
        template_loader = TemplateLoader()
        path = template_loader.get_template_path("xlsx", template_name)
        
        # 验证模板存在
        if not os.path.exists(path):
            raise FileNotFoundError(f"模板文件不存在: {path}")
            
        return path
    except FileNotFoundError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"获取模板路径失败: {str(e)}")
        raise RuntimeError("无法获取模板路径")


def _convert_internal(
    yaml_data: dict,
    md_body: str,
    template_path: str,
    output_path: str,
    progress_callback,
    cancel_event
) -> tuple:
    """
    内部转换实现
    
    参数:
        yaml_data: YAML数据字典
        md_body: Markdown正文
        template_path: Excel模板路径
        output_path: 输出文件路径
        progress_callback: 进度回调
        cancel_event: 取消事件
    
    返回:
        tuple: (转换是否成功, 警告消息列表)
    """
    logger.info("开始Excel转换 (通用处理逻辑)")
    logger.info(f"模板路径: {template_path}")
    logger.info(f"输出路径: {output_path}")
    
    warning_messages = []
    
    try:
        # 1. 提取模板占位符
        if progress_callback:
            progress_callback(t('conversion.progress.loading_excel_template'))
        logger.debug("从模板提取占位符...")
        template_placeholders = extract_template_placeholders(template_path)
        logger.info(f"从模板中提取到 {len(template_placeholders)} 个占位符")
        
        # 2. 创建YAML字段替换配置
        logger.debug("创建YAML字段替换配置...")
        replacements = create_replacements(yaml_data, template_placeholders)
        logger.info(f"生成 {len(replacements)} 个YAML字段替换项")
        
        if cancel_event and cancel_event.is_set():
            return (False, warning_messages)
        
        # 3. 提取所有表格数据
        if progress_callback:
            progress_callback(t('conversion.progress.extracting_markdown_tables'))
        logger.debug("从MD正文提取所有表格数据...")
        table_dict = extract_all_tables_to_dict(md_body)
        logger.info(f"提取到 {len(table_dict)} 个表头的数据")
        
        # 4. 加载Excel模板
        logger.debug("加载Excel模板...")
        wb = load_workbook(template_path)
        
        # 保存所有工作表的合并单元格信息
        merged_ranges_info = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            merged_ranges_info[sheet_name] = [m.coord for m in ws.merged_cells.ranges]
        
        total_merged = sum(len(ranges) for ranges in merged_ranges_info.values())
        logger.info(f"模板包含 {total_merged} 个合并单元格")
        
        if cancel_event and cancel_event.is_set():
            return (False, warning_messages)
        
        # 5. 执行YAML字段占位符替换
        if progress_callback:
            progress_callback(t('conversion.progress.filling_yaml_data'))
        logger.debug("开始替换所有工作表中的YAML字段占位符...")
        process_cell_replacements(wb, replacements)
        logger.info("所有工作表的YAML字段占位符替换完成")
        
        if cancel_event and cancel_event.is_set():
            return (False, warning_messages)
        
        # 6. 执行表格数据填充
        if progress_callback:
            progress_callback(t('conversion.progress.filling_table_data'))
        logger.debug("开始填充表格数据到Excel...")
        table_warnings = fill_dict_data_to_excel(wb, table_dict, template_placeholders)
        warning_messages.extend(table_warnings)
        logger.info("表格数据填充完成")
        
        # 7. 恢复所有工作表的合并单元格
        logger.debug("恢复所有工作表的合并单元格...")
        for sheet_name, merged_ranges in merged_ranges_info.items():
            ws = wb[sheet_name]
            for merged in merged_ranges:
                ws.merge_cells(merged)
        logger.info(f"成功恢复 {total_merged} 个合并单元格")
        
        # 8. 处理图片占位符
        if progress_callback:
            progress_callback(t('conversion.progress.processing_images'))
        logger.debug("开始处理图片占位符...")
        from .placeholder_processor import process_image_placeholders
        process_image_placeholders(wb)
        logger.info("图片占位符处理完成")
        
        # 9. 清理剩余占位符
        logger.debug("开始清理剩余占位符...")
        clean_remaining_placeholders(wb)
        logger.info("剩余占位符清理完成")
        
        if cancel_event and cancel_event.is_set():
            return (False, warning_messages)
        
        # 10. 保存结果
        if progress_callback:
            progress_callback(t('conversion.progress.saving_excel'))
        logger.debug("保存Excel文件...")
        wb.save(output_path)
        logger.info(f"Excel文件保存成功: {output_path}")
        
        return (True, warning_messages)
        
    except Exception as e:
        error_msg = f"Excel转换失败: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return (False, [error_msg])


def process_cell_value(raw_value):
    """
    处理单个单元格值，应用统一规则
    
    新增功能：
    - 启用配置化链接处理
    
    参数:
        raw_value: 原始单元格值
        
    返回:
        tuple: (处理后的值, 单元格格式)
    """
    if not raw_value or raw_value.strip() == "":
        return ("", None)
    
    # 去除前后空格
    cleaned_value = raw_value.strip()
    
    # 分数处理
    if _is_fraction(cleaned_value):
        return (f"={cleaned_value}", None)
    
    # 数字处理
    if is_pure_number(cleaned_value):
        # 计算数字字符总数（移除小数点、分数符号、负号等非数字字符）
        digit_chars = re.sub(r'[./-]', '', cleaned_value)
        # 如果数字字符数超过15位，设置为文本格式防止科学计数法
        if len(digit_chars) > 15:
            return (cleaned_value, numbers.FORMAT_TEXT)
        else:
            # 短数字保持数字格式
            return (cleaned_value, None)
    
    # 文本清理（HTML、零宽字符等）
    # 注意：链接已在expand_embedded_content()中统一处理，这里只做基础清理
    return (clean_text(cleaned_value), None)


def extract_all_tables_to_dict(md_body):
    """
    提取MD中所有表格并合并为字典（使用共用表格解析代码）
    
    参数:
        md_body: Markdown正文内容
        
    返回:
        dict: {表头: [(值1, 格式1), (值2, 格式2), ...]}
    """
    logger.debug("提取所有MD表格数据...")
    table_dict = {}
    
    try:
        # 使用共用的表格提取函数
        tables = extract_markdown_tables(md_body)
        logger.info(f"找到 {len(tables)} 个表格")
        
        for table_index, table_data in enumerate(tables):
            try:
                headers = table_data['headers']
                rows = table_data['rows']
                
                # 处理每一行数据
                for row_data in rows:
                    # 将数据添加到对应表头的列表中
                    for i, header in enumerate(headers):
                        if i < len(row_data):
                            processed_value, cell_format = process_cell_value(row_data[i])
                            if header not in table_dict:
                                table_dict[header] = []
                            table_dict[header].append((processed_value, cell_format))
                            logger.debug(f"表格 {table_index+1}: {header} -> ({processed_value}, {cell_format})")
                
                logger.debug(f"表格 {table_index+1} 处理完成")
                
            except Exception as e:
                logger.warning(f"跳过表格 {table_index+1}: {str(e)}")
                continue
                
    except Exception as e:
        logger.error(f"表格提取失败: {str(e)}")
    
    logger.info(f"提取到 {len(table_dict)} 个表头的数据")
    return table_dict


def is_cell_writable(ws, row, col):
    """
    检查单元格是否可写入（不是合并单元格的非左上角）
    
    参数:
        ws: 工作表对象
        row: 行号
        col: 列号
        
    返回:
        bool: 如果可写入返回True，否则False
    """
    # 检查所有合并单元格范围
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        # 如果单元格在合并范围内
        if min_row <= row <= max_row and min_col <= col <= max_col:
            # 只有左上角单元格可写入
            if row == min_row and col == min_col:
                return True
            else:
                return False
    return True


def is_cell_protected(ws, cell):
    """
    检查单元格是否被保护
    
    参数:
        ws: 工作表对象
        cell: 单元格对象
        
    返回:
        bool: 如果被保护返回True，否则False
    """
    try:
        # 检查工作表保护状态
        if ws.protection.sheet:
            # 检查单元格保护状态
            if cell.protection.locked:
                return True
    except Exception as e:
        logger.debug(f"检查单元格保护状态时出错: {str(e)}")
    return False


def add_cell_comment(cell, comment_text, consumed_value=None):
    """
    给单元格添加批注
    
    参数:
        cell: 单元格对象
        comment_text: 批注文本
        consumed_value: 消耗的元素值（可选）
    """
    try:
        # 如果单元格已有批注，先清除
        if cell.comment:
            cell.comment = None
        
        # 构建完整的批注文本
        full_comment = comment_text
        if consumed_value is not None:
            full_comment += f"，消耗了元素: {consumed_value}"
        
        # 添加新批注
        from openpyxl.comments import Comment
        comment = Comment(full_comment, "系统")
        cell.comment = comment
        logger.debug(f"添加批注: {cell.coordinate} -> {full_comment}")
    except Exception as e:
        logger.warning(f"无法添加批注到单元格 {cell.coordinate}: {str(e)}")


def fill_dict_data_to_excel(wb, table_dict, template_placeholders_dict):
    """
    将字典数据填充到Excel中（智能合并单元格处理）
    
    参数:
        wb: 工作簿对象
        table_dict: 表格数据字典 {表头: [(值1, 格式1), (值2, 格式2), ...]}
        template_placeholders_dict: 分类后的占位符字典
        
    返回:
        list: 警告消息列表
    """
    logger.debug("开始填充表格数据到Excel...")
    
    filled_count = 0
    warning_messages = set()  # 用于收集警告消息，避免重复
    
    # 处理列填充占位符
    for placeholder in template_placeholders_dict['column_placeholders']:
        # 提取字段名（去掉↓符号）
        field_name = placeholder[3:-2]  # 移除 {{↓ 和 }}
        
        if field_name not in table_dict:
            logger.debug(f"表格数据中没有字段: {field_name}")
            continue
            
        values_with_format = table_dict[field_name]
        if not values_with_format:
            continue
            
        # 查找占位符位置
        positions = find_placeholder_positions(wb, [placeholder])
        
        if placeholder not in positions or not positions[placeholder]:
            logger.debug(f"未找到占位符位置: {placeholder}")
            continue
        
        # 对每个占位符位置进行列填充
        for position in positions[placeholder]:
            ws = wb[position.sheet_name]
            
            # 从占位符位置开始向下填充数据
            current_row = position.row
            data_index = 0
            
            # 使用while循环实现智能跳过机制
            while data_index < len(values_with_format):
                # 检查当前单元格是否可写入
                if not is_cell_writable(ws, current_row, position.col):
                    # 合并单元格的非左上角：跳过单元格，不消耗数据项
                    current_row += 1
                    continue
                
                # 可写入单元格
                target_cell = ws.cell(row=current_row, column=position.col)
                value, cell_format = values_with_format[data_index]
                
                # 检查单元格是否被保护
                if is_cell_protected(ws, target_cell):
                    # 保护单元格：添加批注，消耗数据项（不填充）
                    add_cell_comment(target_cell, "这是保护单元格不能成功填充", value)
                    data_index += 1
                    current_row += 1
                    warning_msg = f"无法填充 {field_name} 到 {position.sheet_name}[{current_row-1},{position.col}]：单元格只读，消耗了元素: {value}"
                    logger.warning(warning_msg)
                    if warning_msg not in warning_messages:
                        warning_messages.add(warning_msg)
                else:
                    # 正常单元格：填充数据，消耗数据项
                    try:
                        target_cell.value = value
                        
                        # 如果值包含换行符，设置自动换行
                        if isinstance(value, str) and '\n' in value:
                            # 保留原有对齐设置，仅添加换行
                            current_alignment = target_cell.alignment
                            target_cell.alignment = Alignment(
                                horizontal=current_alignment.horizontal if current_alignment else None,
                                vertical=current_alignment.vertical if current_alignment else None,
                                wrap_text=True
                            )
                            logger.debug(f"设置单元格 {target_cell.coordinate} 自动换行")
                        
                        # 设置单元格格式
                        if cell_format:
                            target_cell.number_format = cell_format
                            logger.debug(f"设置单元格 {target_cell.coordinate} 格式为 {cell_format}")
                        
                        # 如果是公式（以等号开头），设置数值格式
                        elif isinstance(value, str) and value.startswith('='):
                            target_cell.number_format = numbers.FORMAT_NUMBER_00
                        
                        filled_count += 1
                        logger.debug(f"列填充: {position.sheet_name}[{current_row},{position.col}] -> {value} (格式: {cell_format})")
                        data_index += 1
                        current_row += 1
                        
                    except (AttributeError, PermissionError) as e:
                        # 处理其他只读情况
                        warning_msg = f"无法填充 {field_name} 到 {position.sheet_name}[{current_row},{position.col}]：单元格只读"
                        logger.warning(warning_msg)
                        if warning_msg not in warning_messages:
                            warning_messages.add(warning_msg)
                        data_index += 1
                        current_row += 1
    
    # 处理行填充占位符
    for placeholder in template_placeholders_dict['row_placeholders']:
        # 提取字段名（去掉→符号）
        field_name = placeholder[3:-2]  # 移除 {{→ 和 }}
        
        if field_name not in table_dict:
            logger.debug(f"表格数据中没有字段: {field_name}")
            continue
            
        values_with_format = table_dict[field_name]
        if not values_with_format:
            continue
            
        # 查找占位符位置
        positions = find_placeholder_positions(wb, [placeholder])
        
        if placeholder not in positions or not positions[placeholder]:
            logger.debug(f"未找到占位符位置: {placeholder}")
            continue
        
        # 对每个占位符位置进行行填充
        for position in positions[placeholder]:
            ws = wb[position.sheet_name]
            
            # 从占位符位置开始向右填充数据
            current_col = position.col
            data_index = 0
            
            # 使用while循环实现智能跳过机制
            while data_index < len(values_with_format):
                # 检查当前单元格是否可写入
                if not is_cell_writable(ws, position.row, current_col):
                    # 合并单元格的非左上角：跳过单元格，不消耗数据项
                    current_col += 1
                    continue
                
                # 可写入单元格
                target_cell = ws.cell(row=position.row, column=current_col)
                value, cell_format = values_with_format[data_index]
                
                # 检查单元格是否被保护
                if is_cell_protected(ws, target_cell):
                    # 保护单元格：添加批注，消耗数据项（不填充）
                    add_cell_comment(target_cell, "这是保护单元格不能成功填充", value)
                    data_index += 1
                    current_col += 1
                    warning_msg = f"无法填充 {field_name} 到 {position.sheet_name}[{position.row},{current_col-1}]：单元格只读，消耗了元素: {value}"
                    logger.warning(warning_msg)
                    if warning_msg not in warning_messages:
                        warning_messages.add(warning_msg)
                else:
                    # 正常单元格：填充数据，消耗数据项
                    try:
                        target_cell.value = value
                        
                        # 如果值包含换行符，设置自动换行
                        if isinstance(value, str) and '\n' in value:
                            # 保留原有对齐设置，仅添加换行
                            current_alignment = target_cell.alignment
                            target_cell.alignment = Alignment(
                                horizontal=current_alignment.horizontal if current_alignment else None,
                                vertical=current_alignment.vertical if current_alignment else None,
                                wrap_text=True
                            )
                            logger.debug(f"设置单元格 {target_cell.coordinate} 自动换行")
                        
                        # 设置单元格格式
                        if cell_format:
                            target_cell.number_format = cell_format
                            logger.debug(f"设置单元格 {target_cell.coordinate} 格式为 {cell_format})")
                        
                        # 如果是公式（以等号开头），设置数值格式
                        elif isinstance(value, str) and value.startswith('='):
                            target_cell.number_format = numbers.FORMAT_NUMBER_00
                        
                        filled_count += 1
                        logger.debug(f"行填充: {position.sheet_name}[{position.row},{current_col}] -> {value} (格式: {cell_format})")
                        data_index += 1
                        current_col += 1
                        
                    except (AttributeError, PermissionError) as e:
                        # 处理其他只读情况
                        warning_msg = f"无法填充 {field_name} 到 {position.sheet_name}[{position.row},{current_col}]：单元格只读"
                        logger.warning(warning_msg)
                        if warning_msg not in warning_messages:
                            warning_messages.add(warning_msg)
                        data_index += 1
                        current_col += 1
    
    logger.info(f"成功填充 {filled_count} 个单元格数据")
    return list(warning_messages)
