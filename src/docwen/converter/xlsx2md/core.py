"""
表格文件 (xlsx, csv) 转 Markdown 核心模块

注意：只处理标准格式（xlsx, csv）
旧格式（xls, et）应在策略层预处理
"""

import os
import pandas as pd
import logging
import openpyxl
from typing import List, Optional, Callable
from docwen.utils.path_utils import generate_output_path
from .image_processor import (
    extract_images_from_xlsx,
    replace_image_markers
)
from docwen.utils.yaml_utils import generate_basic_yaml_frontmatter
from docwen.i18n import t

logger = logging.getLogger(__name__)

def _find_blocks(df: pd.DataFrame) -> List[pd.DataFrame]:
    """
    在DataFrame中查找所有不相连的数据“块”。
    一个“块”被定义为一个周围被空行/空列或DataFrame边缘包围的数据区域。
    """
    if df.empty:
        return []

    rows, cols = df.shape
    visited = set()
    blocks = []

    for r in range(rows):
        for c in range(cols):
            # 检查单元格是否有效：非空且非空字符串
            cell_value = df.iat[r, c]
            if (r, c) not in visited and pd.notna(cell_value) and cell_value != '':
                min_r, max_r, min_c, max_c = r, r, c, c
                
                # 使用广度优先搜索 (BFS) 查找连通块的边界
                q = [(r, c)]
                visited.add((r, c))
                
                while q:
                    curr_r, curr_c = q.pop(0)
                    
                    # 检查8个方向的邻居
                    for dr in [-1, 0, 1]:
                        for dc in [-1, 0, 1]:
                            if dr == 0 and dc == 0:
                                continue
                            
                            nr, nc = curr_r + dr, curr_c + dc
                            
                            if 0 <= nr < rows and 0 <= nc < cols and (nr, nc) not in visited:
                                neighbor_value = df.iat[nr, nc]
                                if pd.notna(neighbor_value) and neighbor_value != '':
                                    visited.add((nr, nc))
                                    q.append((nr, nc))
                                    min_r = min(min_r, nr)
                                    max_r = max(max_r, nr)
                                    min_c = min(min_c, nc)
                                    max_c = max(max_c, nc)
                
                # 提取块并清理
                block = df.iloc[min_r:max_r+1, min_c:max_c+1].copy()
                block.dropna(axis=0, how='all', inplace=True)
                block.dropna(axis=1, how='all', inplace=True)
                blocks.append(block)
                
    return blocks

def _fill_merged_cells(ws):
    """
    填充合并单元格：将合并区域的所有单元格都设置为左上角单元格的值
    """
    # 获取所有合并单元格区域
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        # 获取合并区域的边界
        min_row, min_col, max_row, max_col = (
            merged_range.min_row, merged_range.min_col, 
            merged_range.max_row, merged_range.max_col
        )
        
        # 获取左上角单元格的值
        top_left_value = ws.cell(row=min_row, column=min_col).value
        
        # 先取消合并单元格
        ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        
        # 将合并区域的所有单元格都设置为左上角的值
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value

def _worksheet_to_dataframe(ws, sheet_images: List[dict] = None):
    """
    将openpyxl工作表转换为DataFrame，并处理合并单元格
    
    参数:
        ws: openpyxl工作表对象
        sheet_images: 该工作表的图片信息列表（可选）
    
    返回:
        pd.DataFrame: 转换后的DataFrame
    """
    # 先填充合并单元格
    _fill_merged_cells(ws)
    
    # 将工作表数据转换为二维列表
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    # 转换为DataFrame
    df = pd.DataFrame(data)
    
    # 如果有图片信息，在对应单元格插入图片标记
    if sheet_images:
        # 首先计算需要的最大行和列，以便扩展DataFrame
        max_row_needed = len(df) - 1  # 当前最大行索引（从0开始）
        max_col_needed = len(df.columns) - 1 if len(df.columns) > 0 else -1  # 当前最大列索引
        
        for img_info in sheet_images:
            row = img_info.get('row')
            col = img_info.get('col')
            if row is not None and col is not None:
                # 转换为DataFrame索引（从1开始转为从0开始）
                df_row = row - 1
                df_col = col - 1
                max_row_needed = max(max_row_needed, df_row)
                max_col_needed = max(max_col_needed, df_col)
        
        # 扩展DataFrame以容纳所有图片位置
        current_rows = len(df)
        current_cols = len(df.columns) if len(df.columns) > 0 else 0
        
        # 如果需要更多行
        if max_row_needed >= current_rows:
            rows_to_add = max_row_needed - current_rows + 1
            # 创建空行并添加到DataFrame
            empty_rows = pd.DataFrame(
                [[None] * max(current_cols, max_col_needed + 1) for _ in range(rows_to_add)]
            )
            if current_cols > 0:
                empty_rows.columns = df.columns if len(df.columns) == max(current_cols, max_col_needed + 1) else range(max(current_cols, max_col_needed + 1))
            df = pd.concat([df, empty_rows], ignore_index=True)
            logger.debug(f"扩展DataFrame行数: {current_rows} -> {len(df)}")
        
        # 如果需要更多列
        if max_col_needed >= current_cols:
            cols_to_add = max_col_needed - current_cols + 1
            for i in range(cols_to_add):
                df[current_cols + i] = None
            logger.debug(f"扩展DataFrame列数: {current_cols} -> {len(df.columns)}")
        
        # 现在插入图片标记
        for img_info in sheet_images:
            row = img_info.get('row')
            col = img_info.get('col')
            filename = img_info.get('filename')
            
            if row is not None and col is not None and filename:
                # 转换为DataFrame索引（从1开始转为从0开始）
                df_row = row - 1
                df_col = col - 1
                
                # 检查索引是否在DataFrame范围内（扩展后应该都在范围内）
                if 0 <= df_row < len(df) and 0 <= df_col < len(df.columns):
                    # 在单元格中插入图片标记
                    original_value = df.iat[df_row, df_col]
                    # 如果单元格有内容，保留原内容并在后面加图片
                    if pd.notna(original_value) and str(original_value).strip():
                        df.iat[df_row, df_col] = f"{original_value} {{{{IMAGE:{filename}}}}}"
                    else:
                        # 否则直接插入图片标记
                        df.iat[df_row, df_col] = f"{{{{IMAGE:{filename}}}}}"
                    logger.debug(f"在单元格({row},{col})插入图片标记: {filename}")
                else:
                    logger.warning(f"图片位置超出DataFrame范围: 行{row}, 列{col}, DataFrame大小: {len(df)}x{len(df.columns)}")
    
    return df

def _process_cell_newlines(df: pd.DataFrame) -> pd.DataFrame:
    """
    处理DataFrame中单元格的换行符，替换为<br>标签
    同时转义管道符，避免破坏Markdown表格结构
    
    参数:
        df: 输入的DataFrame
    
    返回:
        处理后的DataFrame副本
    """
    df_copy = df.copy()
    
    for col in df_copy.columns:
        df_copy[col] = df_copy[col].apply(
            lambda x: (
                str(x)
                .replace('\r\n', '<br>')
                .replace('\n', '<br>')
                .replace('\r', '<br>')
                .replace('|', '\\|')  # 转义管道符
            ) if pd.notna(x) and x != '' else x
        )
    
    return df_copy

def convert_spreadsheet_to_md(
    file_path: str,
    extract_image: bool = True,
    extract_ocr: bool = False,
    output_folder: Optional[str] = None,
    original_file_path: Optional[str] = None,
    progress_callback: Optional[Callable[[str], None]] = None,
    cancel_event=None
) -> str:
    """
    将标准表格文件转换为Markdown格式
    
    支持格式：xlsx, csv
    
    参数:
        file_path: 输入文件路径（必须是标准格式，可能是临时副本）
        extract_image: 是否提取图片（默认True）
        extract_ocr: 是否进行OCR识别（默认False）
        output_folder: 输出文件夹路径，用于保存图片（可选）
        original_file_path: 原始文件路径（用于图片命名，可选）
    
    返回:
        str: Markdown格式的文档内容
    
    说明:
        - 生成YAML front matter（标题、别名）
        - 文件名作为YAML的标题和第一个别名
        - xlsx: 每个工作表一个一级标题
        - csv: 文件名作为一级标题
        - 工作表内的每个数据"块"是一个独立的Markdown表格
        - 支持提取图片并根据选项处理（类似文档转MD）
    
    注意:
        旧格式（xls, et）应在策略层预处理为 xlsx
    """
    # 优先使用原始文件路径，用于生成正确的标题和别名
    path_for_naming = original_file_path or file_path
    file_name = os.path.basename(path_for_naming)
    file_stem, file_ext = os.path.splitext(file_name)
    file_ext = file_ext.lower()

    logger.info(f"开始转换表格文件: {file_path}")
    logger.info(f"导出选项 - 提取图片: {extract_image}, OCR识别: {extract_ocr}")

    try:
        # 生成YAML front matter（使用原始文件名）
        md_content = generate_basic_yaml_frontmatter(file_stem)

        if file_ext == '.csv':
            # CSV格式不支持图片
            if (extract_image or extract_ocr):
                logger.info("CSV格式不支持图片提取")
            
            # 添加文件名作为一级标题（使用原始文件名）
            md_content += f"# {file_stem}\n\n"
            
            df = pd.read_csv(file_path, header=None, keep_default_na=False)
            blocks = _find_blocks(df)
            for i, block in enumerate(blocks):
                # 处理单元格换行符
                block = _process_cell_newlines(block)
                
                # 使用第一行作为表头
                if block.shape[0] > 0:
                    headers = block.iloc[0].tolist()
                    data = block.iloc[1:]
                    md_content += data.to_markdown(index=False, headers=headers)
                else:
                    md_content += block.to_markdown(index=False, headers=[] if block.shape[1] > 0 and pd.isna(block.columns).all() else 'keys')
                md_content += "\n\n"
        else:  # xlsx
            # 加载工作簿（只加载一次，避免重复I/O）
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # 提取图片（如果需要）
            images_info = []
            if (extract_image or extract_ocr) and output_folder:
                if progress_callback:
                    progress_callback(t('conversion.progress.extracting_images_start'))
                
                logger.info("开始从XLSX提取图片...")
                # 优先使用原始文件路径进行图片命名，如未提供则使用file_path
                path_for_naming = original_file_path or file_path
                # 传递workbook对象，避免重复加载文件
                images_info = extract_images_from_xlsx(
                    wb, output_folder, path_for_naming,
                    progress_callback=progress_callback
                )
                logger.info(f"图片提取完成，共 {len(images_info)} 张")
                
                if progress_callback:
                    progress_callback(t('conversion.progress.images_extracted', count=len(images_info)))
            
            # 创建图片字典，按工作表分组
            images_by_sheet = {}
            for img_info in images_info:
                sheet = img_info.get('sheet_name')
                if sheet:
                    if sheet not in images_by_sheet:
                        images_by_sheet[sheet] = []
                    images_by_sheet[sheet].append(img_info)
            
            for sheet_name in wb.sheetnames:
                md_content += f"# {sheet_name}\n\n"
                ws = wb[sheet_name]
                
                # 获取该工作表的图片信息
                sheet_images = images_by_sheet.get(sheet_name, [])
                
                # 转换DataFrame并插入图片标记
                df = _worksheet_to_dataframe(ws, sheet_images)
                blocks = _find_blocks(df)
                
                if not blocks:
                    md_content += " (此工作表为空)\n\n"
                else:
                    for i, block in enumerate(blocks):
                        # 处理单元格换行符
                        block = _process_cell_newlines(block)
                        
                        # 使用第一行作为表头
                        if block.shape[0] > 0:
                            headers = block.iloc[0].tolist()
                            data = block.iloc[1:]
                            block_md = data.to_markdown(index=False, headers=headers)
                        else:
                            block_md = block.to_markdown(index=False, headers=[] if block.shape[1] > 0 and pd.isna(block.columns).all() else 'keys')
                        
                        # 替换图片标记为实际链接
                        block_md = replace_image_markers(
                            block_md, sheet_images, extract_image, extract_ocr, output_folder,
                            progress_callback=progress_callback,
                            cancel_event=cancel_event
                        )
                        
                        md_content += block_md
                        md_content += "\n\n"

        return md_content.strip()

    except Exception as e:
        logger.error(f"处理表格文件 {file_name} 失败: {e}", exc_info=True)
        raise
