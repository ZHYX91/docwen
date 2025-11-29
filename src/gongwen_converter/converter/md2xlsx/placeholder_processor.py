"""
占位符处理模块
处理Excel模板中的占位符提取、替换和清理
"""

import re
import logging
from openpyxl.styles import numbers
from copy import copy
from collections import namedtuple
from gongwen_converter.utils.text_utils import clean_text

# 定义占位符位置命名元组
PlaceholderPosition = namedtuple('PlaceholderPosition', ['sheet_name', 'row', 'col', 'is_whole_cell'])

# 配置日志
logger = logging.getLogger(__name__)


def _is_fraction(s: str) -> bool:
    """检查字符串是否为分数，例如 '1/2' 或 '-1/2'"""
    if not isinstance(s, str):
        return False
    
    s_to_check = s
    # 处理可选的前导符号
    if s.startswith(('+', '-')):
        s_to_check = s[1:]
        
    if s_to_check.count('/') == 1:
        parts = s_to_check.split('/')
        # 确保分子和分母都是数字且不为空
        return len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit()
    return False


def extract_template_placeholders(template_path: str) -> dict:
    """
    从Excel模板的所有工作表中提取占位符（符号方案）
    参数:
        template_path: Excel模板路径
    返回:
        dict: 分类后的占位符字典
        {
            'yaml_placeholders': ['{{字段名}}'],
            'column_placeholders': ['{{↓字段名}}'], 
            'row_placeholders': ['{{→字段名}}']
        }
    """
    from openpyxl import load_workbook
    
    logger.debug(f"从模板提取占位符: {template_path}")
    wb = load_workbook(template_path)
    yaml_placeholders = set()
    column_placeholders = set()
    row_placeholders = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # 使用正则匹配所有 {{...}} 模式
                    matches = re.findall(r'{{([^}]+)}}', cell.value)
                    for match in matches:
                        placeholder = f"{{{{{match}}}}}"
                        # 分类占位符
                        if match.startswith('↓'):
                            column_placeholders.add(placeholder)
                        elif match.startswith('→'):
                            row_placeholders.add(placeholder)
                        else:
                            yaml_placeholders.add(placeholder)
    
    wb.close()
    
    result = {
        'yaml_placeholders': list(yaml_placeholders),
        'column_placeholders': list(column_placeholders),
        'row_placeholders': list(row_placeholders)
    }
    
    logger.info(f"从模板中提取到占位符: YAML={len(yaml_placeholders)}, 列填充={len(column_placeholders)}, 行填充={len(row_placeholders)}")
    logger.debug(f"YAML占位符: {yaml_placeholders}")
    logger.debug(f"列填充占位符: {column_placeholders}")
    logger.debug(f"行填充占位符: {row_placeholders}")
    
    return result


def create_replacements(data, template_placeholders_dict):
    """
    基于模板占位符创建替换配置（符号方案）
    参数:
        data: YAML数据字典
        template_placeholders_dict: 分类后的占位符字典
    返回:
        dict: {占位符: (处理后的值, 单元格格式)}, 仅包含YAML中存在的字段
    """
    logger.debug("创建基于模板的字段替换配置...")
    replacements = {}
    
    # 只处理YAML占位符
    for placeholder in template_placeholders_dict.get('yaml_placeholders', []):
        field_name = placeholder[2:-2]
        
        if field_name in data:
            value = data[field_name]
            processed_value = ""
            cell_format = None
            
            if value is None:
                processed_value = ""
            elif isinstance(value, str):
                if _is_fraction(value):
                    processed_value = f"={value}"
                elif value.isdigit():
                    # 字符串形式的数字或长数字字符串，设置为文本格式
                    processed_value = value
                    cell_format = numbers.FORMAT_TEXT
                else:
                    processed_value = clean_text(str(value)) # 普通字符串
            elif isinstance(value, (int, float)):
                # 如果纯数字位数过长，也设置为文本格式
                if len(str(value).replace('.', '')) > 15:
                    processed_value = value
                    cell_format = numbers.FORMAT_TEXT
                else:
                    processed_value = value # 普通数字
            elif isinstance(value, list):
                cleaned_items = [clean_text(str(item)) for item in value]
                processed_value = "、".join(cleaned_items)
            else:
                processed_value = clean_text(str(value))
            
            replacements[placeholder] = (processed_value, cell_format)
            logger.debug(f"添加替换项: {placeholder} -> ({processed_value}, {cell_format})")
    
    logger.info(f"生成 {len(replacements)} 个YAML字段替换项")
    return replacements


def find_placeholder_positions(wb, template_placeholders):
    """
    在Excel模板中查找所有占位符的位置（符号方案）
    参数:
        wb: 工作簿对象
        template_placeholders: 占位符列表
    返回:
        dict: {占位符: [PlaceholderPosition, ...]}
    """
    logger.debug("查找占位符位置...")
    positions = {}
    
    for placeholder in template_placeholders:
        positions[placeholder] = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    for placeholder in template_placeholders:
                        if placeholder in cell.value:
                            # 检查占位符是否占据整个单元格
                            is_whole_cell = (cell.value.strip() == placeholder)
                            position = PlaceholderPosition(
                                sheet_name=sheet_name,
                                row=row_idx,
                                col=col_idx,
                                is_whole_cell=is_whole_cell
                            )
                            positions[placeholder].append(position)
                            logger.debug(f"找到占位符位置: {placeholder} -> {sheet_name}[{row_idx},{col_idx}] (完整单元格: {is_whole_cell})")
    
    return positions



def process_cell_replacements(wb, replacements):
    """
    在工作簿的所有工作表中执行占位符替换
    参数:
        wb: 工作簿对象
        replacements: 替换配置字典 {placeholder: (value, format)}
    """
    logger.debug("处理所有工作表的单元格占位符替换...")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.debug(f"处理工作表: {sheet_name}")
        
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue

                original_value = cell.value
                
                # 优化：处理完整的单元格占位符
                is_single_placeholder = original_value.strip().startswith('{{') and original_value.strip().endswith('}}')
                
                if is_single_placeholder:
                    placeholder = original_value.strip()
                    if placeholder in replacements:
                        value_to_replace, cell_format = replacements[placeholder]
                        cell.value = value_to_replace
                        if cell_format:
                            cell.number_format = cell_format
                            logger.debug(f"设置单元格 {cell.coordinate} 格式为 {cell_format}")
                        logger.debug(f"完整替换单元格 {cell.coordinate}: {original_value[:50]} -> {str(value_to_replace)[:50]}")
                        continue  # 处理完，进入下一个单元格

                # 处理部分替换或混合内容的单元格
                new_value_str = original_value
                cell_modified = False
                final_format = None
                
                for placeholder, (value_to_replace, cell_format) in replacements.items():
                    if placeholder in new_value_str:
                        new_value_str = new_value_str.replace(placeholder, str(value_to_replace))
                        cell_modified = True
                        if cell_format == numbers.FORMAT_TEXT:
                            final_format = numbers.FORMAT_TEXT  # 只要有一个替换项需要文本，整个单元格就设为文本

                if cell_modified:
                    cell.value = new_value_str
                    if final_format:
                         cell.number_format = final_format
                         logger.debug(f"设置单元格 {cell.coordinate} 格式为 {final_format}")
                    logger.debug(f"部分替换单元格 {cell.coordinate}: {original_value[:50]} -> {new_value_str[:50]}")

    logger.info("所有工作表的单元格占位符替换完成")


def clean_remaining_placeholders(wb):
    """
    清理工作簿中所有剩余的占位符，将其替换为空字符串
    参数:
        wb: 工作簿对象
    """
    logger.debug("开始清理剩余占位符...")
    cleaned_count = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.debug(f"清理工作表: {sheet_name}")
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # 使用正则替换所有 {{...}} 模式为空字符串
                    original_text = cell.value
                    cleaned_text = re.sub(r'\{\{[^}]+\}\}', '', original_text)
                    
                    if cleaned_text != original_text:
                        cell.value = cleaned_text
                        cleaned_count += 1
                        logger.debug(f"清理占位符: {original_text} -> {cleaned_text}")
    
    logger.info(f"成功清理 {cleaned_count} 个剩余占位符")


def process_image_placeholders(wb):
    """
    处理工作簿中的图片占位符 {{IMAGE:路径}}
    
    处理流程：
    1. 遍历所有工作表的单元格
    2. 查找包含 {{IMAGE:路径}} 的单元格
    3. 提取图片路径
    4. 检查图片文件是否存在
    5. 在单元格位置插入图片
    6. 清空单元格中的占位符文本
    
    参数:
        wb: Workbook对象
    
    返回:
        int: 成功插入的图片数量
    """
    import os
    from openpyxl.drawing.image import Image
    
    logger.info("开始处理图片占位符...")
    
    # 图片占位符正则：{{IMAGE:路径}}
    image_pattern = r'\{\{IMAGE:([^}]+)\}\}'
    inserted_count = 0
    
    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.debug(f"处理工作表: {sheet_name}")
        
        # 遍历所有单元格
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                
                # 查找图片占位符
                matches = re.findall(image_pattern, cell.value)
                if not matches:
                    continue
                
                logger.debug(f"在单元格 {cell.coordinate} 找到 {len(matches)} 个图片占位符")
                
                # 处理每个图片占位符
                for image_path in matches:
                    placeholder = f"{{{{IMAGE:{image_path}}}}}"
                    
                    # 检查图片文件是否存在
                    if not os.path.exists(image_path):
                        logger.warning(f"图片文件不存在，跳过: {image_path}")
                        # 删除占位符
                        cell.value = cell.value.replace(placeholder, "")
                        continue
                    
                    # 检查文件扩展名
                    ext = os.path.splitext(image_path)[1].lower()
                    supported_formats = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff'}
                    if ext not in supported_formats:
                        logger.warning(f"不支持的图片格式: {ext}，跳过: {image_path}")
                        # 删除占位符
                        cell.value = cell.value.replace(placeholder, "")
                        continue
                    
                    try:
                        # 创建图片对象
                        img = Image(image_path)
                        
                        # 调整图片大小（最大宽度200像素，保持宽高比）
                        max_width = 200
                        if img.width > max_width:
                            scale = max_width / img.width
                            img.width = max_width
                            img.height = int(img.height * scale)
                        
                        # 将图片锚定到单元格
                        # 格式：单元格坐标（如 'A1'）
                        img.anchor = cell.coordinate
                        
                        # 添加图片到工作表
                        ws.add_image(img)
                        
                        # 清空单元格文本（删除占位符）
                        cell.value = cell.value.replace(placeholder, "")
                        
                        # 调整单元格行高（为图片留出空间）
                        row_height = img.height * 0.75  # 像素转点（约）
                        if ws.row_dimensions[cell.row].height is None or ws.row_dimensions[cell.row].height < row_height:
                            ws.row_dimensions[cell.row].height = row_height
                            logger.debug(f"调整行 {cell.row} 高度为 {row_height:.2f}")
                        
                        # 调整单元格列宽（为图片留出空间）
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(cell.column)
                        col_width = img.width / 7.0  # 像素转字符宽度（1字符≈7像素）
                        
                        # 只在当前列宽不足时调整
                        current_width = ws.column_dimensions[col_letter].width
                        if current_width is None or current_width < col_width:
                            ws.column_dimensions[col_letter].width = col_width
                            logger.debug(f"调整列 {col_letter} 宽度为 {col_width:.2f}")
                        
                        inserted_count += 1
                        logger.info(f"成功插入图片 ({inserted_count}) 到 {cell.coordinate}: {os.path.basename(image_path)}")
                        
                    except Exception as e:
                        logger.error(f"插入图片失败: {image_path} | 错误: {str(e)}")
                        # 删除占位符，避免后续清理时出错
                        cell.value = cell.value.replace(placeholder, "")
    
    logger.info(f"图片占位符处理完成 | 成功插入: {inserted_count} 个")
    return inserted_count
