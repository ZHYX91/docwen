"""
表格文件 (xlsx, csv) 转 Markdown 核心模块

注意：只处理标准格式（xlsx, csv）
旧格式（xls, et）应在策略层预处理
"""

import os
import pandas as pd
import logging
import openpyxl
from typing import List, Optional
from gongwen_converter.utils.path_utils import generate_output_path

logger = logging.getLogger(__name__)

def _find_blocks(df: pd.DataFrame) -> List[pd.DataFrame]:
    """
    在DataFrame中查找所有不相连的数据“块”。
    一个“块”被定义为一个周围被空行/空列或DataFrame边缘包围的数据区域。
    """
    if df.empty:
        return []

    rows, cols = df.shape
    visited = set()
    blocks = []

    for r in range(rows):
        for c in range(cols):
            # 检查单元格是否有效：非空且非空字符串
            cell_value = df.iat[r, c]
            if (r, c) not in visited and pd.notna(cell_value) and cell_value != '':
                min_r, max_r, min_c, max_c = r, r, c, c
                
                # 使用广度优先搜索 (BFS) 查找连通块的边界
                q = [(r, c)]
                visited.add((r, c))
                
                while q:
                    curr_r, curr_c = q.pop(0)
                    
                    # 检查8个方向的邻居
                    for dr in [-1, 0, 1]:
                        for dc in [-1, 0, 1]:
                            if dr == 0 and dc == 0:
                                continue
                            
                            nr, nc = curr_r + dr, curr_c + dc
                            
                            if 0 <= nr < rows and 0 <= nc < cols and (nr, nc) not in visited:
                                neighbor_value = df.iat[nr, nc]
                                if pd.notna(neighbor_value) and neighbor_value != '':
                                    visited.add((nr, nc))
                                    q.append((nr, nc))
                                    min_r = min(min_r, nr)
                                    max_r = max(max_r, nr)
                                    min_c = min(min_c, nc)
                                    max_c = max(max_c, nc)
                
                # 提取块并清理
                block = df.iloc[min_r:max_r+1, min_c:max_c+1].copy()
                block.dropna(axis=0, how='all', inplace=True)
                block.dropna(axis=1, how='all', inplace=True)
                blocks.append(block)
                
    return blocks

def _fill_merged_cells(ws):
    """
    填充合并单元格：将合并区域的所有单元格都设置为左上角单元格的值
    """
    # 获取所有合并单元格区域
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        # 获取合并区域的边界
        min_row, min_col, max_row, max_col = (
            merged_range.min_row, merged_range.min_col, 
            merged_range.max_row, merged_range.max_col
        )
        
        # 获取左上角单元格的值
        top_left_value = ws.cell(row=min_row, column=min_col).value
        
        # 先取消合并单元格
        ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        
        # 将合并区域的所有单元格都设置为左上角的值
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value

def _worksheet_to_dataframe(ws, sheet_images: List[dict] = None):
    """
    将openpyxl工作表转换为DataFrame，并处理合并单元格
    
    参数:
        ws: openpyxl工作表对象
        sheet_images: 该工作表的图片信息列表（可选）
    
    返回:
        pd.DataFrame: 转换后的DataFrame
    """
    # 先填充合并单元格
    _fill_merged_cells(ws)
    
    # 将工作表数据转换为二维列表
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    # 转换为DataFrame
    df = pd.DataFrame(data)
    
    # 如果有图片信息，在对应单元格插入图片标记
    if sheet_images:
        for img_info in sheet_images:
            row = img_info.get('row')
            col = img_info.get('col')
            filename = img_info.get('filename')
            
            if row is not None and col is not None and filename:
                # 转换为DataFrame索引（从1开始转为从0开始）
                df_row = row - 1
                df_col = col - 1
                
                # 检查索引是否在DataFrame范围内
                if 0 <= df_row < len(df) and 0 <= df_col < len(df.columns):
                    # 在单元格中插入图片标记
                    original_value = df.iat[df_row, df_col]
                    # 如果单元格有内容，保留原内容并在后面加图片
                    if pd.notna(original_value) and str(original_value).strip():
                        df.iat[df_row, df_col] = f"{original_value} {{{{IMAGE:{filename}}}}}"
                    else:
                        # 否则直接插入图片标记
                        df.iat[df_row, df_col] = f"{{{{IMAGE:{filename}}}}}"
                    logger.debug(f"在单元格({row},{col})插入图片标记: {filename}")
    
    return df

def _process_cell_newlines(df: pd.DataFrame) -> pd.DataFrame:
    """
    处理DataFrame中单元格的换行符，替换为<br>标签
    同时转义管道符，避免破坏Markdown表格结构
    
    参数:
        df: 输入的DataFrame
    
    返回:
        处理后的DataFrame副本
    """
    df_copy = df.copy()
    
    for col in df_copy.columns:
        df_copy[col] = df_copy[col].apply(
            lambda x: (
                str(x)
                .replace('\r\n', '<br>')
                .replace('\n', '<br>')
                .replace('\r', '<br>')
                .replace('|', '\\|')  # 转义管道符
            ) if pd.notna(x) and x != '' else x
        )
    
    return df_copy

def convert_spreadsheet_to_md(
    file_path: str,
    extract_image: bool = True,
    extract_ocr: bool = False,
    output_folder: Optional[str] = None,
    original_file_path: Optional[str] = None
) -> str:
    """
    将标准表格文件转换为Markdown格式
    
    支持格式：xlsx, csv
    
    参数:
        file_path: 输入文件路径（必须是标准格式，可能是临时副本）
        extract_image: 是否提取图片（默认True）
        extract_ocr: 是否进行OCR识别（默认False）
        output_folder: 输出文件夹路径，用于保存图片（可选）
        original_file_path: 原始文件路径（用于图片命名，可选）
    
    返回:
        str: Markdown格式的文档内容
    
    说明:
        - 文件名作为一级标题
        - 每个工作表一个二级标题（CSV除外）
        - 工作表内的每个数据"块"是一个独立的Markdown表格
        - 支持提取图片并根据选项处理（类似文档转MD）
    
    注意:
        旧格式（xls, et）应在策略层预处理为 xlsx
    """
    file_name = os.path.basename(file_path)
    file_stem, file_ext = os.path.splitext(file_name)
    file_ext = file_ext.lower()

    # 只处理标准格式
    if file_ext not in ['.xlsx', '.csv']:
        raise ValueError(
            f"不支持的文件格式: {file_ext}\n"
            f"仅支持标准格式（.xlsx, .csv）\n"
            f"旧格式（.xls, .et）应在策略层预处理"
        )

    logger.info(f"开始转换表格文件: {file_path}")
    logger.info(f"导出选项 - 提取图片: {extract_image}, OCR识别: {extract_ocr}")

    try:
        md_content = f"# {file_stem}\n\n"

        if file_ext == '.csv':
            # CSV格式不支持图片
            if (extract_image or extract_ocr):
                logger.info("CSV格式不支持图片提取")
            
            df = pd.read_csv(file_path, header=None, keep_default_na=False)
            blocks = _find_blocks(df)
            for i, block in enumerate(blocks):
                # 处理单元格换行符
                block = _process_cell_newlines(block)
                
                # 使用第一行作为表头
                if block.shape[0] > 0:
                    headers = block.iloc[0].tolist()
                    data = block.iloc[1:]
                    md_content += data.to_markdown(index=False, headers=headers)
                else:
                    md_content += block.to_markdown(index=False, headers=[] if block.shape[1] > 0 and pd.isna(block.columns).all() else 'keys')
                md_content += "\n\n"
        else:  # xlsx
            # 加载工作簿（只加载一次，避免重复I/O）
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # 提取图片（如果需要）
            images_info = []
            if (extract_image or extract_ocr) and output_folder:
                logger.info("开始从XLSX提取图片...")
                # 优先使用原始文件路径进行图片命名，如未提供则使用file_path
                path_for_naming = original_file_path or file_path
                # 传递workbook对象，避免重复加载文件
                images_info = _extract_images_from_xlsx(wb, output_folder, path_for_naming)
                logger.info(f"图片提取完成，共 {len(images_info)} 张")
            
            # 创建图片字典，按工作表分组
            images_by_sheet = {}
            for img_info in images_info:
                sheet = img_info.get('sheet_name')
                if sheet:
                    if sheet not in images_by_sheet:
                        images_by_sheet[sheet] = []
                    images_by_sheet[sheet].append(img_info)
            
            for sheet_name in wb.sheetnames:
                md_content += f"## {sheet_name}\n\n"
                ws = wb[sheet_name]
                
                # 获取该工作表的图片信息
                sheet_images = images_by_sheet.get(sheet_name, [])
                
                # 转换DataFrame并插入图片标记
                df = _worksheet_to_dataframe(ws, sheet_images)
                blocks = _find_blocks(df)
                
                if not blocks:
                    md_content += " (此工作表为空)\n\n"
                else:
                    for i, block in enumerate(blocks):
                        # 处理单元格换行符
                        block = _process_cell_newlines(block)
                        
                        # 使用第一行作为表头
                        if block.shape[0] > 0:
                            headers = block.iloc[0].tolist()
                            data = block.iloc[1:]
                            block_md = data.to_markdown(index=False, headers=headers)
                        else:
                            block_md = block.to_markdown(index=False, headers=[] if block.shape[1] > 0 and pd.isna(block.columns).all() else 'keys')
                        
                        # 替换图片标记为实际链接
                        block_md = _replace_image_markers(block_md, sheet_images, extract_image, extract_ocr, output_folder)
                        
                        md_content += block_md
                        md_content += "\n\n"

        return md_content.strip()

    except Exception as e:
        logger.error(f"处理表格文件 {file_name} 失败: {e}", exc_info=True)
        raise


def _extract_images_from_xlsx(workbook, output_folder: str, original_file_path: str) -> List[dict]:
    """
    从XLSX工作簿对象中提取所有图片
    
    设计说明：
        此函数采用与文档图片提取一致的设计模式，接收已加载的工作簿对象，
        实现了职责分离（文件加载 vs 图片提取）和性能优化（避免重复I/O）。
    
    参数:
        workbook: openpyxl的Workbook对象（已加载）
        output_folder: 图片保存的文件夹路径
        original_file_path: 原始文件路径（用于生成规范的图片文件名）
    
    返回:
        List[dict]: 图片信息列表，每个元素包含：
            - filename: 图片文件名（如 '销售数据_image1_20250125_091530_fromXlsx.png'）
            - image_path: 图片完整路径
            - sheet_name: 所在工作表名称
            - row: 图片所在行号（可选）
            - col: 图片所在列号（可选）
    """
    images_info = []
    image_counter = 1
    
    logger.info(f"开始从XLSX工作簿提取图片到: {output_folder}")
    
    # 确保输出文件夹存在
    os.makedirs(output_folder, exist_ok=True)
    
    try:
        # 使用传入的workbook对象，避免重复加载
        wb = workbook
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 检查工作表中的图片
            if hasattr(ws, '_images') and ws._images:
                logger.debug(f"工作表 '{sheet_name}' 包含 {len(ws._images)} 张图片")
                
                for img in ws._images:
                    try:
                        # 获取图片数据
                        image_data = img._data()
                        
                        # 确定图片格式
                        # openpyxl的图片对象可能有format属性
                        ext = 'png'  # 默认格式
                        if hasattr(img, 'format'):
                            ext = img.format.lower()
                        elif hasattr(img, '_format'):
                            ext = img._format.lower()
                        
                        # 使用统一的路径生成工具生成规范文件名
                        # 使用original_file_path进行命名，确保文件名基于原始文件而非临时副本
                        image_path = generate_output_path(
                            input_path=original_file_path,
                            output_dir=output_folder,
                            section=f"image{image_counter}",
                            add_timestamp=True,
                            description="fromXlsx",
                            file_type=ext
                        )
                        filename = os.path.basename(image_path)
                        
                        # 保存图片
                        with open(image_path, 'wb') as f:
                            f.write(image_data)
                        
                        # 获取图片位置
                        row, col = None, None
                        if hasattr(img, 'anchor') and img.anchor:
                            if hasattr(img.anchor, '_from'):
                                row = img.anchor._from.row + 1  # openpyxl从0开始，转为从1开始
                                col = img.anchor._from.col + 1
                        
                        logger.debug(f"提取图片: {filename} (工作表 '{sheet_name}', 行{row}, 列{col})")
                        
                        # 记录图片信息（包括位置）
                        img_info = {
                            'filename': filename,
                            'image_path': image_path,
                            'sheet_name': sheet_name
                        }
                        if row is not None and col is not None:
                            img_info['row'] = row
                            img_info['col'] = col
                        
                        images_info.append(img_info)
                        
                        image_counter += 1
                        
                    except Exception as e:
                        logger.warning(f"提取图片失败 (工作表 '{sheet_name}'): {e}")
                        continue
        
        logger.info(f"XLSX图片提取完成，共提取 {len(images_info)} 张图片")
        
    except Exception as e:
        logger.error(f"提取XLSX图片时出错: {e}", exc_info=True)
    
    return images_info


def _process_image_with_ocr(img: dict, keep_images: bool, enable_ocr: bool, output_folder: str) -> str:
    """
    根据选项处理图片，生成对应的Markdown链接（与文档转MD逻辑一致）
    
    参数:
        img: 图片信息字典，包含 'filename', 'image_path' 等
        keep_images: 是否保留图片
        enable_ocr: 是否启用OCR识别
        output_folder: 输出文件夹路径
    
    返回:
        str: Markdown链接（图片链接或图片md文件链接）
    """
    from gongwen_converter.config.config_manager import config_manager
    from gongwen_converter.utils.markdown_utils import format_image_link, format_md_file_link
    
    filename = img['filename']
    image_path = img['image_path']
    
    logger.debug(f"处理表格图片: {filename}, 保留图片: {keep_images}, OCR: {enable_ocr}")
    
    # 获取链接格式配置
    link_settings = config_manager.get_markdown_link_style_settings()
    image_format = link_settings.get("image_link_format", "wiki")
    image_embed = link_settings.get("image_embed", True)
    md_file_format = link_settings.get("md_file_link_format", "wiki")
    md_file_embed = link_settings.get("md_file_embed", True)
    
    # 场景1：只勾选提取图片
    if keep_images and not enable_ocr:
        logger.info(f"场景1：只保留图片 - {filename}")
        return format_image_link(filename, image_format, image_embed)
    
    # 场景2：只勾选OCR
    elif not keep_images and enable_ocr:
        logger.info(f"场景2：只OCR识别 - {filename}")
        # 创建图片md文件（只包含OCR文本）
        md_filename = _create_image_md_file(image_path, filename, output_folder, 
                                            include_image=False, include_ocr=True)
        return format_md_file_link(md_filename, md_file_format, md_file_embed)
    
    # 场景3：两者都勾选
    elif keep_images and enable_ocr:
        logger.info(f"场景3：图片 + OCR - {filename}")
        # 创建图片md文件（包含图片链接和OCR文本）
        md_filename = _create_image_md_file(image_path, filename, output_folder, 
                                            include_image=True, include_ocr=True)
        return format_md_file_link(md_filename, md_file_format, md_file_embed)
    
    # 都不勾选（不应该出现这种情况，但作为兜底）
    else:
        logger.warning(f"都不勾选时不应该调用此函数 - {filename}")
        return format_image_link(filename, image_format, image_embed)


def _replace_image_markers(markdown_text: str, sheet_images: List[dict], 
                           keep_images: bool, enable_ocr: bool, output_folder: str) -> str:
    """
    替换Markdown文本中的图片标记为实际的图片链接
    
    参数:
        markdown_text: 包含图片标记的Markdown文本
        sheet_images: 工作表的图片信息列表
        keep_images: 是否保留图片
        enable_ocr: 是否启用OCR
        output_folder: 输出文件夹路径
    
    返回:
        str: 替换后的Markdown文本
    """
    import re
    
    # 查找所有图片标记: {{IMAGE:filename}}
    for img_info in sheet_images:
        filename = img_info.get('filename')
        if not filename:
            continue
        
        # 创建标记模式
        marker = f"{{{{IMAGE:{filename}}}}}"
        
        # 如果Markdown中包含这个标记，替换为实际链接
        if marker in markdown_text:
            image_link = _process_image_with_ocr(img_info, keep_images, enable_ocr, output_folder)
            markdown_text = markdown_text.replace(marker, image_link)
            logger.debug(f"替换图片标记: {marker} -> {image_link}")
    
    return markdown_text


def _create_image_md_file(image_path: str, image_filename: str, output_folder: str, 
                          include_image: bool, include_ocr: bool) -> str:
    """
    创建图片的markdown文件（与文档转MD逻辑一致）
    
    参数:
        image_path: 图片文件的完整路径
        image_filename: 图片文件名（如 'image_1.png'）
        output_folder: 输出文件夹路径
        include_image: 是否在md文件中包含图片链接
        include_ocr: 是否在md文件中包含OCR识别文本
    
    返回:
        str: 创建的md文件名（如 'image_1.md'）
    """
    # 生成md文件名（与图片同名，扩展名改为.md）
    base_name = os.path.splitext(image_filename)[0]
    md_filename = f"{base_name}.md"
    md_path = os.path.join(output_folder, md_filename)
    
    try:
        lines = []
        
        # 包含图片链接
        if include_image:
            lines.append(f"![{image_filename}]({image_filename})")
            lines.append("")  # 空行
        
        # 包含OCR识别文本
        if include_ocr:
            logger.info(f"开始OCR识别: {image_filename}")
            from gongwen_converter.utils.ocr_utils import extract_text_simple
            ocr_text = extract_text_simple(image_path)
            
            if ocr_text:
                lines.append(ocr_text)
                logger.info(f"OCR识别完成: {image_filename}, 识别出 {len(ocr_text)} 个字符")
        
        # 写入文件
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        logger.info(f"创建图片md文件: {md_filename}")
        return md_filename
        
    except Exception as e:
        logger.error(f"创建图片md文件失败: {md_filename}, 错误: {e}", exc_info=True)
        # 失败时返回图片链接作为兜底
        return image_filename
