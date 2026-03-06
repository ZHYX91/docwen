"""table_merger 单元测试。"""

from __future__ import annotations

import os
import shutil
from pathlib import Path

import openpyxl
import pytest

import docwen.table_merger.core as table_core
from docwen.table_merger.core import TableMerger

pytestmark = pytest.mark.unit


def test_save_result_generates_unique_suffix_on_collision(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    output_dir = tmp_path / "out"
    output_dir.mkdir()

    monkeypatch.setattr(table_core, "t", lambda *_args, **_kwargs: "merged_table", raising=True)
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.get_output_directory", lambda *_args, **_kwargs: str(output_dir), raising=True
    )
    monkeypatch.setattr("docwen.utils.path_utils.generate_timestamp", lambda: "20200101_000000", raising=True)

    merger = TableMerger()
    merger.temp_dir = str(tmp_path / "temp")
    Path(merger.temp_dir).mkdir()

    base_file = tmp_path / "base.xlsx"
    base_file.write_bytes(b"x")

    wb1 = openpyxl.Workbook()
    p1 = merger._save_result(wb1, str(base_file))
    wb1.close()

    wb2 = openpyxl.Workbook()
    p2 = merger._save_result(wb2, str(base_file))
    wb2.close()

    assert Path(p1).exists() is True
    assert Path(p2).exists() is True
    assert p1 != p2
    assert Path(p2).stem.endswith("_001") is True


def test_save_result_falls_back_when_target_move_fails(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    bad_output_dir = tmp_path / "bad_out"
    bad_output_dir.mkdir()

    monkeypatch.setattr(table_core, "t", lambda *_args, **_kwargs: "merged_table", raising=True)
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.get_output_directory",
        lambda *_args, **_kwargs: str(bad_output_dir),
        raising=True,
    )
    monkeypatch.setattr("docwen.utils.path_utils.generate_timestamp", lambda: "20200101_000000", raising=True)

    base_dir = tmp_path / "base_dir"
    base_dir.mkdir()
    base_file = base_dir / "base.xlsx"
    base_file.write_bytes(b"x")

    calls: list[str] = []

    def _stub_move_file_with_retry(source: str, destination: str, *_args, **_kwargs):
        calls.append(destination)

        if os.path.normcase(destination).startswith(os.path.normcase(str(bad_output_dir))):
            return None

        Path(destination).parent.mkdir(parents=True, exist_ok=True)
        shutil.move(source, destination)
        return destination

    monkeypatch.setattr("docwen.utils.workspace_manager.move_file_with_retry", _stub_move_file_with_retry, raising=True)

    merger = TableMerger()
    merger.temp_dir = str(tmp_path / "temp")
    Path(merger.temp_dir).mkdir()

    wb = openpyxl.Workbook()
    out = merger._save_result(wb, str(base_file))
    wb.close()

    assert Path(out).exists() is True
    assert os.path.normcase(str(base_dir)) in os.path.normcase(out)
    assert any(os.path.normcase(str(bad_output_dir)) in os.path.normcase(c) for c in calls) is True
