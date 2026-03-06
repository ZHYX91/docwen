"""table_merger 单元测试。"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.cell.cell import MergedCell

import docwen.table_merger.core as table_core
from docwen.table_merger.core import TableMerger

pytestmark = pytest.mark.unit


def _write_xlsx(path: Path, value: str = "x") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = value
    wb.save(path)
    wb.close()


def test_preprocess_table_xlsx_copies_into_temp_dir(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    src = tmp_path / "a.xlsx"
    _write_xlsx(src)

    temp_dir = tmp_path / "tmp"
    temp_dir.mkdir()

    merger = TableMerger()
    merger.temp_dir = str(temp_dir)

    monkeypatch.setattr(table_core, "detect_actual_file_format", lambda *_args, **_kwargs: "xlsx", raising=True)
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.prepare_input_file", lambda p, *_args, **_kwargs: p, raising=True
    )

    out = merger._preprocess_table(str(src), is_base=True)
    out_path = Path(out)
    assert out_path.exists() is True
    assert out_path.name.endswith("_base.xlsx")
    assert out in merger.temp_files


def test_preprocess_table_csv_converts_and_removes_intermediate(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    src = tmp_path / "a.csv"
    src.write_text("a,b\n1,2\n", encoding="utf-8")

    temp_dir = tmp_path / "tmp"
    temp_dir.mkdir()

    def _stub_csv_to_xlsx(_csv_path: str, output_path: str | None = None, **_kwargs) -> str:
        assert output_path is not None
        _write_xlsx(Path(output_path), value="mid")
        return output_path

    merger = TableMerger()
    merger.temp_dir = str(temp_dir)

    monkeypatch.setattr(table_core, "detect_actual_file_format", lambda *_args, **_kwargs: "csv", raising=True)
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.prepare_input_file", lambda p, *_args, **_kwargs: p, raising=True
    )
    monkeypatch.setattr(table_core, "csv_to_xlsx", _stub_csv_to_xlsx, raising=True)

    out = merger._preprocess_table(str(src), is_base=False)
    out_path = Path(out)
    assert out_path.exists() is True
    assert out_path.name.endswith("_collect.xlsx")


def test_preprocess_table_unsupported_format_raises(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    src = tmp_path / "a.bin"
    src.write_bytes(b"x")

    temp_dir = tmp_path / "tmp"
    temp_dir.mkdir()

    merger = TableMerger()
    merger.temp_dir = str(temp_dir)

    monkeypatch.setattr(table_core, "detect_actual_file_format", lambda *_args, **_kwargs: "pdf", raising=True)
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.prepare_input_file", lambda p, *_args, **_kwargs: p, raising=True
    )

    with pytest.raises(ValueError) as excinfo:
        merger._preprocess_table(str(src), is_base=False)

    assert "不支持的表格格式" in str(excinfo.value)


def test_preprocess_table_ods_converts_to_xlsx(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    src = tmp_path / "a.ods"
    src.write_bytes(b"x")

    temp_dir = tmp_path / "tmp"
    temp_dir.mkdir()

    merger = TableMerger()
    merger.temp_dir = str(temp_dir)

    monkeypatch.setattr(table_core, "detect_actual_file_format", lambda *_args, **_kwargs: "ods", raising=True)
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.prepare_input_file", lambda p, *_args, **_kwargs: p, raising=True
    )

    def _stub_ods_to_xlsx(_input_path: str, output_path: str, **_kwargs) -> str:
        _write_xlsx(Path(output_path), value="ods")
        return output_path

    monkeypatch.setattr("docwen.converter.formats.spreadsheet.ods_to_xlsx", _stub_ods_to_xlsx, raising=True)

    out = merger._preprocess_table(str(src), is_base=True)
    out_path = Path(out)
    assert out_path.exists() is True
    assert out_path.name.endswith("_base.xlsx")


def test_preprocess_table_xls_passes_actual_format_to_office_to_xlsx(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    src = tmp_path / "a.xls"
    src.write_bytes(b"x")

    temp_dir = tmp_path / "tmp"
    temp_dir.mkdir()

    merger = TableMerger()
    merger.temp_dir = str(temp_dir)

    monkeypatch.setattr(table_core, "detect_actual_file_format", lambda *_args, **_kwargs: "et", raising=True)
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.prepare_input_file", lambda p, *_args, **_kwargs: p, raising=True
    )

    captured = {}

    def _stub_office_to_xlsx(input_path: str, output_path: str, actual_format: str | None = None, **_kwargs) -> str:
        captured["actual_format"] = actual_format
        _write_xlsx(Path(output_path), value="et")
        return output_path

    monkeypatch.setattr("docwen.converter.formats.spreadsheet.office_to_xlsx", _stub_office_to_xlsx, raising=True)

    out = merger._preprocess_table(str(src), is_base=True)
    out_path = Path(out)
    assert captured["actual_format"] == "et"
    assert out_path.exists() is True
    assert out_path.name.endswith("_base.xlsx")


def test_merge_tables_smoke_flow(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    base = tmp_path / "base.xlsx"
    collect = tmp_path / "collect.xlsx"
    _write_xlsx(base, value="base")
    _write_xlsx(collect, value="collect")

    merger = TableMerger()

    monkeypatch.setattr(table_core, "t", lambda *_args, **_kwargs: "merged", raising=True)
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.get_output_directory", lambda *_args, **_kwargs: str(tmp_path), raising=True
    )

    monkeypatch.setattr(TableMerger, "_preprocess_table", lambda self, file_path, **_kwargs: file_path, raising=True)
    monkeypatch.setattr(TableMerger, "_unmerge_all_cells", lambda *_args, **_kwargs: None, raising=True)
    monkeypatch.setattr(TableMerger, "_find_best_offset", lambda *_args, **_kwargs: (0, 0), raising=True)
    monkeypatch.setattr(TableMerger, "_merge_by_row", lambda *_args, **_kwargs: None, raising=True)

    ok, msg, out_path = merger.merge_tables(str(base), [str(collect)], mode=TableMerger.MODE_BY_ROW)

    assert ok is True
    assert "成功汇总" in msg
    assert out_path is not None
    assert Path(out_path).exists() is True
    assert Path(out_path).suffix == ".xlsx"
    assert merger.temp_dir is not None
    assert Path(merger.temp_dir).exists() is False


def test_merge_cell_values_rules() -> None:
    merger = TableMerger()

    assert merger._merge_cell_values(None, None) is None
    assert merger._merge_cell_values("", "x") == "x"
    assert merger._merge_cell_values("x", "") == "x"
    assert merger._merge_cell_values("1", 2) == 3
    assert merger._merge_cell_values("A", "A") == "A"
    assert merger._merge_cell_values("A", "B") == "A,B"


def test_try_convert_to_number() -> None:
    merger = TableMerger()

    assert merger._try_convert_to_number(1) == 1
    assert merger._try_convert_to_number(1.5) == 1.5
    assert merger._try_convert_to_number(" 2 ") == 2
    assert merger._try_convert_to_number("2.0") == 2
    assert merger._try_convert_to_number("2.5") == 2.5
    assert merger._try_convert_to_number("x") is None
    assert merger._try_convert_to_number(None) is None


def test_calculate_overlap_and_find_best_offset() -> None:
    base_wb = openpyxl.Workbook()
    base_ws = base_wb.active
    base_ws["B2"] = "X"

    collect_wb = openpyxl.Workbook()
    collect_ws = collect_wb.active
    collect_ws["A1"] = "X"

    merger = TableMerger()

    assert merger._calculate_overlap(base_ws, collect_ws, row_offset=1, col_offset=1) == 1
    assert merger._calculate_overlap(base_ws, collect_ws, row_offset=0, col_offset=0) == 0
    assert merger._find_best_offset(base_ws, collect_ws) == (1, 1)


def test_unmerge_all_cells_fills_value_to_all_cells() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "V"
    ws.merge_cells("A1:B2")

    merger = TableMerger()
    merger._unmerge_all_cells(ws)

    assert len(list(ws.merged_cells.ranges)) == 0
    assert ws["A1"].value == "V"
    assert ws["B1"].value == "V"
    assert ws["A2"].value == "V"
    assert ws["B2"].value == "V"
    assert not isinstance(ws["B2"], MergedCell)
    wb.close()


def test_merge_by_cell_does_not_write_into_merged_cell() -> None:
    base_wb = openpyxl.Workbook()
    base_ws = base_wb.active
    base_ws["A1"] = 1
    base_ws.merge_cells("A1:A2")

    collect_wb = openpyxl.Workbook()
    collect_ws = collect_wb.active
    collect_ws["A2"] = 2

    merger = TableMerger()
    merger._merge_by_cell(base_ws, collect_ws, row_offset=0, col_offset=0)

    assert base_ws["A1"].value == 1
    assert isinstance(base_ws["A2"], MergedCell)
    base_wb.close()
    collect_wb.close()

    base_wb.close()
    collect_wb.close()


def test_find_best_offset_out_of_range_uses_value_matching() -> None:
    base_wb = openpyxl.Workbook()
    base_ws = base_wb.active
    base_ws.cell(50, 50).value = "X"

    collect_wb = openpyxl.Workbook()
    collect_ws = collect_wb.active
    collect_ws["A1"] = "X"

    merger = TableMerger()
    assert merger._find_best_offset(base_ws, collect_ws) == (49, 49)

    base_wb.close()
    collect_wb.close()


def test_find_best_offset_normalizes_numbers() -> None:
    base_wb = openpyxl.Workbook()
    base_ws = base_wb.active
    base_ws["B2"] = 2

    collect_wb = openpyxl.Workbook()
    collect_ws = collect_wb.active
    collect_ws["A1"] = "2.0"

    merger = TableMerger()
    assert merger._find_best_offset(base_ws, collect_ws) == (1, 1)

    base_wb.close()
    collect_wb.close()


def test_check_row_coverage_normalizes_numbers() -> None:
    base_wb = openpyxl.Workbook()
    base_ws = base_wb.active
    base_ws["A1"] = "K"
    base_ws["B1"] = 2

    collect_wb = openpyxl.Workbook()
    collect_ws = collect_wb.active
    collect_ws["A1"] = "K"
    collect_ws["B1"] = "2.0"
    collect_ws["C1"] = "X"

    merger = TableMerger()
    assert merger._check_row_coverage(base_ws, 1, collect_ws, 1, col_offset=0) == "collect_covers_base"

    base_wb.close()
    collect_wb.close()


def test_check_column_coverage_normalizes_numbers() -> None:
    base_wb = openpyxl.Workbook()
    base_ws = base_wb.active
    base_ws["A1"] = "K"
    base_ws["A2"] = 2

    collect_wb = openpyxl.Workbook()
    collect_ws = collect_wb.active
    collect_ws["A1"] = "K"
    collect_ws["A2"] = "2.0"
    collect_ws["A3"] = "X"

    merger = TableMerger()
    assert merger._check_column_coverage(base_ws, 1, collect_ws, 1, row_offset=0) == "collect_covers_base"

    base_wb.close()
    collect_wb.close()


def test_find_best_offset_respects_offset_range() -> None:
    base_wb = openpyxl.Workbook()
    base_ws = base_wb.active
    for r in (15, 16, 17):
        base_ws.cell(r, 1).value = "X"
    for r in (21, 22, 23, 24):
        base_ws.cell(r, 1).value = "X"

    collect_wb = openpyxl.Workbook()
    collect_ws = collect_wb.active
    for r in (1, 2, 3, 4):
        collect_ws.cell(r, 1).value = "X"

    merger = TableMerger()
    assert merger._find_best_offset(base_ws, collect_ws) == (14, 0)

    merger.offset_range = 25
    assert merger._find_best_offset(base_ws, collect_ws) == (20, 0)

    base_wb.close()
    collect_wb.close()
