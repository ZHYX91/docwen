"""xlsx2md 转换与图片提取的单元测试。"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.drawing.image import Image as XlsxImage
from PIL import Image

from docwen.converter.xlsx2md.core import convert_spreadsheet_to_md
from docwen.converter.xlsx2md.image_processor import extract_images_from_xlsx

pytestmark = pytest.mark.unit


def test_convert_spreadsheet_to_md_uses_file_path_actual_format_not_original_extension(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "in.csv"
    csv_path.write_text("A,B\n1,2\n", encoding="utf-8")

    monkeypatch.setattr(
        "docwen.utils.file_type_utils.detect_actual_file_format",
        lambda *_args, **_kwargs: "csv",
        raising=True,
    )

    def _should_not_load_workbook(*_args, **_kwargs):
        raise AssertionError("openpyxl.load_workbook should not be called for CSV input")

    monkeypatch.setattr("openpyxl.load_workbook", _should_not_load_workbook, raising=True)

    md = convert_spreadsheet_to_md(
        str(csv_path),
        extract_image=False,
        extract_ocr=False,
        original_file_path=str(tmp_path / "renamed.xls"),
    )

    assert "# renamed" in md
    assert "A" in md
    assert "B" in md


def test_convert_spreadsheet_to_md_cancel_event_skips_ocr(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    png_path = tmp_path / "img.png"
    Image.new("RGB", (1, 1), (0, 255, 0)).save(png_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.add_image(XlsxImage(str(png_path)), "A1")

    xlsx_path = tmp_path / "book.xlsx"
    wb.save(xlsx_path)
    wb.close()

    out_dir = tmp_path / "out"

    class _ToggleCancel:
        def __init__(self) -> None:
            self._calls = 0

        def is_set(self) -> bool:
            self._calls += 1
            return self._calls >= 2

    cancel_event = _ToggleCancel()

    called = {"ocr": 0}

    def _raise_ocr(*_args, **_kwargs):
        called["ocr"] += 1
        raise AssertionError("OCR should not run when cancel_event is set")

    monkeypatch.setattr("docwen.utils.ocr_utils.extract_text_simple", _raise_ocr, raising=True)

    md = convert_spreadsheet_to_md(
        str(xlsx_path),
        extract_image=False,
        extract_ocr=True,
        output_folder=str(out_dir),
        original_file_path=str(xlsx_path),
        cancel_event=cancel_event,
    )

    assert "{{IMAGE:" not in md
    assert called["ocr"] == 0


def test_worksheet_to_dataframe_limits_iter_rows_to_effective_used_range(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from openpyxl.worksheet.worksheet import Worksheet

    from docwen.converter.xlsx2md import core as xlsx2md_core

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "X"
    ws["Z1000"] = ""

    original_iter_rows = Worksheet.iter_rows
    called: dict[str, object] = {}

    def _spy_iter_rows(self: Worksheet, *args, **kwargs):
        if self is ws:
            called["kwargs"] = dict(kwargs)
        return original_iter_rows(self, *args, **kwargs)

    monkeypatch.setattr(Worksheet, "iter_rows", _spy_iter_rows, raising=True)

    df = xlsx2md_core._worksheet_to_dataframe(ws)

    wb.close()

    assert df.shape == (1, 1)
    assert called["kwargs"] == {"min_row": 1, "max_row": 1, "min_col": 1, "max_col": 1}


def test_convert_spreadsheet_to_md_reads_gbk_csv(tmp_path: Path) -> None:
    csv_path = tmp_path / "in.csv"
    csv_path.write_bytes("列1,列2\n1,2\n".encode("gbk"))

    md = convert_spreadsheet_to_md(
        str(csv_path),
        extract_image=False,
        extract_ocr=False,
        original_file_path=str(csv_path),
    )

    assert "列1" in md
    assert "列2" in md
    assert "1" in md
    assert "2" in md


def test_convert_spreadsheet_to_md_reads_semicolon_csv(tmp_path: Path) -> None:
    csv_path = tmp_path / "in.csv"
    csv_path.write_text("A;B\n1;2\n", encoding="utf-8")

    md = convert_spreadsheet_to_md(
        str(csv_path),
        extract_image=False,
        extract_ocr=False,
        original_file_path=str(csv_path),
    )

    assert "A" in md
    assert "B" in md
    assert "1" in md
    assert "2" in md


@pytest.mark.slow
def test_extract_images_from_xlsx_extracts_one(tmp_path: Path) -> None:
    png_path = tmp_path / "img.png"
    Image.new("RGB", (1, 1), (0, 255, 0)).save(png_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.add_image(XlsxImage(str(png_path)), "A1")

    xlsx_path = tmp_path / "book.xlsx"
    wb.save(xlsx_path)
    wb.close()

    loaded = openpyxl.load_workbook(xlsx_path, data_only=True)
    out_dir = tmp_path / "images"
    images = extract_images_from_xlsx(loaded, str(out_dir), str(xlsx_path))
    loaded.close()

    assert len(images) == 1
    assert Path(images[0]["image_path"]).exists() is True
