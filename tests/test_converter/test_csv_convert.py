"""converter 单元测试。"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from docwen.converter.formats.spreadsheet.csv_convert import csv_to_xlsx, xlsx_to_csv


pytestmark = pytest.mark.unit

@pytest.mark.unit
def test_csv_to_xlsx_creates_workbook(tmp_path: Path) -> None:
    csv_path = tmp_path / "a.csv"
    csv_path.write_text("1,2,3\n4,5,6\n", encoding="utf-8")

    out = csv_to_xlsx(str(csv_path))
    out_path = Path(out)
    assert out_path.exists() is True

    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == ["Sheet1"]
    ws = wb.active
    assert ws.cell(row=1, column=1).value == 1
    assert ws.cell(row=2, column=3).value == 6


@pytest.mark.unit
def test_xlsx_to_csv_creates_one_file_per_sheet(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "S1"
    ws1.append([1, 2])
    ws1.append([3, 4])
    ws2 = wb.create_sheet("S 2")
    ws2.append(["a", "b"])
    wb.save(xlsx_path)

    outs = xlsx_to_csv(
        str(xlsx_path),
        output_dir=str(tmp_path),
        original_basename="book",
        unified_timestamp_desc="20250101_000000_fromXlsx",
    )

    assert len(outs) == 2
    out_paths = [Path(p) for p in outs]
    assert all(p.exists() for p in out_paths) is True

    names = sorted(p.name for p in out_paths)
    assert names[0].startswith("book_S1_20250101_000000_fromXlsx")
    assert names[1].startswith("book_S_2_20250101_000000_fromXlsx")

    text = out_paths[0].read_text(encoding="utf-8-sig")
    assert "1,2" in text
