"""services 单元测试。"""

from __future__ import annotations

import json
import threading
from pathlib import Path

import openpyxl
import pytest

import docwen.cli.executor as executor

pytestmark = pytest.mark.unit


def _write_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "A"
    ws["B1"] = "B"
    ws["A2"] = 1
    ws["B2"] = 2
    wb.save(path)
    wb.close()


def test_cli_spreadsheet_convert_to_md_smoke(
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
    tmp_path: Path,
) -> None:
    xlsx = tmp_path / "a.xlsx"
    _write_xlsx(xlsx)

    out_dir = tmp_path / "out"
    out_dir.mkdir()
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.get_output_directory",
        lambda *_args, **_kwargs: str(out_dir),
        raising=True,
    )

    exit_code = executor.execute_action("convert", str(xlsx), options={"target_format": "md"}, json_mode=True)
    payload = json.loads(capsys.readouterr().out)

    assert exit_code == 0
    assert payload["success"] is True
    assert Path(payload["data"]["output_file"]).exists() is True
    assert Path(payload["data"]["output_file"]).suffix == ".md"


def test_spreadsheet_to_md_strategy_passes_cancel_event(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from docwen.services.strategies.spreadsheet.to_markdown import SpreadsheetToMarkdownStrategy

    xlsx = tmp_path / "a.xlsx"
    _write_xlsx(xlsx)

    out_dir = tmp_path / "out"
    out_dir.mkdir()
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.get_output_directory",
        lambda *_args, **_kwargs: str(out_dir),
        raising=True,
    )

    def _stub_prepare_input_file(input_path: str, temp_dir: str, actual_format: str) -> str:
        dst = Path(temp_dir) / f"input.{actual_format}"
        dst.write_bytes(Path(input_path).read_bytes())
        return str(dst)

    monkeypatch.setattr(
        "docwen.utils.workspace_manager.prepare_input_file",
        _stub_prepare_input_file,
        raising=True,
    )

    captured = {}

    def _stub_convert_spreadsheet_to_md(_file_path: str, **kwargs) -> str:
        captured["cancel_event"] = kwargs.get("cancel_event")
        return "# ok"

    monkeypatch.setattr(
        "docwen.services.strategies.spreadsheet.to_markdown.convert_spreadsheet_to_md",
        _stub_convert_spreadsheet_to_md,
        raising=True,
    )

    cancel_event = threading.Event()
    strategy = SpreadsheetToMarkdownStrategy()
    result = strategy.execute(str(xlsx), options={"actual_format": "xlsx", "cancel_event": cancel_event})

    assert result.success is True
    assert captured["cancel_event"] is cancel_event


def test_spreadsheet_to_txt_is_not_supported() -> None:
    from docwen.errors import StrategyNotFoundError
    from docwen.services.strategies import get_strategy

    with pytest.raises(StrategyNotFoundError):
        get_strategy(source_format="xlsx", target_format="txt")


def test_cli_spreadsheet_convert_csv_smoke(
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
    tmp_path: Path,
) -> None:
    xlsx = tmp_path / "a.xlsx"
    _write_xlsx(xlsx)

    out_dir = tmp_path / "out"
    out_dir.mkdir()
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.get_output_directory",
        lambda *_args, **_kwargs: str(out_dir),
        raising=True,
    )

    exit_code = executor.execute_action(
        "convert",
        str(xlsx),
        options={"target_format": "csv"},
        json_mode=True,
    )
    payload = json.loads(capsys.readouterr().out)

    assert exit_code == 0
    assert payload["success"] is True
    assert Path(payload["data"]["output_file"]).exists() is True
    assert Path(payload["data"]["output_file"]).suffix == ".csv"


def test_cli_spreadsheet_merge_tables_smoke(
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
    tmp_path: Path,
) -> None:
    base = tmp_path / "base.xlsx"
    collect = tmp_path / "collect.xlsx"
    _write_xlsx(base)
    _write_xlsx(collect)

    monkeypatch.setattr(
        "docwen.utils.workspace_manager.get_output_directory",
        lambda *_args, **_kwargs: str(tmp_path),
        raising=True,
    )

    exit_code = executor.execute_action(
        "merge_tables",
        str(base),
        options={"mode": 1, "file_list": [str(base), str(collect)]},
        json_mode=True,
    )
    payload = json.loads(capsys.readouterr().out)

    assert exit_code == 0
    assert payload["success"] is True
    assert Path(payload["data"]["output_file"]).exists() is True
    assert Path(payload["data"]["output_file"]).suffix == ".xlsx"


def test_spreadsheet_to_pdf_strategy_smoke_without_office(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from docwen.services.strategies.spreadsheet.to_pdf import SpreadsheetToPdfStrategy

    xlsx = tmp_path / "a.xlsx"
    _write_xlsx(xlsx)

    out_dir = tmp_path / "out"
    out_dir.mkdir()
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.get_output_directory",
        lambda *_args, **_kwargs: str(out_dir),
        raising=True,
    )

    def _stub_prepare_input_file(input_path: str, temp_dir: str, actual_format: str) -> str:
        dst = Path(temp_dir) / f"input.{actual_format}"
        dst.write_bytes(Path(input_path).read_bytes())
        return str(dst)

    monkeypatch.setattr(
        "docwen.utils.workspace_manager.prepare_input_file",
        _stub_prepare_input_file,
        raising=True,
    )

    def _stub_xlsx_to_pdf(_input_path: str, output_path: str, cancel_event=None):
        Path(output_path).write_bytes(b"%PDF-1.4\n%stub\n")
        return output_path

    monkeypatch.setattr(
        "docwen.converter.formats.pdf_export.xlsx_to_pdf",
        _stub_xlsx_to_pdf,
        raising=True,
    )

    strategy = SpreadsheetToPdfStrategy()
    result = strategy.execute(str(xlsx), options={"actual_format": "xlsx"})

    assert result.success is True
    assert result.output_path is not None
    assert Path(result.output_path).exists() is True
    assert Path(result.output_path).suffix == ".pdf"


def test_spreadsheet_to_pdf_detects_actual_format_when_missing(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from docwen.services.strategies.spreadsheet.to_pdf import SpreadsheetToPdfStrategy

    xlsx = tmp_path / "a.xlsx"
    _write_xlsx(xlsx)

    out_dir = tmp_path / "out"
    out_dir.mkdir()
    monkeypatch.setattr(
        "docwen.utils.workspace_manager.get_output_directory",
        lambda *_args, **_kwargs: str(out_dir),
        raising=True,
    )
    monkeypatch.setattr(
        "docwen.utils.file_type_utils.detect_actual_file_format",
        lambda *_args, **_kwargs: "xls",
        raising=True,
    )

    captured = {}

    def _stub_prepare_input_file(input_path: str, temp_dir: str, actual_format: str) -> str:
        captured["actual_format"] = actual_format
        dst = Path(temp_dir) / f"input.{actual_format}"
        dst.write_bytes(Path(input_path).read_bytes())
        return str(dst)

    monkeypatch.setattr(
        "docwen.utils.workspace_manager.prepare_input_file",
        _stub_prepare_input_file,
        raising=True,
    )

    def _stub_xlsx_to_pdf(_input_path: str, output_path: str, cancel_event=None):
        Path(output_path).write_bytes(b"%PDF-1.4\n%stub\n")
        return output_path

    monkeypatch.setattr(
        "docwen.converter.formats.pdf_export.xlsx_to_pdf",
        _stub_xlsx_to_pdf,
        raising=True,
    )

    strategy = SpreadsheetToPdfStrategy()
    result = strategy.execute(str(xlsx), options={})

    assert captured["actual_format"] == "xls"
    assert result.success is True
