"""Capture FA-10 Merge Tables final-artifact parity evidence.

The orchestrator prepares immutable inputs under an external evidence directory,
then invokes each project's public route where it is operable.  The old PySide6
public CLI defect is captured explicitly and its production TableMerger is used
for the artifact projection.  Reference repositories remain read-only.  The N1
XLSX is patched at the OOXML layer so cached formula values from the externally
calculated source are preserved.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import subprocess
import sys
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECTS = ("docwen-ref-tk", "docwen-ref-pyside6", "docwen-current")
EXPECTED_MANDATORY_PROJECTIONS: dict[str, dict[str, Any]] = {
    "B1": {
        "sheetnames": ["Formula", "BaseExtra"],
        "active_sheet": "Formula",
        "dimensions": "A1:D4",
        "used_values": [
            ["Item", "Value", "Base note,Collect note", "Base note,Collect note"],
            ["Alpha", 12, None, None],
            ["Beta", 8, None, None],
            ["Total", 20, None, None],
        ],
        "contains_formula_text": False,
        "merged_ranges": [],
        "sheet_protection": True,
        "base_style": {
            "a1_style_id": 1,
            "a1_fill_type": "solid",
            "a1_fill_rgb": "00FFCC00",
            "a1_bold": True,
        },
    },
    "N1": {
        "sheetnames": ["Data", "Lookup"],
        "active_sheet": "Data",
        "dimensions": "A1:E7",
        "used_values": [
            ["VIS-105 LibreOffice SmartSheet Matrix"] * 5,
            ["Item", "Region", "Quantity", "Unit Price", "Amount"],
            ["Alpha", "North", 4, 25, 50],
            ["Beta", "South", 10, 14.5, 72.5],
            ["Gamma", "East", 6, 9.5, 28.5],
            ["Delta", "West", 8, 10, 40],
            ["Total", None, 28, None, 191],
        ],
        "contains_formula_text": False,
        "merged_ranges": [],
        "sheet_protection": True,
        "base_style": {
            "a1_style_id": 4,
            "a1_fill_type": "solid",
            "a1_fill_rgb": "FF1F4E78",
            "a1_bold": True,
        },
    },
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _print_json(value: Any) -> None:
    payload = (json.dumps(value, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    sys.stdout.buffer.write(payload)


def _write_deterministic_xlsx(path: Path, members: dict[str, bytes]) -> None:
    with zipfile.ZipFile(path, "w") as target:
        for name in sorted(members):
            info = zipfile.ZipInfo(name, date_time=(2026, 7, 22, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            target.writestr(info, members[name])


def _write_cached_formula_workbook(path: Path, *, base: bool) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.properties.created = datetime(2026, 7, 22, 0, 0, 0)
    wb.properties.modified = datetime(2026, 7, 22, 0, 0, 0)
    ws = wb.active
    assert ws is not None
    ws.title = "Formula"
    rows = (
        [["Item", "Value"], ["Alpha", 10], ["Beta", 5], ["Total", "=SUM(B2:B3)"]]
        if base
        else [["Item", "Value"], ["Alpha", 2], ["Beta", 3], ["Total", "=SUM(B2:B3)"]]
    )
    for row in rows:
        ws.append(row)
    ws.merge_cells("C1:D1")
    ws["C1"] = "Base note" if base else "Collect note"
    ws["A1"].fill = PatternFill("solid", fgColor="FFCC00" if base else "99CCFF")
    ws["A1"].font = Font(bold=base, italic=not base)
    ws.protection.sheet = base
    if base:
        extra = wb.create_sheet("BaseExtra")
        extra["A1"] = "base-extra-kept"
    wb.save(path)
    wb.close()

    cached_value = 15 if base else 5
    with zipfile.ZipFile(path, "r") as source:
        members = {name: source.read(name) for name in source.namelist()}
    core_member = "docProps/core.xml"
    core_xml = members[core_member].decode("utf-8")
    core_xml = re.sub(
        r"(<dcterms:modified[^>]*>)[^<]+(</dcterms:modified>)",
        r"\g<1>2026-07-22T00:00:00Z\g<2>",
        core_xml,
    )
    members[core_member] = core_xml.encode("utf-8")
    xml = members["xl/worksheets/sheet1.xml"].decode("utf-8")
    xml, count = re.subn(
        r'(<c r="B4"[^>]*><f>SUM\(B2:B3\)</f>)<v></v>(</c>)',
        rf"\g<1><v>{cached_value}</v>\g<2>",
        xml,
    )
    if count != 1:
        raise RuntimeError(f"could not inject B1 cached value into {path}")
    members["xl/worksheets/sheet1.xml"] = xml.encode("utf-8")
    _write_deterministic_xlsx(path, members)


def _add_sheet_protection_without_recalculation(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as source:
        members = {name: source.read(name) for name in source.namelist()}
    member = "xl/worksheets/sheet1.xml"
    xml = members[member].decode("utf-8")
    if "<sheetProtection" not in xml:
        xml, count = re.subn(
            r"</(?P<prefix>[A-Za-z0-9_]+:)?sheetData>",
            lambda match: (
                f"</{match.group('prefix') or ''}sheetData>"
                f"<{match.group('prefix') or ''}sheetProtection "
                'sheet="1" objects="1" scenarios="1"/>'
            ),
            xml,
            count=1,
        )
        if count != 1:
            raise RuntimeError(f"could not add sheet protection to {path}")
        members[member] = xml.encode("utf-8")
        _write_deterministic_xlsx(path, members)


def _formula_cache_projection(path: Path) -> dict[str, Any]:
    import openpyxl

    values = openpyxl.load_workbook(path, data_only=True)
    formulas = openpyxl.load_workbook(path, data_only=False)
    try:
        cached: dict[str, Any] = {}
        formula_text: dict[str, str] = {}
        for sheet_name in formulas.sheetnames:
            formula_sheet = formulas[sheet_name]
            value_sheet = values[sheet_name]
            for row in formula_sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        key = f"{sheet_name}!{cell.coordinate}"
                        formula_text[key] = cell.value
                        cached[key] = value_sheet[cell.coordinate].value
        return {
            "formula_count": len(formula_text),
            "formula_text": formula_text,
            "cached_values": cached,
        }
    finally:
        values.close()
        formulas.close()


def _prepare_inputs(args: argparse.Namespace, output_dir: Path) -> dict[str, Any]:
    input_dir = output_dir / "inputs"
    b1_dir = input_dir / "b1"
    n1_dir = input_dir / "n1"
    b1_dir.mkdir(parents=True, exist_ok=True)
    n1_dir.mkdir(parents=True, exist_ok=True)

    b1_base = b1_dir / "cached-base.xlsx"
    b1_collect = b1_dir / "cached-collect.xlsx"
    _write_cached_formula_workbook(b1_base, base=True)
    _write_cached_formula_workbook(b1_collect, base=False)

    rich_xlsx = Path(args.rich_xlsx).resolve()
    legacy_xls = Path(args.legacy_xls).resolve()
    n1_base = n1_dir / "protected-base.xlsx"
    n1_collect = n1_dir / "legacy-collect.xls"
    shutil.copyfile(rich_xlsx, n1_base)
    shutil.copyfile(legacy_xls, n1_collect)
    _add_sheet_protection_without_recalculation(n1_base)

    files = (b1_base, b1_collect, n1_base, n1_collect)
    manifest = {
        "stage": "FA-10",
        "probe_id": "FA-10-B1-N1-2026-07-22",
        "contract": {
            "sets": ["B1", "N1"],
            "projects": list(PROJECTS),
            "expected_final_workbooks": 6,
            "mandatory_fields": [
                "used_values",
                "cached_formula_values",
                "base_style",
                "merged_ranges",
                "sheet_protection",
                "active_sheet",
                "old_format_preconversion",
                "final_placement",
            ],
            "pass_threshold": "100% mandatory normalized fields across all three projects",
            "collision_evidence": (
                "existing current end-to-end legacy-suffix guard; no duplicate timestamped "
                "reference outputs are generated"
            ),
            "external_condition": (
                "VIS105 source already contains six cached values; Excel native-dialog attempt "
                "was not accepted and is not counted as external acceptance"
            ),
        },
        "sources": {
            "rich_xlsx": {"path": str(rich_xlsx), "sha256": _sha256(rich_xlsx)},
            "legacy_xls": {"path": str(legacy_xls), "sha256": _sha256(legacy_xls)},
        },
        "prepared_inputs": {
            path.relative_to(output_dir).as_posix(): {
                "sha256": _sha256(path),
                "size": path.stat().st_size,
                **(_formula_cache_projection(path) if path.suffix.lower() == ".xlsx" else {}),
            }
            for path in files
        },
    }
    _write_json(output_dir / "input-manifest.json", manifest)
    return manifest


def _task_processes() -> list[dict[str, Any]]:
    completed = subprocess.run(
        ["tasklist", "/fo", "csv", "/nh"],
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    tracked = {"excel.exe", "wps.exe", "et.exe", "soffice.exe", "soffice.bin"}
    rows: list[dict[str, Any]] = []
    for row in csv.reader(completed.stdout.splitlines()):
        if len(row) >= 2 and row[0].lower() in tracked:
            rows.append({"name": row[0], "pid": row[1]})
    return rows


def _projection(path: Path) -> dict[str, Any]:
    import openpyxl

    values = openpyxl.load_workbook(path, data_only=True)
    formulas = openpyxl.load_workbook(path, data_only=False)
    try:
        value_sheet = values.active
        formula_sheet = formulas.active
        assert value_sheet is not None and formula_sheet is not None
        max_row = value_sheet.max_row
        max_column = value_sheet.max_column
        used_values = [
            [value_sheet.cell(row, column).value for column in range(1, max_column + 1)]
            for row in range(1, max_row + 1)
        ]
        formula_values = [
            [formula_sheet.cell(row, column).value for column in range(1, max_column + 1)]
            for row in range(1, max_row + 1)
        ]
        a1 = value_sheet["A1"]
        return {
            "sheetnames": values.sheetnames,
            "active_sheet": value_sheet.title,
            "dimensions": value_sheet.calculate_dimension(),
            "used_values": used_values,
            "formula_projection": formula_values,
            "contains_formula_text": any(
                isinstance(value, str) and value.startswith("=") for row in formula_values for value in row
            ),
            "merged_ranges": sorted(str(item) for item in value_sheet.merged_cells.ranges),
            "sheet_protection": bool(value_sheet.protection.sheet),
            "base_style": {
                "a1_style_id": a1.style_id,
                "a1_fill_type": a1.fill.fill_type,
                "a1_fill_rgb": a1.fill.fgColor.rgb,
                "a1_bold": bool(a1.font.bold),
            },
        }
    finally:
        values.close()
        formulas.close()


def _mandatory_projection(projection: dict[str, Any]) -> dict[str, Any]:
    return {key: projection[key] for key in EXPECTED_MANDATORY_PROJECTIONS["B1"]}


def _run_case(
    *,
    project: str,
    executable: Path,
    project_root: Path,
    case: str,
    source_base: Path,
    source_collect: Path,
    output_dir: Path,
) -> dict[str, Any]:
    case_dir = output_dir / project / case.lower()
    if case_dir.exists():
        if not case_dir.resolve().is_relative_to(output_dir.resolve()):
            raise RuntimeError(f"refusing to replace evidence outside output directory: {case_dir}")
        shutil.rmtree(case_dir)
    case_dir.mkdir(parents=True, exist_ok=True)
    base = case_dir / source_base.name
    collect = case_dir / source_collect.name
    shutil.copyfile(source_base, base)
    shutil.copyfile(source_collect, collect)
    before_files = {path.resolve() for path in case_dir.rglob("*.xlsx")}
    before_processes = _task_processes()

    worker_result = case_dir / "worker-result.json"
    if project == "docwen-ref-pyside6":
        command = [
            str(project_root / ".venv" / "Scripts" / "python.exe"),
            str(Path(__file__).resolve()),
            "--pyside-worker",
            "--worker-base",
            str(base),
            "--worker-collect",
            str(collect),
            "--worker-result",
            str(worker_result),
        ]
    elif project == "docwen-current":
        final_dir = case_dir / "final"
        command = [
            str(executable),
            "run",
            str(base),
            str(collect),
            "--to",
            "xlsx",
            "--action",
            "merge_tables",
            "--mode",
            "cell",
            "--output",
            str(final_dir),
            "--json",
            "--quiet",
        ]
    else:
        command = [
            str(executable),
            "merge-tables",
            str(base),
            str(collect),
            "--mode",
            "cell",
            "--json",
            "--quiet",
        ]
    completed = subprocess.run(
        command,
        cwd=project_root,
        capture_output=True,
        timeout=240,
    )
    stdout = completed.stdout.decode("utf-8", errors="replace")
    stderr = completed.stderr.decode("utf-8", errors="replace")
    immediate_processes = _task_processes()
    after_processes = immediate_processes
    deadline = time.monotonic() + 10.0
    while any(item not in before_processes for item in after_processes) and time.monotonic() < deadline:
        time.sleep(0.5)
        after_processes = _task_processes()
    if completed.returncode != 0:
        raise RuntimeError(f"{project} {case} failed ({completed.returncode})\nSTDOUT:\n{stdout}\nSTDERR:\n{stderr}")
    if project == "docwen-ref-pyside6":
        worker_payload = json.loads(worker_result.read_text(encoding="utf-8"))
        candidates = [Path(worker_payload["output_path"]).resolve()]
    else:
        candidates = [
            path.resolve()
            for path in case_dir.rglob("*.xlsx")
            if path.resolve() not in before_files and path.resolve() != base.resolve()
        ]
    if len(candidates) != 1:
        raise RuntimeError(f"{project} {case} expected one final XLSX, found {candidates}")
    artifact = candidates[0]
    return {
        "command": command,
        "returncode": completed.returncode,
        "stdout": stdout,
        "stderr": stderr,
        "artifact": str(artifact),
        "artifact_sha256": _sha256(artifact),
        "final_placement": {
            "inside_case_dir": artifact.is_relative_to(case_dir.resolve()),
            "suffix": artifact.suffix.lower(),
            "relative_path": artifact.relative_to(case_dir.resolve()).as_posix(),
        },
        "processes": {
            "before": before_processes,
            "immediate": immediate_processes,
            "after": after_processes,
            "residue_added": [item for item in after_processes if item not in before_processes],
        },
        "projection": _projection(artifact),
    }


def _pyside_worker(args: argparse.Namespace) -> int:
    from docwen.converter import TableMerger

    merger = TableMerger()
    success, message, output_path = merger.merge_tables(
        base_file=str(Path(args.worker_base).resolve()),
        collect_files=[str(Path(args.worker_collect).resolve())],
        mode=3,
    )
    payload = {"success": success, "message": message, "output_path": output_path}
    _write_json(Path(args.worker_result).resolve(), payload)
    return 0 if success else 1


def _capture_pyside_public_cli_defect(
    *, project_root: Path, source_base: Path, source_collect: Path, output_dir: Path
) -> dict[str, Any]:
    defect_dir = output_dir / "docwen-ref-pyside6" / "public-cli-defect"
    if defect_dir.exists():
        if not defect_dir.resolve().is_relative_to(output_dir.resolve()):
            raise RuntimeError(f"refusing to replace evidence outside output directory: {defect_dir}")
        shutil.rmtree(defect_dir)
    defect_dir.mkdir(parents=True)
    base = defect_dir / source_base.name
    collect = defect_dir / source_collect.name
    shutil.copyfile(source_base, base)
    shutil.copyfile(source_collect, collect)
    command = [
        str(project_root / ".venv" / "Scripts" / "docwen.exe"),
        "merge-tables",
        str(base),
        str(collect),
        "--mode",
        "cell",
        "--json",
        "--quiet",
    ]
    completed = subprocess.run(command, cwd=project_root, capture_output=True, timeout=120)
    stdout = completed.stdout.decode("utf-8", errors="replace")
    stderr = completed.stderr.decode("utf-8", errors="replace")
    return {
        "classification": "old_reference_public_cli_file_list_wiring_defect",
        "expected_failure_observed": completed.returncode != 0 and '"success": false' in stdout,
        "command": command,
        "returncode": completed.returncode,
        "stdout": stdout,
        "stderr": stderr,
        "source_evidence": [
            "src/docwen/services/execution.py builds ConversionRequest.file_list",
            "src/docwen/services/use_cases.py does not copy request.file_list into typed.merge.file_list",
            "src/docwen/services/strategies/operations/merge_tables.py reads typed.merge.file_list",
        ],
        "artifact_probe_fallback": "production docwen.converter.TableMerger in isolated reference venv",
    }


def _orchestrate(args: argparse.Namespace) -> int:
    if not args.rich_xlsx or not args.legacy_xls or not args.output_dir:
        raise SystemExit("--rich-xlsx, --legacy-xls, and --output-dir are required")
    current_root = Path(args.current_root).resolve()
    tk_root = Path(args.tk_root).resolve()
    pyside_root = Path(args.pyside_root).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    manifest = _prepare_inputs(args, output_dir)
    if args.prepare_only:
        _print_json(manifest)
        return 0

    inputs = output_dir / "inputs"
    specs = (
        ("docwen-ref-tk", tk_root, tk_root / ".venv" / "Scripts" / "docwen.exe"),
        ("docwen-ref-pyside6", pyside_root, pyside_root / ".venv" / "Scripts" / "docwen.exe"),
        ("docwen-current", current_root, current_root / ".venv" / "Scripts" / "docwen.exe"),
    )
    pyside_cli_defect = _capture_pyside_public_cli_defect(
        project_root=pyside_root,
        source_base=inputs / "b1" / "cached-base.xlsx",
        source_collect=inputs / "b1" / "cached-collect.xlsx",
        output_dir=output_dir,
    )
    projects: dict[str, Any] = {}
    for project, root, executable in specs:
        projects[project] = {
            "B1": _run_case(
                project=project,
                executable=executable,
                project_root=root,
                case="B1",
                source_base=inputs / "b1" / "cached-base.xlsx",
                source_collect=inputs / "b1" / "cached-collect.xlsx",
                output_dir=output_dir,
            ),
            "N1": _run_case(
                project=project,
                executable=executable,
                project_root=root,
                case="N1",
                source_base=inputs / "n1" / "protected-base.xlsx",
                source_collect=inputs / "n1" / "legacy-collect.xls",
                output_dir=output_dir,
            ),
        }

    comparisons: dict[str, Any] = {}
    for case in ("B1", "N1"):
        projections = {project: projects[project][case]["projection"] for project in PROJECTS}
        baseline = projections["docwen-ref-tk"]
        mandatory = {project: _mandatory_projection(projection) for project, projection in projections.items()}
        comparisons[case] = {
            "all_equal": all(projection == baseline for projection in projections.values()),
            "matches_expected": {
                project: projection == EXPECTED_MANDATORY_PROJECTIONS[case] for project, projection in mandatory.items()
            },
            "expected_mandatory_projection": EXPECTED_MANDATORY_PROJECTIONS[case],
            "projections": projections,
        }
    prepared = manifest["prepared_inputs"]
    cache_requirements_met = (
        prepared["inputs/b1/cached-base.xlsx"]["cached_values"] == {"Formula!B4": 15}
        and prepared["inputs/b1/cached-collect.xlsx"]["cached_values"] == {"Formula!B4": 5}
        and prepared["inputs/n1/protected-base.xlsx"]["formula_count"] == 6
        and all(value is not None for value in prepared["inputs/n1/protected-base.xlsx"]["cached_values"].values())
    )
    final_placement_met = all(
        projects[project][case]["final_placement"]["inside_case_dir"]
        and projects[project][case]["final_placement"]["suffix"] == ".xlsx"
        for project in PROJECTS
        for case in ("B1", "N1")
    )
    result = {
        "probe_id": manifest["probe_id"],
        "input_manifest": str(output_dir / "input-manifest.json"),
        "projects": projects,
        "old_reference_dispositions": {"docwen-ref-pyside6": pyside_cli_defect},
        "comparisons": comparisons,
        "acceptance": {
            "cache_requirements_met": cache_requirements_met,
            "final_placement_met": final_placement_met,
            "reference_defect_observed": pyside_cli_defect["expected_failure_observed"],
        },
        "pass": cache_requirements_met
        and final_placement_met
        and pyside_cli_defect["expected_failure_observed"]
        and all(value["all_equal"] for value in comparisons.values())
        and all(all(value["matches_expected"].values()) for value in comparisons.values())
        and all(
            not projects[project][case]["processes"]["residue_added"] for project in PROJECTS for case in ("B1", "N1")
        ),
    }
    _write_json(output_dir / "probe-result.json", result)
    _print_json(result)
    return 0 if result["pass"] else 1


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pyside-worker", action="store_true")
    parser.add_argument("--worker-base")
    parser.add_argument("--worker-collect")
    parser.add_argument("--worker-result")
    parser.add_argument("--current-root", default=".")
    parser.add_argument("--tk-root", default="../docwen-ref-tk")
    parser.add_argument("--pyside-root", default="../docwen-ref-pyside6")
    parser.add_argument("--rich-xlsx")
    parser.add_argument("--legacy-xls")
    parser.add_argument("--output-dir")
    parser.add_argument("--prepare-only", action="store_true")
    return parser


def main() -> int:
    args = _parser().parse_args()
    return _pyside_worker(args) if args.pyside_worker else _orchestrate(args)


if __name__ == "__main__":
    raise SystemExit(main())
